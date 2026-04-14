# Author: Sa-tish Chauhan

"""
Schema Validation Tab
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from pathlib import Path
import threading
import sys
import pyodbc
import json
import time
from datetime import datetime

# Add parent directories to path
parent_dir = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(parent_dir))

from gui.utils.excel_utils import read_excel_file, create_sample_excel
from gui.utils.database_utils import connect_with_msal_cache, connect_to_any_database
from gui.utils.schema_remap import pair_tables_for_schema_remap
from gui.utils.canvas_mousewheel import bind_canvas_vertical_scroll

# Import DB2-agnostic schema helpers
try:
    from gui.utils.db2_schema import (
        is_db2_connection, fetch_tables_generic, fetch_columns_generic, 
        table_exists_generic, get_tables_query, get_columns_query
    )
    _DB2_HELPERS_AVAILABLE = True
except ImportError:
    _DB2_HELPERS_AVAILABLE = False
    is_db2_connection = None
    fetch_tables_generic = None
    fetch_columns_generic = None
    table_exists_generic = None

# Import DB2 to SQL type mapping validation
try:
    from gui.utils.db2_type_mapping import (
        validate_type_mapping, compare_columns_with_type_mapping,
        get_expected_sql_type, get_type_mapping_summary
    )
    _TYPE_MAPPING_AVAILABLE = True
except ImportError:
    _TYPE_MAPPING_AVAILABLE = False
    validate_type_mapping = None
    compare_columns_with_type_mapping = None

# Retry helper for deadlock errors
def retry_on_deadlock(func, max_retries=3, delay=1.0, log_callback=None):
    """
    Retry a database operation on deadlock errors (SQL Server error 1205 / 40001).
    
    Args:
        func: Function to execute (should be a lambda or callable)
        max_retries: Maximum number of retry attempts (default: 3)
        delay: Initial delay in seconds between retries (default: 1.0, doubles each retry)
        log_callback: Optional callback function(str) to log retry attempts
    
    Returns:
        Result of func()
    
    Raises:
        Last exception if all retries fail
    """
    last_exception = None
    current_delay = delay
    
    for attempt in range(max_retries):
        try:
            return func()
        except pyodbc.Error as e:
            error_code = e.args[0] if e.args else ""
            error_msg = str(e.args[1]) if len(e.args) > 1 else str(e)
            
            # Check if it's a deadlock error (1205 or 40001)
            is_deadlock = (
                "1205" in error_msg or 
                "40001" in error_code or 
                "deadlock" in error_msg.lower() or
                "deadlocked" in error_msg.lower()
            )
            
            if is_deadlock and attempt < max_retries - 1:
                last_exception = e
                if log_callback:
                    log_callback(f"Deadlock detected (attempt {attempt + 1}/{max_retries}), retrying in {current_delay:.1f}s...")
                time.sleep(current_delay)
                current_delay *= 2  # Exponential backoff
                continue
            else:
                # Not a deadlock or out of retries
                raise
        except Exception as e:
            # Non-pyodbc exceptions or non-deadlock errors - don't retry
            raise
    
    # If we get here, all retries failed
    if last_exception:
        raise last_exception
    raise Exception("Retry failed for unknown reason")


# Import signature-based matching for indexes and constraints
try:
    from gui.utils.schema_matching import (
        compare_indexes_with_signatures,
        compare_foreign_keys_with_signatures,
        compare_default_constraints_with_signatures,
        is_auto_generated_constraint_name,
        build_index_signature,
        build_fk_signature
    )
    _SIGNATURE_MATCHING_AVAILABLE = True
except ImportError:
    _SIGNATURE_MATCHING_AVAILABLE = False
    compare_indexes_with_signatures = None
    compare_foreign_keys_with_signatures = None
    compare_default_constraints_with_signatures = None
    is_auto_generated_constraint_name = None

# Import driver utilities for handling missing driver errors
try:
    from utils.driver_utils import (
        check_sql_server_odbc_driver,
        install_odbc_via_powershell,
        get_manual_install_instructions,
    )
    DRIVER_UTILS_AVAILABLE = True
except ImportError:
    DRIVER_UTILS_AVAILABLE = False

# First segment of "Type.Schema.Name" programmable objects (not "Schema.Table.Column")
_SCHEMA_VALIDATION_OBJECT_TYPE_PREFIXES = frozenset(
    {
        "VIEW",
        "SQL_STORED_PROCEDURE",
        "SQL_STOR",
        "SQL_SCALAR_FUNCTION",
        "SQL_TABLE_VALUED_FUNCTION",
        "SQL_INLINE_TABLE_VALUED_FUNCTION",
        "SQL_FUNCTION",
        "CLR_STORED_PROCEDURE",
        "CLR_SCALAR_FUNCTION",
        "CLR_TABLE_VALUED_FUNCTION",
        "CLR_AGGREGATE_FUNCTION",
        "SYNONYM",
        "SEQUENCE",
    }
)


def is_driver_missing_error(error_msg: str) -> bool:
    """Check if an error is related to missing ODBC driver."""
    error_str = str(error_msg).upper()
    driver_error_indicators = [
        "IM002",  # Data source name not found
        "IM003",  # Driver not found
        "IM004",  # Driver's SQLAllocHandle failed
        "01000",  # Driver not capable
        "DATA SOURCE NAME NOT FOUND",
        "NO DEFAULT DRIVER",
        "DRIVER NOT FOUND",
        "ODBC DRIVER",
    ]
    return any(indicator in error_str for indicator in driver_error_indicators)


class BulkSchemaComparisonDialog:
    """Dialog for comparing and fixing multiple missing objects at once."""
    
    def __init__(self, parent, missing_objects_info, fetch_code_callback, deploy_callback, get_drop_callback=None, get_table_script_callback=None):
        """
        Initialize bulk comparison dialog.

        Args:
            parent: Parent window
            missing_objects_info: List of tuples (obj_name, db_name, status) with status "Missing" or "Mismatch"
            fetch_code_callback: Function to fetch source/dest code: (obj_name, is_source) -> code
            deploy_callback: Function to deploy script: (obj_name, sql_script) -> (success, message)
            get_drop_callback: Optional (obj_name, db_name) -> DROP SQL string; used for Mismatch to prepend DROP before source DDL
            get_table_script_callback: Optional (db_name, schema, table) -> (script, error). Used to replicate full table index/PK layout when index/PK Mismatch.
        """
        self.parent = parent
        self.missing_objects_info = missing_objects_info
        self.fetch_code_callback = fetch_code_callback
        self.deploy_callback = deploy_callback
        self.get_drop_callback = get_drop_callback
        self.get_table_script_callback = get_table_script_callback
        self.object_data = {}  # Store code for each object
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Bulk Schema Comparison - {len(missing_objects_info)} Objects")
        self.dialog.geometry("1400x800")
        self.dialog.transient(parent)
        
        self._create_widgets()
        self._load_all_objects()
        
    def _create_widgets(self):
        """Create dialog widgets."""
        # Header
        header_frame = ttk.Frame(self.dialog, padding=10)
        header_frame.pack(fill=tk.X)
        
        tk.Label(header_frame, text=f"Comparing {len(self.missing_objects_info)} Missing Object(s)", 
                font=("Arial", 12, "bold")).pack(side=tk.LEFT)
        
        # Main content - split into list and comparison view
        main_frame = ttk.Frame(self.dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel - Object list
        left_panel = ttk.Frame(main_frame, width=300)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5))
        left_panel.pack_propagate(False)
        
        tk.Label(left_panel, text="Objects to fix (Missing / Mismatch)", font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))
        
        # Listbox with scrollbar
        list_frame = ttk.Frame(left_panel)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        list_scroll = ttk.Scrollbar(list_frame)
        list_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.objects_listbox = tk.Listbox(list_frame, yscrollcommand=list_scroll.set, font=("Consolas", 9))
        self.objects_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scroll.config(command=self.objects_listbox.yview)
        
        self.objects_listbox.bind('<<ListboxSelect>>', self._on_object_select)
        
        # Populate listbox; each item is (obj_name, db_name, status)
        for obj_name, db_name, status in self.missing_objects_info:
            display_name = f"{obj_name}\n  [{db_name}] ({status})"
            self.objects_listbox.insert(tk.END, display_name)
        
        # Right panel - Comparison view
        right_panel = ttk.Frame(main_frame)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Comparison labels
        labels_frame = ttk.Frame(right_panel)
        labels_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(labels_frame, text="Source Code", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Label(labels_frame, text="Destination Code (Editable)", font=("Arial", 9, "bold")).pack(side=tk.RIGHT, padx=5)
        
        # Text widgets for comparison
        text_frame = ttk.Frame(right_panel)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Source text (read-only)
        source_frame = ttk.Frame(text_frame)
        source_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        source_scroll = ttk.Scrollbar(source_frame)
        source_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.source_text = tk.Text(source_frame, wrap=tk.NONE, yscrollcommand=source_scroll.set,
                                   font=("Consolas", 9), bg="#f5f5f5", fg="black", state=tk.DISABLED)
        self.source_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        source_scroll.config(command=self.source_text.yview)
        
        # Destination text (editable)
        dest_frame = ttk.Frame(text_frame)
        dest_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        dest_scroll = ttk.Scrollbar(dest_frame)
        dest_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.dest_text = tk.Text(dest_frame, wrap=tk.NONE, yscrollcommand=dest_scroll.set,
                                 font=("Consolas", 9), bg="white", fg="black", insertbackground="black")
        self.dest_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        dest_scroll.config(command=self.dest_text.yview)
        
        # Sync scrolling and detect edits
        def dest_key_handler(e):
            self._sync_scroll(e)
            self._on_dest_edit()
        
        def dest_click_handler(e):
            self._sync_scroll(e)
            self._on_dest_edit()
        
        self.source_text.bind('<KeyRelease>', self._sync_scroll)
        self.source_text.bind('<Button-1>', self._sync_scroll)
        self.source_text.bind('<MouseWheel>', self._sync_scroll)
        self.dest_text.bind('<KeyRelease>', dest_key_handler)
        self.dest_text.bind('<Button-1>', dest_click_handler)
        self.dest_text.bind('<MouseWheel>', self._sync_scroll)
        
        # Buttons frame
        btn_frame = ttk.Frame(self.dialog, padding=10)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="Refresh All", command=self._refresh_all, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Deploy All", command=self._deploy_all, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Deploy Selected", command=self._deploy_selected, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Close", command=self.dialog.destroy, width=15).pack(side=tk.RIGHT, padx=5)
        
        # Status label
        self.status_label = tk.Label(btn_frame, text="Ready", fg="green")
        self.status_label.pack(side=tk.LEFT, padx=20)
        
        # Select first object
        if self.objects_listbox.size() > 0:
            self.objects_listbox.selection_set(0)
            self.objects_listbox.event_generate("<<ListboxSelect>>")
    
    def _sync_scroll(self, event=None):
        """Sync scrolling between source and destination text widgets."""
        try:
            # Get scroll position from source
            source_pos = self.source_text.yview()[0]
            # Apply to destination
            self.dest_text.yview_moveto(source_pos)
        except:
            pass
    
    def _on_object_select(self, event=None):
        """Handle object selection from listbox."""
        selection = self.objects_listbox.curselection()
        if not selection:
            return
        
        idx = selection[0]
        obj_name, db_name, source_status = self.missing_objects_info[idx]
        
        # Save current edits if any
        if hasattr(self, 'current_obj_name') and self.current_obj_name:
            self._save_current_edits()
        
        # Load object data
        self.current_obj_name = obj_name
        self._load_object_data(obj_name)
    
    def _save_current_edits(self):
        """Save current edits to object_data."""
        if hasattr(self, 'current_obj_name') and self.current_obj_name:
            dest_code = self.dest_text.get("1.0", tk.END).rstrip()
            if self.current_obj_name in self.object_data:
                self.object_data[self.current_obj_name]['dest_code'] = dest_code
                self.object_data[self.current_obj_name]['edited'] = True
    
    def _load_object_data(self, obj_name):
        """Load source and destination code for an object. For Mismatch, destination panel shows full fix script (DROP + source)."""
        # Check if already loaded
        if obj_name in self.object_data:
            data = self.object_data[obj_name]
            self._display_object_data(data['source_code'], data['dest_code'])
            return
        
        # Resolve db_name and status for this object (for Mismatch we show DROP + source as dest)
        db_name = None
        status = None
        for o, d, s in self.missing_objects_info:
            if o == obj_name:
                db_name = d
                status = s
                break

        # Fetch from callback
        self.status_label.config(text=f"Loading {obj_name}...", fg="blue")
        self.dialog.update()
        
        try:
            # Fetch source code
            source_code = self.fetch_code_callback(obj_name, True) or ""
            
            # Fetch destination code (may fail if object doesn't exist, which is expected)
            dest_code = ""
            try:
                dest_code = self.fetch_code_callback(obj_name, False) or ""
            except Exception:
                dest_code = ""
            
            # If destination code is empty/None and source code exists, use source as template
            if not dest_code or dest_code.strip() == "":
                dest_code = source_code
            else:
                # When object is missing, destination may return a different object type (e.g. index name
                # matches a column name). Use source script so we deploy the correct object.
                src_is_index = source_code.strip().startswith("-- Index:") or ("CREATE " in source_code and " INDEX " in source_code)
                dest_is_column = dest_code.strip().startswith("-- Column:") or ("ALTER TABLE" in dest_code and " ADD " in dest_code)
                src_is_column = source_code.strip().startswith("-- Column:") or ("ALTER TABLE" in source_code and " ADD " in source_code)
                dest_is_index = dest_code.strip().startswith("-- Index:") or ("CREATE " in dest_code and " INDEX " in dest_code)
                if (src_is_index and dest_is_column) or (src_is_column and dest_is_index):
                    dest_code = source_code

            # For Mismatch: show full fix script (DROP then CREATE from source) in destination panel so user sees what will run
            if status == "Mismatch" and db_name and getattr(self, "get_drop_callback", None):
                drop_sql = (self.get_drop_callback(obj_name, db_name) or "").strip()
                if drop_sql and source_code and source_code.strip():
                    dest_code = drop_sql + "\n\n" + source_code.strip()
            
            # Store in object_data
            self.object_data[obj_name] = {
                'source_code': source_code,
                'dest_code': dest_code,
                'edited': False
            }
            
            self._display_object_data(source_code, dest_code)
            self.status_label.config(text="Ready", fg="green")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}", fg="red")
            source_code = self.object_data.get(obj_name, {}).get('source_code', f"Error loading: {str(e)}")
            self._display_object_data(source_code, "")
    
    def _display_object_data(self, source_code, dest_code):
        """Display source and destination code in text widgets."""
        # Enable source text for editing
        self.source_text.config(state=tk.NORMAL)
        self.source_text.delete("1.0", tk.END)
        if source_code and source_code.strip():
            self.source_text.insert("1.0", source_code)
        else:
            self.source_text.insert("1.0", "(Object not found in source)")
        self.source_text.config(state=tk.DISABLED)
        
        # Enable destination text for editing
        self.dest_text.config(state=tk.NORMAL)
        self.dest_text.delete("1.0", tk.END)
        if dest_code and dest_code.strip():
            self.dest_text.insert("1.0", dest_code)
        else:
            # If no destination code, use source code as template (object doesn't exist in destination)
            if source_code and source_code.strip():
                self.dest_text.insert("1.0", source_code)
            else:
                self.dest_text.insert("1.0", "(Object does not exist in destination - edit script above to create it)")
        self.dest_text.config(state=tk.NORMAL)
        
        # Highlight differences
        self._highlight_differences()
    
    def _on_dest_edit(self):
        """Handle destination text edits - update highlighting and mark as edited."""
        # Re-highlight differences after edit
        self._highlight_differences()
        
        # Mark as edited
        if hasattr(self, 'current_obj_name') and self.current_obj_name:
            if self.current_obj_name in self.object_data:
                self.object_data[self.current_obj_name]['edited'] = True
    
    def _highlight_differences(self):
        """Highlight differences between source and destination."""
        source_lines = self.source_text.get("1.0", tk.END).split('\n')
        dest_lines = self.dest_text.get("1.0", tk.END).split('\n')
        
        max_lines = max(len(source_lines), len(dest_lines))
        
        # Clear existing tags
        self.source_text.tag_delete("diff")
        self.dest_text.tag_delete("diff")
        
        for i in range(max_lines):
            source_line = source_lines[i] if i < len(source_lines) else ""
            dest_line = dest_lines[i] if i < len(dest_lines) else ""
            
            if source_line != dest_line:
                start_pos = f"{i+1}.0"
                end_pos = f"{i+1}.end"
                self.source_text.tag_add("diff", start_pos, end_pos)
                self.dest_text.tag_add("diff", start_pos, end_pos)
        
        # Configure diff tag
        self.source_text.tag_config("diff", background="#ffcccc", foreground="black")
        self.dest_text.tag_config("diff", background="#ffcccc", foreground="black")
    
    def _load_all_objects(self):
        """Load all objects in background. For Mismatch, dest_code = DROP + source (full fix script)."""
        def load_all():
            for obj_name, db_name, status in self.missing_objects_info:
                if obj_name not in self.object_data:
                    try:
                        source_code = self.fetch_code_callback(obj_name, True) or ""
                        dest_code = ""
                        try:
                            dest_code = self.fetch_code_callback(obj_name, False) or ""
                        except Exception:
                            dest_code = ""
                        if not dest_code or not dest_code.strip():
                            dest_code = source_code
                        else:
                            src_is_index = source_code.strip().startswith("-- Index:") or ("CREATE " in source_code and " INDEX " in source_code)
                            dest_is_column = dest_code.strip().startswith("-- Column:") or ("ALTER TABLE" in dest_code and " ADD " in dest_code)
                            src_is_column = source_code.strip().startswith("-- Column:") or ("ALTER TABLE" in source_code and " ADD " in source_code)
                            dest_is_index = dest_code.strip().startswith("-- Index:") or ("CREATE " in dest_code and " INDEX " in dest_code)
                            if (src_is_index and dest_is_column) or (src_is_column and dest_is_index):
                                dest_code = source_code
                        if status == "Mismatch" and db_name and getattr(self, "get_drop_callback", None):
                            drop_sql = (self.get_drop_callback(obj_name, db_name) or "").strip()
                            if drop_sql and source_code and source_code.strip():
                                dest_code = drop_sql + "\n\n" + source_code.strip()
                        self.object_data[obj_name] = {
                            'source_code': source_code,
                            'dest_code': dest_code,
                            'edited': False
                        }
                    except Exception as e:
                        self.object_data[obj_name] = {
                            'source_code': f"-- Error: {str(e)}",
                            'dest_code': "",
                            'edited': False
                        }
        threading.Thread(target=load_all, daemon=True).start()
    
    def _refresh_all(self):
        """Refresh all object data."""
        self.object_data.clear()
        self._load_all_objects()
        if hasattr(self, 'current_obj_name') and self.current_obj_name:
            self._load_object_data(self.current_obj_name)
        self.status_label.config(text="Refreshed", fg="green")
    
    def _deploy_all(self):
        """Deploy all objects."""
        self._save_current_edits()
        
        if not messagebox.askyesno("Confirm Deployment", 
                                   f"Deploy all {len(self.missing_objects_info)} objects to destination?\n\n"
                                   "This will execute the SQL scripts for all objects."):
            return
        
        self._deploy_objects(self.missing_objects_info)
    
    def _deploy_selected(self):
        """Deploy selected objects."""
        selection = self.objects_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select at least one object to deploy.")
            return
        
        self._save_current_edits()
        
        selected_objects = [self.missing_objects_info[idx] for idx in selection]
        
        if not messagebox.askyesno("Confirm Deployment", 
                                   f"Deploy {len(selected_objects)} selected object(s) to destination?"):
            return
        
        self._deploy_objects(selected_objects)
    
    def _deploy_objects(self, objects_to_deploy):
        """Deploy objects to destination. For Mismatch, prepend DROP then use source DDL.
        Deploy primary keys before other indexes so CLUSTERED index can be created (exact replication).
        """
        self.status_label.config(text="Deploying...", fg="blue")
        self.dialog.update()

        # Sort so primary keys (PK_*) are deployed first; then CLUSTERED indexes can succeed
        def _deploy_order(item):
            obj_name = item[0]
            parts = obj_name.split(".")
            last_part = parts[-1] if len(parts) >= 3 else ""
            is_pk = last_part.startswith("PK_")
            return (0 if is_pk else 1, obj_name)
        objects_to_deploy = sorted(objects_to_deploy, key=_deploy_order)

        success_count = 0
        fail_count = 0
        errors = []
        # Tables for which we will deploy one combined "table index layout" script (drop FKs, drop indexes/PK, recreate from source, restore FKs)
        tables_use_table_script = set()
        for obj_name, db_name, status in objects_to_deploy:
            p = obj_name.split(".")
            if len(p) == 3 and status in ("Mismatch", "Missing"):
                third = p[2]
                if third.startswith("PK_") or third.startswith("IX_") or third.startswith("IDX_") or third.startswith("AK_"):
                    tables_use_table_script.add((db_name, p[0], p[1]))
        deployed_tables = set()  # (db_name, schema, table) already deployed with table script

        for obj_name, db_name, status in objects_to_deploy:
            try:
                parts = obj_name.split(".")
                if len(parts) == 3:
                    schema, table, third = parts[0], parts[1], parts[2]
                    table_key = (db_name, schema, table)
                    is_index_or_pk = third.startswith("PK_") or third.startswith("IX_") or third.startswith("IDX_") or third.startswith("AK_")
                    if is_index_or_pk and table_key in tables_use_table_script and getattr(self, "get_table_script_callback", None):
                        if table_key in deployed_tables:
                            success_count += 1
                            continue
                        script, err = self.get_table_script_callback(db_name, schema, table)
                        if err:
                            errors.append(f"{schema}.{table} (table index layout): {err}")
                            fail_count += 1
                            continue
                        if not script or not script.strip():
                            errors.append(f"{schema}.{table}: No table script generated")
                            fail_count += 1
                            continue
                        success, message = self.deploy_callback(obj_name, script)
                        if success:
                            deployed_tables.add(table_key)
                            success_count += 1
                        else:
                            fail_count += 1
                            errors.append(f"{schema}.{table} (table index layout): {message}")
                        continue
                    if table_key in deployed_tables:
                        success_count += 1
                        continue

                # Per-object deploy
                if obj_name in self.object_data:
                    script = self.object_data[obj_name].get('dest_code', '')
                    if script and "(Object does not exist in destination" in script:
                        script = self.object_data[obj_name].get('source_code', '')
                    if not script or not script.strip():
                        script = self.object_data[obj_name].get('source_code', '')
                else:
                    script = self.fetch_code_callback(obj_name, True) or ""

                if not script or not script.strip():
                    errors.append(f"{obj_name}: No script available")
                    fail_count += 1
                    continue

                # For Mismatch (definition differs): prepend DROP on destination then recreate from source
                if status == "Mismatch" and getattr(self, 'get_drop_callback', None):
                    drop_sql = self.get_drop_callback(obj_name, db_name)
                    if drop_sql and drop_sql.strip():
                        script = drop_sql.strip() + "\n\n" + script.strip()

                success, message = self.deploy_callback(obj_name, script)
                if success:
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"{obj_name}: {message}")

            except Exception as e:
                fail_count += 1
                errors.append(f"{obj_name}: {str(e)}")
        
        # Show results
        result_msg = f"Deployment Complete:\n\nSuccess: {success_count}\nFailed: {fail_count}"
        if errors:
            result_msg += f"\n\nErrors:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                result_msg += f"\n... and {len(errors) - 10} more"
        
        if fail_count == 0:
            messagebox.showinfo("Success", result_msg)
            self.status_label.config(text=f"Deployed {success_count} objects", fg="green")
        else:
            messagebox.showwarning("Deployment Complete", result_msg)
            self.status_label.config(text=f"{success_count} succeeded, {fail_count} failed", fg="orange")


class SchemaComparisonDialog:
    """Dialog for comparing source and destination schema objects with diff highlighting."""
    
    def __init__(self, parent, obj_name, source_code, dest_code, deploy_callback):
        """
        Initialize comparison dialog.
        
        Args:
            parent: Parent window
            obj_name: Name of the object being compared
            source_code: Source code (left side)
            dest_code: Destination code (right side, may be empty)
            deploy_callback: Callback function to execute when deploy is clicked
                            Should accept (obj_name, sql_script) and return (success, message)
        """
        self.parent = parent
        self.obj_name = obj_name
        self.source_code = source_code or ""
        self.dest_code = dest_code or ""
        self.deploy_callback = deploy_callback
        self.sql_script = None
        self.current_step = "compare"  # "compare" or "preview"
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Schema Comparison: {obj_name}")
        self.dialog.geometry("1200x700")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self._create_widgets()
        self._highlight_differences()
        
    def _create_widgets(self):
        """Create dialog widgets."""
        # Header frame
        header_frame = ttk.Frame(self.dialog, padding=10)
        header_frame.pack(fill=tk.X)
        
        tk.Label(header_frame, text=f"Object: {self.obj_name}", 
                font=("Arial", 12, "bold")).pack(side=tk.LEFT)
        
        # Main content frame
        content_frame = ttk.Frame(self.dialog)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Comparison view (shown initially)
        self.compare_frame = ttk.Frame(content_frame)
        self.compare_frame.pack(fill=tk.BOTH, expand=True)
        
        # Labels for source and destination
        labels_frame = ttk.Frame(self.compare_frame)
        labels_frame.pack(fill=tk.X)
        
        tk.Label(labels_frame, text="Source (Left)", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Label(labels_frame, text="Destination (Right)", font=("Arial", 10, "bold")).pack(side=tk.RIGHT, padx=5)
        
        # Text widgets for comparison
        text_frame = ttk.Frame(self.compare_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Source text widget (left)
        source_frame = ttk.Frame(text_frame)
        source_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        tk.Label(source_frame, text="Source Code", font=("Arial", 9)).pack(anchor=tk.W)
        source_scroll = ttk.Scrollbar(source_frame)
        source_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.source_text = tk.Text(source_frame, wrap=tk.NONE, yscrollcommand=source_scroll.set,
                                   font=("Consolas", 10), bg="white", fg="black")
        self.source_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        source_scroll.config(command=self.source_text.yview)
        
        # Sync scrolling
        self.source_text.bind('<KeyRelease>', self._sync_scroll)
        self.source_text.bind('<Button-1>', self._sync_scroll)
        self.source_text.bind('<MouseWheel>', self._sync_scroll)
        
        # Destination text widget (right)
        dest_frame = ttk.Frame(text_frame)
        dest_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        tk.Label(dest_frame, text="Destination Code", font=("Arial", 9)).pack(anchor=tk.W)
        dest_scroll = ttk.Scrollbar(dest_frame)
        dest_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.dest_text = tk.Text(dest_frame, wrap=tk.NONE, yscrollcommand=dest_scroll.set,
                                 font=("Consolas", 10), bg="white", fg="black")
        self.dest_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        dest_scroll.config(command=self.dest_text.yview)
        
        # Sync scrolling
        self.dest_text.bind('<KeyRelease>', self._sync_scroll)
        self.dest_text.bind('<Button-1>', self._sync_scroll)
        self.dest_text.bind('<MouseWheel>', self._sync_scroll)
        
        # Preview view (shown after clicking Next)
        self.preview_frame = ttk.Frame(content_frame)
        
        tk.Label(self.preview_frame, text="SQL Script to Execute on Destination", 
                font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))
        
        preview_scroll = ttk.Scrollbar(self.preview_frame)
        preview_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.preview_text = scrolledtext.ScrolledText(self.preview_frame, wrap=tk.WORD,
                                                      font=("Consolas", 10), height=30)
        self.preview_text.pack(fill=tk.BOTH, expand=True)
        
        # Buttons frame
        btn_frame = ttk.Frame(self.dialog, padding=10)
        btn_frame.pack(fill=tk.X)
        
        self.next_btn = ttk.Button(btn_frame, text="Next: Preview Script", 
                                  command=self._show_preview, width=20)
        self.next_btn.pack(side=tk.LEFT, padx=5)
        
        self.back_btn = ttk.Button(btn_frame, text="Back: Comparison", 
                                  command=self._show_comparison, width=20, state=tk.DISABLED)
        self.back_btn.pack(side=tk.LEFT, padx=5)
        
        self.deploy_btn = ttk.Button(btn_frame, text="Deploy", 
                                    command=self._deploy, width=15, state=tk.DISABLED)
        self.deploy_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="Cancel", command=self.dialog.destroy, width=15).pack(side=tk.RIGHT, padx=5)
        
        # Populate text widgets
        self.source_text.insert("1.0", self.source_code)
        self.dest_text.insert("1.0", self.dest_code if self.dest_code else "(Object does not exist in destination)")
        
        # Make text widgets read-only
        self.source_text.config(state=tk.DISABLED)
        self.dest_text.config(state=tk.DISABLED)
        
    def _sync_scroll(self, event=None):
        """Sync scrolling between source and destination text widgets."""
        try:
            # Get scroll position from source
            source_pos = self.source_text.yview()[0]
            # Apply to destination
            self.dest_text.yview_moveto(source_pos)
        except:
            pass
    
    def _highlight_differences(self):
        """Highlight differences between source and destination code."""
        # Enable text widgets for editing
        self.source_text.config(state=tk.NORMAL)
        self.dest_text.config(state=tk.NORMAL)
        
        source_lines = self.source_code.split('\n')
        dest_lines = (self.dest_code if self.dest_code else "").split('\n')
        
        # Simple line-by-line comparison
        max_lines = max(len(source_lines), len(dest_lines))
        
        for i in range(max_lines):
            source_line = source_lines[i] if i < len(source_lines) else ""
            dest_line = dest_lines[i] if i < len(dest_lines) else ""
            
            # If lines are different, highlight in red
            if source_line != dest_line:
                # Highlight source line
                start_pos = f"{i+1}.0"
                end_pos = f"{i+1}.end"
                self.source_text.tag_add("diff", start_pos, end_pos)
                self.dest_text.tag_add("diff", start_pos, end_pos)
        
        # Configure diff tag
        self.source_text.tag_config("diff", background="#ffcccc", foreground="black")
        self.dest_text.tag_config("diff", background="#ffcccc", foreground="black")
        
        # Make text widgets read-only again
        self.source_text.config(state=tk.DISABLED)
        self.dest_text.config(state=tk.DISABLED)
    
    def _show_preview(self):
        """Show SQL script preview."""
        # Generate SQL script from source code
        # Use source code as the script (it should already be in CREATE/ALTER format)
        self.sql_script = self.source_code
        
        # If source code is empty or invalid, show error
        if not self.sql_script or not self.sql_script.strip():
            messagebox.showerror("Error", "No SQL script available to preview!")
            return
        
        # Show preview frame, hide comparison frame
        self.compare_frame.pack_forget()
        self.preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Populate preview
        self.preview_text.delete("1.0", tk.END)
        self.preview_text.insert("1.0", self.sql_script)
        self.preview_text.config(state=tk.DISABLED)
        
        # Update buttons
        self.next_btn.config(state=tk.DISABLED)
        self.back_btn.config(state=tk.NORMAL)
        self.deploy_btn.config(state=tk.NORMAL)
        self.current_step = "preview"
    
    def _show_comparison(self):
        """Show comparison view."""
        # Show comparison frame, hide preview frame
        self.preview_frame.pack_forget()
        self.compare_frame.pack(fill=tk.BOTH, expand=True)
        
        # Update buttons
        self.next_btn.config(state=tk.NORMAL)
        self.back_btn.config(state=tk.DISABLED)
        self.deploy_btn.config(state=tk.DISABLED)
        self.current_step = "compare"
    
    def _deploy(self):
        """Deploy the script to destination."""
        if not self.sql_script:
            messagebox.showerror("Error", "No SQL script to deploy!")
            return
        
        # Confirm deployment
        if not messagebox.askyesno("Confirm Deployment", 
                                   f"Are you sure you want to deploy this script to destination?\n\n"
                                   f"Object: {self.obj_name}"):
            return
        
        # Disable deploy button
        self.deploy_btn.config(state=tk.DISABLED)
        
        # Call deploy callback
        try:
            success, message = self.deploy_callback(self.obj_name, self.sql_script)
            if success:
                messagebox.showinfo("Success", f"Deployment successful!\n\n{message}")
                self.dialog.destroy()
            else:
                messagebox.showerror("Deployment Failed", f"Deployment failed:\n\n{message}")
                self.deploy_btn.config(state=tk.NORMAL)
        except Exception as e:
            messagebox.showerror("Error", f"Error during deployment:\n\n{str(e)}")
            self.deploy_btn.config(state=tk.NORMAL)


class SchemaValidationTab:
    """Schema validation tab."""
    
    def __init__(self, parent, main_window):
        self.main_window = main_window
        self.frame = ttk.Frame(parent)
        self.project_path = None
        
        self._create_widgets()
        
    def set_project_path(self, project_path):
        """Set the current project path."""
        self.project_path = project_path
        
    def _create_widgets(self):
        """Create UI widgets."""
        # Create scrollable canvas
        canvas = tk.Canvas(self.frame)
        scrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        def on_canvas_configure(event):
            canvas_width = event.width
            canvas.itemconfig(canvas_frame, width=canvas_width)
        
        canvas.bind('<Configure>', on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Title
        title_label = tk.Label(
            scrollable_frame,
            text="Schema Validation",
            font=("Arial", 16, "bold")
        )
        title_label.pack(pady=10)
        
        # Create two-column layout
        main_frame = ttk.Frame(scrollable_frame)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        bind_canvas_vertical_scroll(canvas, scrollable_frame)

        # Store reference
        self.scrollable_frame = scrollable_frame
        
        # Left column - Source
        left_frame = ttk.LabelFrame(main_frame, text="Source Database", padding=10)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        # Use shared variables from main window
        self.src_server_var = self.main_window.shared_src_server
        self.src_db_var = self.main_window.shared_src_db
        self.src_auth_var = self.main_window.shared_src_auth
        self.src_user_var = self.main_window.shared_src_user
        self.src_password_var = self.main_window.shared_src_password
        
        # DB type, port, and schema for DB2 support
        self.src_db_type_var = tk.StringVar(value="sqlserver")
        self.src_port_var = tk.StringVar(value="50000")
        self.src_schema_var = tk.StringVar(value="")
        
        # Create connection widget for source
        from gui.widgets.connection_widget import ConnectionWidget
        self.src_connection_widget = ConnectionWidget(
            parent=left_frame,
            server_var=self.src_server_var,
            db_var=self.src_db_var,
            auth_var=self.src_auth_var,
            user_var=self.src_user_var,
            password_var=self.src_password_var,
            label_text="",
            row_start=0,
            db_type_var=self.src_db_type_var,
            port_var=self.src_port_var,
            schema_var=self.src_schema_var
        )
        
        # Right column - Destination (using shared variables)
        right_frame = ttk.LabelFrame(main_frame, text="Destination Database", padding=10)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        # Use shared variables from main window
        self.dest_server_var = self.main_window.shared_dest_server
        self.dest_db_var = self.main_window.shared_dest_db
        self.dest_auth_var = self.main_window.shared_dest_auth
        self.dest_user_var = self.main_window.shared_dest_user
        self.dest_password_var = self.main_window.shared_dest_password
        
        # DB type, port, and schema for DB2 support
        self.dest_db_type_var = tk.StringVar(value="sqlserver")
        self.dest_port_var = tk.StringVar(value="50000")
        self.dest_schema_var = tk.StringVar(value="")
        
        # Create connection widget for destination
        from gui.widgets.connection_widget import ConnectionWidget
        self.dest_connection_widget = ConnectionWidget(
            parent=right_frame,
            server_var=self.dest_server_var,
            db_var=self.dest_db_var,
            auth_var=self.dest_auth_var,
            user_var=self.dest_user_var,
            password_var=self.dest_password_var,
            label_text="",
            row_start=0,
            db_type_var=self.dest_db_type_var,
            port_var=self.dest_port_var,
            schema_var=self.dest_schema_var
        )
        
        # Options frame
        options_frame = ttk.LabelFrame(scrollable_frame, text="Validation Options", padding=10)
        options_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.validate_tables_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Validate Tables", variable=self.validate_tables_var).pack(anchor=tk.W)
        
        self.validate_columns_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Validate Columns", variable=self.validate_columns_var).pack(anchor=tk.W)
        
        self.validate_indexes_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Validate Indexes", variable=self.validate_indexes_var).pack(anchor=tk.W)
        
        self.validate_constraints_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Validate Constraints", variable=self.validate_constraints_var).pack(anchor=tk.W)
        
        self.validate_programmables_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Validate Programmables (Views, Procedures, Functions)", 
                       variable=self.validate_programmables_var).pack(anchor=tk.W)
        
        # Cross-schema map: e.g. source tables under "userid" compared to "dbo" on destination (SQL Server)
        remap_frame = ttk.LabelFrame(options_frame, text="Schema remap (optional)", padding=6)
        remap_frame.pack(fill=tk.X, pady=(10, 0))
        self.schema_remap_enabled_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            remap_frame,
            text="Map source schema to destination schema (tables/columns; same table name)",
            variable=self.schema_remap_enabled_var,
        ).pack(anchor=tk.W)
        remap_row = ttk.Frame(remap_frame)
        remap_row.pack(anchor=tk.W, pady=4)
        tk.Label(remap_row, text="Source schema:").pack(side=tk.LEFT, padx=(0, 4))
        self.schema_remap_from_var = tk.StringVar(value="")
        ttk.Entry(remap_row, textvariable=self.schema_remap_from_var, width=18).pack(side=tk.LEFT, padx=2)
        tk.Label(remap_row, text="Destination schema:").pack(side=tk.LEFT, padx=(12, 4))
        self.schema_remap_to_var = tk.StringVar(value="dbo")
        ttk.Entry(remap_row, textvariable=self.schema_remap_to_var, width=18).pack(side=tk.LEFT, padx=2)
        tk.Label(
            remap_frame,
            text="Example: userid (source) -> dbo (Azure). Other schemas still match 1:1. "
            "Ignored when source/destination are different platform types (uses cross-platform table-name match). "
            "Index/FK/constraint steps still use physical names on each server.",
            font=("Arial", 8),
            fg="gray",
            wraplength=520,
        ).pack(anchor=tk.W, pady=(4, 0))
        
        # Excel support frame
        excel_frame = ttk.LabelFrame(scrollable_frame, text="Bulk Processing (Excel)", padding=10)
        excel_frame.pack(fill=tk.X, padx=10, pady=10)
        
        excel_btn_frame = ttk.Frame(excel_frame)
        excel_btn_frame.pack(fill=tk.X)
        
        ttk.Button(excel_btn_frame, text="Download Sample Template", 
                  command=lambda: self._download_template("schema_validation")).pack(side=tk.LEFT, padx=5)
        ttk.Button(excel_btn_frame, text="Upload Excel File", 
                  command=self._upload_excel).pack(side=tk.LEFT, padx=5)
        
        self.excel_file_var = tk.StringVar()
        tk.Label(excel_frame, textvariable=self.excel_file_var, fg="gray").pack(anchor=tk.W, pady=5)
        
        self.excel_configs = []
        
        # Buttons
        btn_frame = ttk.Frame(scrollable_frame)
        btn_frame.pack(pady=10)
        
        self.validate_btn = ttk.Button(btn_frame, text="Start Validation", command=self._start_validation, width=20)
        self.validate_btn.pack(side=tk.LEFT, padx=5)
        
        self.bulk_validate_btn = ttk.Button(btn_frame, text="Start Bulk Validation", command=self._start_bulk_validation, 
                                            width=20, state=tk.DISABLED)
        self.bulk_validate_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_btn = ttk.Button(btn_frame, text="Export Report", command=self._export_report, width=20, state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        self.fix_missing_btn = ttk.Button(btn_frame, text="Fix Missing Objects", command=self._fix_missing_objects, 
                                         width=20, state=tk.DISABLED)
        self.fix_missing_btn.pack(side=tk.LEFT, padx=5)
        
        # Results frame
        results_frame = ttk.LabelFrame(scrollable_frame, text="Validation Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Filter frame
        filter_frame = ttk.Frame(results_frame)
        filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(filter_frame, text="Filter by Status:").pack(side=tk.LEFT, padx=5)
        
        self.status_filter_var = tk.StringVar(value="All")
        status_filter_combo = ttk.Combobox(filter_frame, textvariable=self.status_filter_var,
                                          values=["All", "Match OK", "Missing", "Extra only", "Mismatch", "Error", "Auto-Gen"],
                                          state="readonly", width=15)
        status_filter_combo.pack(side=tk.LEFT, padx=5)
        status_filter_combo.bind("<<ComboboxSelected>>", lambda e: self._filter_results())
        
        tk.Label(filter_frame, text="Search:").pack(side=tk.LEFT, padx=(20, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=20)
        search_entry.pack(side=tk.LEFT, padx=5)
        search_entry.bind("<KeyRelease>", lambda e: self._filter_results())
        
        ttk.Button(filter_frame, text="Clear Filter", command=self._clear_filter).pack(side=tk.LEFT, padx=5)
        
        # Store all items for filtering
        self.all_tree_items = []
        
        # Treeview for results
        tree_frame = ttk.Frame(results_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.results_tree = ttk.Treeview(tree_frame, columns=("DB", "Source", "Destination", "Status", "Details"), 
                                         show="tree headings", yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.results_tree.yview)
        
        self.results_tree.heading("#0", text="Object", command=lambda: self._sort_treeview("#0"))
        self.results_tree.heading("DB", text="Database", command=lambda: self._sort_treeview("DB"))
        self.results_tree.heading("Source", text="Source", command=lambda: self._sort_treeview("Source"))
        self.results_tree.heading("Destination", text="Destination", command=lambda: self._sort_treeview("Destination"))
        self.results_tree.heading("Status", text="Status", command=lambda: self._sort_treeview("Status"))
        self.results_tree.heading("Details", text="Details", command=lambda: self._sort_treeview("Details"))
        
        self.results_tree.column("#0", width=200)
        self.results_tree.column("DB", width=250)  # Increased width for database names
        self.results_tree.column("Source", width=150)
        self.results_tree.column("Destination", width=150)
        self.results_tree.column("Status", width=100)
        self.results_tree.column("Details", width=200)
        
        self.treeview_sort_reverse = {}
        
        self.results_tree.pack(fill=tk.BOTH, expand=True)
        
        # Log output
        log_frame = ttk.LabelFrame(scrollable_frame, text="Validation Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.validation_log = scrolledtext.ScrolledText(log_frame, height=10, wrap=tk.WORD)
        self.validation_log.pack(fill=tk.BOTH, expand=True)
        
        self.validation_results = {}
        self._cached_driver = None  # Cache the detected driver
    
    def _get_odbc_driver(self):
        """Auto-detect the best available ODBC driver."""
        if self._cached_driver:
            return self._cached_driver
        
        drivers = pyodbc.drivers()
        preferred = [
            "ODBC Driver 18 for SQL Server",
            "ODBC Driver 17 for SQL Server",
            "ODBC Driver 13 for SQL Server",
            "SQL Server",
        ]
        
        for name in preferred:
            if name in drivers:
                self._cached_driver = name
                return name
        
        # Fallback: find any SQL Server driver
        for d in drivers:
            if "SQL Server" in d:
                self._cached_driver = d
                return d
        
        # Last resort - return the most common one and hope it works
        return "ODBC Driver 17 for SQL Server"
        
    def _get_connection_string(self, server, database, auth, user, password):
        """Build connection string with auto-detected driver."""
        driver_name = self._get_odbc_driver()
        driver = "{" + driver_name + "}"
        
        if auth == "entra_mfa":
            conn_str = f"DRIVER={driver};SERVER={server};DATABASE={database};Authentication=ActiveDirectoryInteractive;UID={user}"
        elif auth == "entra_password":
            conn_str = f"DRIVER={driver};SERVER={server};DATABASE={database};Authentication=ActiveDirectoryPassword;UID={user};PWD={password}"
        elif auth == "sql":
            conn_str = f"DRIVER={driver};SERVER={server};DATABASE={database};UID={user};PWD={password}"
        else:  # windows
            conn_str = f"DRIVER={driver};SERVER={server};DATABASE={database};Trusted_Connection=yes"
            
        return conn_str
        
    def _start_validation(self):
        """Start schema validation in a separate thread."""
        # Validate inputs
        if not self.src_server_var.get():
            messagebox.showerror("Error", "Source server is required!")
            return
        if not self.src_db_var.get():
            messagebox.showerror("Error", "Source database is required!")
            return
        if not self.dest_server_var.get():
            messagebox.showerror("Error", "Destination server is required!")
            return
        if not self.dest_db_var.get():
            messagebox.showerror("Error", "Destination database is required!")
            return
            
        self.validate_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        self.results_tree.delete(*self.results_tree.get_children())
        self.all_tree_items = []  # Clear filter items list
        self.validation_log.delete("1.0", tk.END)
        self.validation_log.insert(tk.END, "Starting schema validation...\n")
        self.validation_results = {}
        
        def run_validation():
            try:
                self.validation_log.insert(tk.END, "Step 1: Connecting to databases...\n")
                self.validation_log.insert(tk.END, f"  Source: {self.src_server_var.get()} | {self.src_db_var.get()}\n")
                self.validation_log.insert(tk.END, f"  Destination: {self.dest_server_var.get()} | {self.dest_db_var.get()}\n")
                self.validation_log.see(tk.END)
                
                try:
                    self.validation_log.insert(tk.END, "  Connecting to source database...\n")
                    self.validation_log.see(tk.END)
                    src_conn = connect_to_any_database(
                        server=self.src_server_var.get(),
                        database=self.src_db_var.get(),
                        auth=self.src_auth_var.get(),
                        user=self.src_user_var.get(),
                        password=self.src_password_var.get() or None,
                        db_type=self.src_db_type_var.get(),
                        port=int(self.src_port_var.get() or 50000),
                        timeout=30
                    )
                    self.validation_log.insert(tk.END, "  [OK] Connected to source database\n")
                    self.validation_log.see(tk.END)
                except Exception as e:
                    self.validation_log.insert(tk.END, f"  [FAILED] Failed to connect to source: {str(e)}\n")
                    self.validation_log.see(tk.END)
                    raise
                
                try:
                    self.validation_log.insert(tk.END, "  Connecting to destination database...\n")
                    self.validation_log.see(tk.END)
                    dest_conn = connect_to_any_database(
                        server=self.dest_server_var.get(),
                        database=self.dest_db_var.get(),
                        auth=self.dest_auth_var.get(),
                        user=self.dest_user_var.get(),
                        password=self.dest_password_var.get() or None,
                        db_type=self.dest_db_type_var.get(),
                        port=int(self.dest_port_var.get() or 50000),
                        timeout=30
                    )
                    self.validation_log.insert(tk.END, "  [OK] Connected to destination database\n")
                    self.validation_log.see(tk.END)
                except Exception as e:
                    self.validation_log.insert(tk.END, f"  [FAILED] Failed to connect to destination: {str(e)}\n")
                    self.validation_log.see(tk.END)
                    if src_conn:
                        src_conn.close()
                    raise
                
                try:
                    src_cur = src_conn.cursor()
                    dest_cur = dest_conn.cursor()
                    
                    # Get db types for appropriate queries
                    src_db_type = self.src_db_type_var.get()
                    dest_db_type = self.dest_db_type_var.get()
                    src_schema_filter = self.src_schema_var.get().strip() if hasattr(self, 'src_schema_var') else None
                    dest_schema_filter = self.dest_schema_var.get().strip() if hasattr(self, 'dest_schema_var') else None
                    
                    # Initialize common_tables in case it's needed later
                    common_tables = set()
                    
                    # Validate tables
                    if self.validate_tables_var.get():
                        self.validation_log.insert(tk.END, "\nStep 2: Validating tables...\n")
                        self.validation_log.insert(tk.END, f"  Source db type: {src_db_type}, Dest db type: {dest_db_type}\n")
                        self.validation_log.insert(tk.END, "  Querying source database for tables...\n")
                        self.validation_log.see(tk.END)
                        
                        # Use DB2 or SQL Server query based on db_type
                        if _DB2_HELPERS_AVAILABLE and src_db_type == 'db2':
                            src_tables_list = fetch_tables_generic(src_cur, 'db2', src_schema_filter)
                        else:
                            def query_src_tables():
                                src_cur.execute("""
                                    SELECT TABLE_SCHEMA, TABLE_NAME
                                    FROM INFORMATION_SCHEMA.TABLES
                                    WHERE TABLE_TYPE = 'BASE TABLE'
                                    ORDER BY TABLE_SCHEMA, TABLE_NAME
                                """)
                                return [(row[0], row[1]) for row in src_cur.fetchall()]
                            
                            src_tables_list = retry_on_deadlock(
                                query_src_tables,
                                max_retries=3,
                                delay=1.0,
                                log_callback=lambda msg: self.validation_log.insert(tk.END, f"  [RETRY] {msg}\n") or self.validation_log.see(tk.END)
                            )
                        
                        # Helper to convert Java strings (from JDBC) to Python strings
                        def to_py_str(val):
                            return str(val).strip() if val is not None else ''
                        
                        # Normalize the table list to Python strings
                        src_tables_list = [(to_py_str(row[0]), to_py_str(row[1])) for row in src_tables_list]
                        src_tables = {f"{row[0]}.{row[1]}" for row in src_tables_list}
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(src_tables)} table(s) in source\n")
                        self.validation_log.see(tk.END)
                        
                        self.validation_log.insert(tk.END, "  Querying destination database for tables...\n")
                        self.validation_log.see(tk.END)
                        
                        if _DB2_HELPERS_AVAILABLE and dest_db_type == 'db2':
                            dest_tables_list = fetch_tables_generic(dest_cur, 'db2', dest_schema_filter)
                        else:
                            def query_dest_tables():
                                dest_cur.execute("""
                                    SELECT TABLE_SCHEMA, TABLE_NAME
                                    FROM INFORMATION_SCHEMA.TABLES
                                    WHERE TABLE_TYPE = 'BASE TABLE'
                                    ORDER BY TABLE_SCHEMA, TABLE_NAME
                                """)
                                return [(row[0], row[1]) for row in dest_cur.fetchall()]
                            
                            dest_tables_list = retry_on_deadlock(
                                query_dest_tables,
                                max_retries=3,
                                delay=1.0,
                                log_callback=lambda msg: self.validation_log.insert(tk.END, f"  [RETRY] {msg}\n") or self.validation_log.see(tk.END)
                            )
                        
                        # Normalize the table list to Python strings
                        dest_tables_list = [(to_py_str(row[0]), to_py_str(row[1])) for row in dest_tables_list]
                        dest_tables = {f"{row[0]}.{row[1]}" for row in dest_tables_list}
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(dest_tables)} table(s) in destination\n")
                        self.validation_log.see(tk.END)
                        
                        self.validation_log.insert(tk.END, "  Comparing table lists...\n")
                        self.validation_log.see(tk.END)
                        
                        # When comparing cross-database types (DB2 to SQL Server), compare by table name only
                        # because schemas are different (e.g., CHAUHS.TABLE vs dbo.TABLE)
                        cross_db_compare = (src_db_type != dest_db_type)
                        
                        if cross_db_compare:
                            self.validation_log.insert(tk.END, "  Note: Cross-database comparison (comparing table names only, ignoring schema)\n")
                            self.validation_log.see(tk.END)
                            
                            # Create maps of table_name -> full_name for matching
                            # (table lists are already normalized to Python strings above)
                            src_table_names = {row[1].upper(): f"{row[0]}.{row[1]}" for row in src_tables_list}
                            dest_table_names = {row[1].upper(): f"{row[0]}.{row[1]}" for row in dest_tables_list}
                            
                            # Find matches by table name (case-insensitive)
                            matching_names = set(src_table_names.keys()) & set(dest_table_names.keys())
                            missing_names = set(src_table_names.keys()) - set(dest_table_names.keys())
                            extra_names = set(dest_table_names.keys()) - set(src_table_names.keys())
                            
                            # Build common_tables with source schema.table for column comparison
                            common_tables = {src_table_names[name] for name in matching_names}
                            missing_in_dest = {src_table_names[name] for name in missing_names}
                            extra_in_dest = {dest_table_names[name] for name in extra_names}
                            
                            # Also store mapping for cross-db column comparison
                            self._cross_db_table_map = {
                                src_table_names[name]: dest_table_names[name] 
                                for name in matching_names
                            }
                        else:
                            schema_remap_active = (
                                self.schema_remap_enabled_var.get()
                                and self.schema_remap_from_var.get().strip()
                                and self.schema_remap_to_var.get().strip()
                            )
                            if schema_remap_active:
                                self.validation_log.insert(
                                    tk.END,
                                    f"  Note: Schema remap ({self.schema_remap_from_var.get().strip()} -> "
                                    f"{self.schema_remap_to_var.get().strip()}): pair tables by name on dest schema\n",
                                )
                                self.validation_log.see(tk.END)
                                try:
                                    common_tables, missing_in_dest, extra_in_dest, self._cross_db_table_map = (
                                        pair_tables_for_schema_remap(
                                            src_tables_list,
                                            dest_tables_list,
                                            self.schema_remap_from_var.get(),
                                            self.schema_remap_to_var.get(),
                                        )
                                    )
                                except ValueError:
                                    common_tables = set()
                                    missing_in_dest = src_tables - dest_tables
                                    extra_in_dest = dest_tables - src_tables
                                    self._cross_db_table_map = {}
                            else:
                                missing_in_dest = src_tables - dest_tables
                                extra_in_dest = dest_tables - src_tables
                                common_tables = src_tables & dest_tables
                                self._cross_db_table_map = {}
                        
                        self.validation_log.insert(tk.END, f"  Results: {len(common_tables)} match, {len(missing_in_dest)} missing in dest, {len(extra_in_dest)} extra in dest\n")
                        self.validation_log.see(tk.END)
                        
                        db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
                        
                        for table in sorted(common_tables):
                            item = self.results_tree.insert("", tk.END, text=table,
                                                   values=(db_name, "Exists", "Exists", "Match OK", ""))
                            self.all_tree_items.append(item)
                        
                        for table in sorted(missing_in_dest):
                            item = self.results_tree.insert("", tk.END, text=table,
                                                   values=(db_name, "Exists", "Missing", "Missing", "Table not in destination"))
                            self.all_tree_items.append(item)
                        
                        for table in sorted(extra_in_dest):
                            item = self.results_tree.insert("", tk.END, text=table,
                                                   values=(db_name, "Missing", "Exists", "Extra only", "Table not in source"))
                            self.all_tree_items.append(item)
                    
                    # Validate columns
                    if self.validate_columns_var.get() and self.validate_tables_var.get():
                        self.validation_log.insert(tk.END, f"\nStep 3: Validating columns for {len(common_tables)} common table(s)...\n")
                        self.validation_log.see(tk.END)
                        
                        table_count = 0
                        for table in sorted(common_tables):
                            table_count += 1
                            self.validation_log.insert(tk.END, f"  [{table_count}/{len(common_tables)}] Validating columns for {table}...\n")
                            self.validation_log.see(tk.END)
                            schema, name = table.split(".", 1)
                            
                            self.validation_log.insert(tk.END, f"    Querying source columns for {table}...\n")
                            self.validation_log.see(tk.END)
                            
                            # Use DB2 or SQL Server query based on db_type
                            if _DB2_HELPERS_AVAILABLE and src_db_type == 'db2':
                                # Fetch DB2 columns with scale info for type mapping
                                src_cur.execute("""
                                    SELECT COLNAME, TYPENAME, LENGTH, SCALE, NULLS
                                    FROM SYSCAT.COLUMNS
                                    WHERE TABSCHEMA = ? AND TABNAME = ?
                                    ORDER BY COLNO
                                """, [schema, name])
                                src_cols = []
                                for row in src_cur.fetchall():
                                    col_name = str(row[0]).strip() if row[0] else ''
                                    data_type = str(row[1]).strip().upper() if row[1] else ''
                                    length = row[2] if row[2] else 0
                                    scale = row[3] if row[3] else 0
                                    nullable = str(row[4]).strip() if row[4] else 'Y'
                                    src_cols.append({
                                        'name': col_name,
                                        'type': data_type,
                                        'length': length,
                                        'scale': scale,
                                        'nullable': 'YES' if nullable == 'Y' else 'NO'
                                    })
                                src_columns = {col['name']: (col['type'], col['length'], col['nullable']) for col in src_cols}
                            else:
                                def query_src_columns():
                                    src_cur.execute("""
                                        SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE
                                        FROM INFORMATION_SCHEMA.COLUMNS
                                        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                                        ORDER BY ORDINAL_POSITION
                                    """, schema, name)
                                    return src_cur.fetchall()
                                
                                src_columns_list = retry_on_deadlock(
                                    query_src_columns,
                                    max_retries=3,
                                    delay=1.0,
                                    log_callback=lambda msg: self.validation_log.insert(tk.END, f"    [RETRY] {msg}\n") or self.validation_log.see(tk.END)
                                )
                                src_columns = {row[0]: row[1:] for row in src_columns_list}
                            self.validation_log.insert(tk.END, f"    [OK] Found {len(src_columns)} column(s) in source\n")
                            
                            self.validation_log.insert(tk.END, f"    Querying destination columns for {table}...\n")
                            self.validation_log.see(tk.END)
                            
                            if _DB2_HELPERS_AVAILABLE and dest_db_type == 'db2':
                                dest_cols = fetch_columns_generic(dest_cur, 'db2', schema, name)
                                dest_columns = {col['name']: (col['type'], col['length'], col['nullable']) for col in dest_cols}
                            else:
                                def query_dest_columns():
                                    dest_cur.execute("""
                                        SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE
                                        FROM INFORMATION_SCHEMA.COLUMNS
                                        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                                        ORDER BY ORDINAL_POSITION
                                    """, schema, name)
                                    return dest_cur.fetchall()
                                
                                dest_columns_list = retry_on_deadlock(
                                    query_dest_columns,
                                    max_retries=3,
                                    delay=1.0,
                                    log_callback=lambda msg: self.validation_log.insert(tk.END, f"    [RETRY] {msg}\n") or self.validation_log.see(tk.END)
                                )
                                dest_columns = {row[0]: row[1:] for row in dest_columns_list}
                            self.validation_log.insert(tk.END, f"    [OK] Found {len(dest_columns)} column(s) in destination\n")
                            
                            self.validation_log.insert(tk.END, f"    Comparing columns...\n")
                            self.validation_log.see(tk.END)
                            db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
                            
                            # Check if this is a cross-database comparison (DB2 to SQL Server)
                            is_cross_db = (src_db_type == 'db2' and dest_db_type == 'sqlserver')
                            
                            # For cross-database, get destination table name from mapping
                            _tbl_map = getattr(self, '_cross_db_table_map', None) or {}
                            if table in _tbl_map:
                                dest_full = _tbl_map[table]
                                dest_schema, dest_name = dest_full.split('.', 1)
                            else:
                                dest_schema, dest_name = schema, name
                            
                            missing_cols = []
                            mismatch_cols = []
                            success_cols = []
                            warning_cols = []
                            
                            # Use type mapping validation for DB2 to SQL Server
                            if is_cross_db and _TYPE_MAPPING_AVAILABLE:
                                # Convert source columns to list of dicts
                                if isinstance(src_cols, list) and len(src_cols) > 0 and isinstance(src_cols[0], dict):
                                    src_cols_list = src_cols
                                else:
                                    # Already fetched as dict
                                    src_cols_list = [{'name': k, 'type': v[0], 'length': v[1], 'nullable': v[2]} 
                                                     for k, v in src_columns.items()]
                                
                                # Get destination columns as list
                                dest_cur.execute("""
                                    SELECT COLUMN_NAME, DATA_TYPE, 
                                           COALESCE(CHARACTER_MAXIMUM_LENGTH, NUMERIC_PRECISION) as LENGTH,
                                           NUMERIC_SCALE as SCALE,
                                           IS_NULLABLE
                                    FROM INFORMATION_SCHEMA.COLUMNS
                                    WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                                    ORDER BY ORDINAL_POSITION
                                """, dest_schema, dest_name)
                                dest_cols_list = []
                                for row in dest_cur.fetchall():
                                    col_name = str(row[0]).strip() if row[0] else ''
                                    data_type = str(row[1]).strip().upper() if row[1] else ''
                                    length = row[2] if row[2] else 0
                                    scale = row[3] if row[3] else 0
                                    nullable = str(row[4]).strip() if row[4] else 'YES'
                                    dest_cols_list.append({
                                        'name': col_name,
                                        'type': data_type,
                                        'length': length,
                                        'scale': scale,
                                        'nullable': nullable
                                    })
                                
                                # Compare with type mapping
                                mapping_results = compare_columns_with_type_mapping(
                                    src_cols_list, dest_cols_list,
                                    src_db_type='db2', dest_db_type='sqlserver'
                                )
                                
                                # Process results
                                for col_info in mapping_results['matching']:
                                    success_cols.append(col_info['column'])
                                    item = self.results_tree.insert("", tk.END, text=f"{table}.{col_info['column']}",
                                                           values=(db_name, col_info['src_type'], col_info['dest_type'], 
                                                                   "Mapped OK", col_info['message']))
                                    self.all_tree_items.append(item)
                                
                                for col_info in mapping_results['type_issues']:
                                    if col_info['status'] == 'WARNING':
                                        warning_cols.append(col_info['column'])
                                        item = self.results_tree.insert("", tk.END, text=f"{table}.{col_info['column']}",
                                                               values=(db_name, col_info['src_type'], col_info['dest_type'], 
                                                                       "Warning", col_info['message']))
                                    else:  # ERROR
                                        mismatch_cols.append(col_info['column'])
                                        item = self.results_tree.insert("", tk.END, text=f"{table}.{col_info['column']}",
                                                               values=(db_name, col_info['src_type'], col_info['dest_type'], 
                                                                       "Mismatch", col_info['message']))
                                    self.all_tree_items.append(item)
                                
                                for col_info in mapping_results['missing_in_dest']:
                                    missing_cols.append(col_info['column'])
                                    item = self.results_tree.insert("", tk.END, text=f"{table}.{col_info['column']}",
                                                           values=(db_name, col_info['src_type'], "Missing", 
                                                                   "Missing", "Column not in destination"))
                                    self.all_tree_items.append(item)
                                
                                for col_info in mapping_results['extra_in_dest']:
                                    item = self.results_tree.insert("", tk.END, text=f"{table}.{col_info['column']}",
                                                           values=(db_name, "Missing", col_info['dest_type'], 
                                                                   "Extra only", "Column only in destination"))
                                    self.all_tree_items.append(item)
                                
                                # Summary message
                                summary = mapping_results['summary']
                                if summary['errors'] == 0 and summary['warnings'] == 0 and len(missing_cols) == 0:
                                    self.validation_log.insert(tk.END, f"    [OK] All {summary['correctly_mapped']} columns correctly mapped\n")
                                else:
                                    self.validation_log.insert(tk.END, 
                                        f"    Results: OK {summary['correctly_mapped']} mapped, WARN {summary['warnings']} warnings, "
                                        f"ERR {summary['errors']} errors, {len(missing_cols)} missing\n")
                            else:
                                # Same database type - direct comparison
                                for col in src_columns:
                                    if col not in dest_columns:
                                        missing_cols.append(col)
                                        item = self.results_tree.insert("", tk.END, text=f"{table}.{col}",
                                                               values=(db_name, "Exists", "Missing", "Missing", "Column not in destination"))
                                        self.all_tree_items.append(item)
                                    elif src_columns[col] != dest_columns[col]:
                                        mismatch_cols.append(col)
                                        src_info = src_columns[col]
                                        dest_info = dest_columns[col]
                                        item = self.results_tree.insert("", tk.END, text=f"{table}.{col}",
                                                               values=(db_name, str(src_info), str(dest_info), "Mismatch", "Column definition differs"))
                                        self.all_tree_items.append(item)
                                    else:
                                        # Columns match - show success for cross-db or if all types match
                                        success_cols.append(col)
                                
                                if missing_cols or mismatch_cols:
                                    self.validation_log.insert(tk.END, f"    [WARN] Issues found: {len(missing_cols)} missing, {len(mismatch_cols)} mismatched\n")
                                else:
                                    self.validation_log.insert(tk.END, f"    [OK] All {len(success_cols)} columns match\n")
                            self.validation_log.see(tk.END)
                    
                    # Validate indexes
                    if self.validate_indexes_var.get():
                        self.validation_log.insert(tk.END, "\nStep 4: Validating indexes...\n")
                        self.validation_log.insert(tk.END, "  Querying source database for indexes with columns...\n")
                        self.validation_log.see(tk.END)
                        
                        # Helper to convert Java strings
                        def to_str(val):
                            return str(val).strip() if val is not None else ''
                        
                        # Fetch indexes with columns for signature matching
                        src_indexes_data = []
                        
                        # Use DB2-specific query if source is DB2
                        if src_db_type == 'db2':
                            schema_filter = src_schema_filter or ''
                            # Fetch indexes with their columns using SYSCAT.INDEXCOLUSE
                            if schema_filter:
                                src_cur.execute("""
                                    SELECT i.INDSCHEMA, i.TABNAME, i.INDNAME, 
                                           CASE WHEN i.INDEXTYPE = 'CLUS' THEN 'CLUSTERED' ELSE 'NONCLUSTERED' END,
                                           CASE WHEN i.UNIQUERULE = 'U' OR i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           CASE WHEN i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           COALESCE(ic.COLNAME, '')
                                    FROM SYSCAT.INDEXES i
                                    LEFT JOIN SYSCAT.INDEXCOLUSE ic ON i.INDSCHEMA = ic.INDSCHEMA AND i.INDNAME = ic.INDNAME
                                    WHERE i.INDSCHEMA = ?
                                    ORDER BY i.INDSCHEMA, i.TABNAME, i.INDNAME, ic.COLSEQ
                                """, [schema_filter])
                            else:
                                src_cur.execute("""
                                    SELECT i.INDSCHEMA, i.TABNAME, i.INDNAME,
                                           CASE WHEN i.INDEXTYPE = 'CLUS' THEN 'CLUSTERED' ELSE 'NONCLUSTERED' END,
                                           CASE WHEN i.UNIQUERULE = 'U' OR i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           CASE WHEN i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           COALESCE(ic.COLNAME, '')
                                    FROM SYSCAT.INDEXES i
                                    LEFT JOIN SYSCAT.INDEXCOLUSE ic ON i.INDSCHEMA = ic.INDSCHEMA AND i.INDNAME = ic.INDNAME
                                    WHERE i.INDSCHEMA NOT LIKE 'SYS%'
                                    ORDER BY i.INDSCHEMA, i.TABNAME, i.INDNAME, ic.COLSEQ
                                """)
                            
                            # Group by index to collect columns
                            index_columns = {}
                            for row in src_cur.fetchall():
                                key = f"{to_str(row[0])}.{to_str(row[1])}.{to_str(row[2])}"
                                if key not in index_columns:
                                    index_columns[key] = {
                                        'name': to_str(row[2]),
                                        'table': to_str(row[1]),
                                        'schema': to_str(row[0]),
                                        'type': to_str(row[3]),
                                        'is_unique': bool(row[4]),
                                        'is_primary': bool(row[5]),
                                        'columns': []
                                    }
                                col = to_str(row[6])
                                if col and col not in index_columns[key]['columns']:
                                    index_columns[key]['columns'].append(col)
                            src_indexes_data = list(index_columns.values())
                        else:
                            # SQL Server - fetch indexes with columns
                            src_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    i.name AS Index_Name,
                                    i.type_desc AS Index_Type,
                                    i.is_unique,
                                    i.is_primary_key,
                                    c.name AS Column_Name
                                FROM sys.indexes i
                                INNER JOIN sys.tables t ON i.object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                LEFT JOIN sys.index_columns ic ON i.object_id = ic.object_id AND i.index_id = ic.index_id
                                LEFT JOIN sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
                                WHERE i.type > 0
                                    AND i.is_primary_key = 0
                                    AND i.is_unique_constraint = 0
                                ORDER BY s.name, t.name, i.name, ic.key_ordinal
                            """)
                            
                            index_columns = {}
                            for row in src_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                if key not in index_columns:
                                    index_columns[key] = {
                                        'name': row[2],
                                        'table': row[1],
                                        'schema': row[0],
                                        'type': row[3],
                                        'is_unique': bool(row[4]),
                                        'is_primary': bool(row[5]),
                                        'columns': []
                                    }
                                if row[6] and row[6] not in index_columns[key]['columns']:
                                    index_columns[key]['columns'].append(row[6])
                            src_indexes_data = list(index_columns.values())
                        
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(src_indexes_data)} index(es) in source\n")
                        self.validation_log.see(tk.END)
                        
                        self.validation_log.insert(tk.END, "  Querying destination database for indexes with columns...\n")
                        self.validation_log.see(tk.END)
                        
                        dest_indexes_data = []
                        
                        # Use DB2-specific query if destination is DB2
                        if dest_db_type == 'db2':
                            schema_filter = dest_schema_filter or ''
                            if schema_filter:
                                dest_cur.execute("""
                                    SELECT i.INDSCHEMA, i.TABNAME, i.INDNAME,
                                           CASE WHEN i.INDEXTYPE = 'CLUS' THEN 'CLUSTERED' ELSE 'NONCLUSTERED' END,
                                           CASE WHEN i.UNIQUERULE = 'U' OR i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           CASE WHEN i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           COALESCE(ic.COLNAME, '')
                                    FROM SYSCAT.INDEXES i
                                    LEFT JOIN SYSCAT.INDEXCOLUSE ic ON i.INDSCHEMA = ic.INDSCHEMA AND i.INDNAME = ic.INDNAME
                                    WHERE i.INDSCHEMA = ?
                                    ORDER BY i.INDSCHEMA, i.TABNAME, i.INDNAME, ic.COLSEQ
                                """, [schema_filter])
                            else:
                                dest_cur.execute("""
                                    SELECT i.INDSCHEMA, i.TABNAME, i.INDNAME,
                                           CASE WHEN i.INDEXTYPE = 'CLUS' THEN 'CLUSTERED' ELSE 'NONCLUSTERED' END,
                                           CASE WHEN i.UNIQUERULE = 'U' OR i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           CASE WHEN i.UNIQUERULE = 'P' THEN 1 ELSE 0 END,
                                           COALESCE(ic.COLNAME, '')
                                    FROM SYSCAT.INDEXES i
                                    LEFT JOIN SYSCAT.INDEXCOLUSE ic ON i.INDSCHEMA = ic.INDSCHEMA AND i.INDNAME = ic.INDNAME
                                    WHERE i.INDSCHEMA NOT LIKE 'SYS%'
                                    ORDER BY i.INDSCHEMA, i.TABNAME, i.INDNAME, ic.COLSEQ
                                """)
                            
                            index_columns = {}
                            for row in dest_cur.fetchall():
                                key = f"{to_str(row[0])}.{to_str(row[1])}.{to_str(row[2])}"
                                if key not in index_columns:
                                    index_columns[key] = {
                                        'name': to_str(row[2]),
                                        'table': to_str(row[1]),
                                        'schema': to_str(row[0]),
                                        'type': to_str(row[3]),
                                        'is_unique': bool(row[4]),
                                        'is_primary': bool(row[5]),
                                        'columns': []
                                    }
                                col = to_str(row[6])
                                if col and col not in index_columns[key]['columns']:
                                    index_columns[key]['columns'].append(col)
                            dest_indexes_data = list(index_columns.values())
                        else:
                            # SQL Server
                            dest_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    i.name AS Index_Name,
                                    i.type_desc AS Index_Type,
                                    i.is_unique,
                                    i.is_primary_key,
                                    c.name AS Column_Name
                                FROM sys.indexes i
                                INNER JOIN sys.tables t ON i.object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                LEFT JOIN sys.index_columns ic ON i.object_id = ic.object_id AND i.index_id = ic.index_id
                                LEFT JOIN sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
                                WHERE i.type > 0
                                    AND i.is_primary_key = 0
                                    AND i.is_unique_constraint = 0
                                ORDER BY s.name, t.name, i.name, ic.key_ordinal
                            """)
                            
                            index_columns = {}
                            for row in dest_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                if key not in index_columns:
                                    index_columns[key] = {
                                        'name': row[2],
                                        'table': row[1],
                                        'schema': row[0],
                                        'type': row[3],
                                        'is_unique': bool(row[4]),
                                        'is_primary': bool(row[5]),
                                        'columns': []
                                    }
                                if row[6] and row[6] not in index_columns[key]['columns']:
                                    index_columns[key]['columns'].append(row[6])
                            dest_indexes_data = list(index_columns.values())
                        
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(dest_indexes_data)} index(es) in destination\n")
                        self.validation_log.insert(tk.END, "  Comparing indexes (using signature matching)...\n")
                        self.validation_log.see(tk.END)
                        
                        db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
                        
                        # Use signature-based matching if available
                        if _SIGNATURE_MATCHING_AVAILABLE and compare_indexes_with_signatures:
                            index_results = compare_indexes_with_signatures(src_indexes_data, dest_indexes_data)
                            
                            summary = index_results['summary']
                            self.validation_log.insert(tk.END, f"  Signature matching: {summary['matched']} matched ({summary['matched_renamed']} renamed), {summary['missing']} missing, {summary['extra']} extra\n")
                            self.validation_log.see(tk.END)
                            
                            # Show matched by name
                            for match_info in index_results.get('matched', []):
                                if match_info.get('match_type') == 'NAME':
                                    src_idx = match_info['source']
                                    idx_name = f"{src_idx.get('schema', '')}.{src_idx.get('table', '')}.{src_idx.get('name', '')}"
                                    # Only show if there are column differences
                                    # item = self.results_tree.insert("", tk.END, text=idx_name,
                                    #                        values=(db_name, "Match", "Match", "Match OK", match_info.get('message', '')))
                                    # self.all_tree_items.append(item)
                            
                            # Show matched by signature (renamed - this is NOT an error!)
                            for match_info in index_results.get('matched', []):
                                if match_info.get('match_type') == 'SIGNATURE':
                                    src_idx = match_info['source']
                                    dest_idx = match_info['dest']
                                    src_name = f"{src_idx.get('schema', '')}.{src_idx.get('table', '')}.{src_idx.get('name', '')}"
                                    dest_name = dest_idx.get('name', '')
                                    cols = ', '.join(src_idx.get('columns', [])[:3])
                                    if len(src_idx.get('columns', [])) > 3:
                                        cols += '...'
                                    item = self.results_tree.insert("", tk.END, text=src_name,
                                                           values=(db_name, src_idx.get('name', ''), dest_name, "Renamed OK", f"Same columns ({cols}), auto-generated names"))
                                    self.all_tree_items.append(item)
                            
                            # Show actually missing (not just renamed)
                            for idx_info in index_results.get('missing_in_dest', []):
                                idx_name = f"{idx_info.get('schema', '')}.{idx_info.get('table', '')}.{idx_info.get('name', '')}"
                                cols = ', '.join(idx_info.get('columns', [])[:3])
                                item = self.results_tree.insert("", tk.END, text=idx_name,
                                                       values=(db_name, "Exists", "Missing", "Missing", f"Columns: {cols}"))
                                self.all_tree_items.append(item)
                            
                            # Show extra in destination
                            for idx_info in index_results.get('extra_in_dest', []):
                                idx_name = f"{idx_info.get('schema', '')}.{idx_info.get('table', '')}.{idx_info.get('name', '')}"
                                cols = ', '.join(idx_info.get('columns', [])[:3])
                                item = self.results_tree.insert("", tk.END, text=idx_name,
                                                       values=(db_name, "Missing", "Exists", "Extra only", f"Columns: {cols}"))
                                self.all_tree_items.append(item)
                        else:
                            # Fallback to name-based comparison
                            self.validation_log.insert(tk.END, "  (Using name-based comparison - signature matching not available)\n")
                            src_indexes = {f"{idx['schema']}.{idx['table']}.{idx['name']}": idx for idx in src_indexes_data}
                            dest_indexes = {f"{idx['schema']}.{idx['table']}.{idx['name']}": idx for idx in dest_indexes_data}
                            
                            missing_in_dest = set(src_indexes.keys()) - set(dest_indexes.keys())
                            extra_in_dest = set(dest_indexes.keys()) - set(src_indexes.keys())
                            
                            for idx in sorted(missing_in_dest):
                                item = self.results_tree.insert("", tk.END, text=idx,
                                                       values=(db_name, "Exists", "Missing", "Missing", "Index not in destination"))
                                self.all_tree_items.append(item)
                            
                            for idx in sorted(extra_in_dest):
                                item = self.results_tree.insert("", tk.END, text=idx,
                                                       values=(db_name, "Missing", "Exists", "Extra only", "Index not in source"))
                                self.all_tree_items.append(item)
                    
                    # Validate constraints (with signature matching for FKs)
                    if self.validate_constraints_var.get():
                        self.validation_log.insert(tk.END, "\nStep 5: Validating constraints (with FK signature matching)...\n")
                        self.validation_log.insert(tk.END, "  Querying source database for foreign keys with columns...\n")
                        self.validation_log.see(tk.END)
                        
                        # Helper to convert Java strings
                        def to_str(val):
                            return str(val).strip() if val is not None else ''
                        
                        # Fetch FKs separately for signature matching
                        src_fks_data = []
                        src_other_constraints = {}
                        src_defaults_data = []  # Default constraints with column info
                        
                        if src_db_type == 'db2':
                            # DB2 doesn't have named default constraints - they're column attributes
                            src_defaults_data = []
                            schema_filter = src_schema_filter or ''
                            # Fetch foreign keys with column details
                            if schema_filter:
                                src_cur.execute("""
                                    SELECT r.TABSCHEMA, r.TABNAME, r.CONSTNAME,
                                           fk.COLNAME, r.REFTABSCHEMA, r.REFTABNAME, pk.COLNAME AS REFCOLNAME,
                                           r.DELETERULE, r.UPDATERULE
                                    FROM SYSCAT.REFERENCES r
                                    LEFT JOIN SYSCAT.KEYCOLUSE fk ON r.TABSCHEMA = fk.TABSCHEMA 
                                        AND r.TABNAME = fk.TABNAME AND r.CONSTNAME = fk.CONSTNAME
                                    LEFT JOIN SYSCAT.KEYCOLUSE pk ON r.REFTABSCHEMA = pk.TABSCHEMA 
                                        AND r.REFTABNAME = pk.TABNAME AND r.REFKEYNAME = pk.CONSTNAME
                                        AND fk.COLSEQ = pk.COLSEQ
                                    WHERE r.TABSCHEMA = ?
                                    ORDER BY r.TABSCHEMA, r.TABNAME, r.CONSTNAME, fk.COLSEQ
                                """, [schema_filter])
                            else:
                                src_cur.execute("""
                                    SELECT r.TABSCHEMA, r.TABNAME, r.CONSTNAME,
                                           fk.COLNAME, r.REFTABSCHEMA, r.REFTABNAME, pk.COLNAME AS REFCOLNAME,
                                           r.DELETERULE, r.UPDATERULE
                                    FROM SYSCAT.REFERENCES r
                                    LEFT JOIN SYSCAT.KEYCOLUSE fk ON r.TABSCHEMA = fk.TABSCHEMA 
                                        AND r.TABNAME = fk.TABNAME AND r.CONSTNAME = fk.CONSTNAME
                                    LEFT JOIN SYSCAT.KEYCOLUSE pk ON r.REFTABSCHEMA = pk.TABSCHEMA 
                                        AND r.REFTABNAME = pk.TABNAME AND r.REFKEYNAME = pk.CONSTNAME
                                        AND fk.COLSEQ = pk.COLSEQ
                                    WHERE r.TABSCHEMA NOT LIKE 'SYS%'
                                    ORDER BY r.TABSCHEMA, r.TABNAME, r.CONSTNAME, fk.COLSEQ
                                """)
                            
                            # Group by FK to collect columns
                            fk_data = {}
                            for row in src_cur.fetchall():
                                key = f"{to_str(row[0])}.{to_str(row[1])}.{to_str(row[2])}"
                                if key not in fk_data:
                                    fk_data[key] = {
                                        'name': to_str(row[2]),
                                        'table': to_str(row[1]),
                                        'schema': to_str(row[0]),
                                        'ref_table': to_str(row[5]),
                                        'ref_schema': to_str(row[4]),
                                        'columns': [],
                                        'ref_columns': [],
                                        'on_delete': to_str(row[7]),
                                        'on_update': to_str(row[8])
                                    }
                                col = to_str(row[3])
                                ref_col = to_str(row[6])
                                if col and col not in fk_data[key]['columns']:
                                    fk_data[key]['columns'].append(col)
                                if ref_col and ref_col not in fk_data[key]['ref_columns']:
                                    fk_data[key]['ref_columns'].append(ref_col)
                            src_fks_data = list(fk_data.values())
                            
                            # Fetch other constraints (check, unique)
                            if schema_filter:
                                src_cur.execute("""
                                    SELECT TABSCHEMA, TABNAME, CONSTNAME, TYPE
                                    FROM SYSCAT.TABCONST
                                    WHERE TABSCHEMA = ? AND TYPE NOT IN ('P', 'F')
                                    ORDER BY TABSCHEMA, TABNAME, CONSTNAME
                                """, [schema_filter])
                            else:
                                src_cur.execute("""
                                    SELECT TABSCHEMA, TABNAME, CONSTNAME, TYPE
                                    FROM SYSCAT.TABCONST
                                    WHERE TABSCHEMA NOT LIKE 'SYS%' AND TYPE NOT IN ('P', 'F')
                                    ORDER BY TABSCHEMA, TABNAME, CONSTNAME
                                """)
                            for row in src_cur.fetchall():
                                key = f"{to_str(row[0])}.{to_str(row[1])}.{to_str(row[2])}"
                                src_other_constraints[key] = to_str(row[3])
                        else:
                            # SQL Server - fetch FKs with columns
                            src_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    fk.name AS FK_Name,
                                    c.name AS Column_Name,
                                    rs.name AS Ref_Schema,
                                    rt.name AS Ref_Table,
                                    rc.name AS Ref_Column,
                                    fk.delete_referential_action_desc,
                                    fk.update_referential_action_desc
                                FROM sys.foreign_keys fk
                                INNER JOIN sys.tables t ON fk.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                INNER JOIN sys.tables rt ON fk.referenced_object_id = rt.object_id
                                INNER JOIN sys.schemas rs ON rt.schema_id = rs.schema_id
                                LEFT JOIN sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
                                LEFT JOIN sys.columns c ON fkc.parent_object_id = c.object_id AND fkc.parent_column_id = c.column_id
                                LEFT JOIN sys.columns rc ON fkc.referenced_object_id = rc.object_id AND fkc.referenced_column_id = rc.column_id
                                ORDER BY s.name, t.name, fk.name, fkc.constraint_column_id
                            """)
                            
                            fk_data = {}
                            for row in src_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                if key not in fk_data:
                                    fk_data[key] = {
                                        'name': row[2],
                                        'table': row[1],
                                        'schema': row[0],
                                        'ref_table': row[5],
                                        'ref_schema': row[4],
                                        'columns': [],
                                        'ref_columns': [],
                                        'on_delete': row[7] or '',
                                        'on_update': row[8] or ''
                                    }
                                if row[3] and row[3] not in fk_data[key]['columns']:
                                    fk_data[key]['columns'].append(row[3])
                                if row[6] and row[6] not in fk_data[key]['ref_columns']:
                                    fk_data[key]['ref_columns'].append(row[6])
                            src_fks_data = list(fk_data.values())
                            
                            # Fetch other constraints (CHECK, UNIQUE - but not DEFAULT, handle separately)
                            src_cur.execute("""
                                SELECT s.name, t.name, c.name, c.type_desc
                                FROM sys.objects c
                                INNER JOIN sys.tables t ON c.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                WHERE c.type IN ('C', 'UQ')
                                ORDER BY s.name, t.name, c.name
                            """)
                            for row in src_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                src_other_constraints[key] = row[3]
                            
                            # Fetch DEFAULT constraints separately with column info for signature matching
                            src_cur.execute("""
                                SELECT s.name AS schema_name, t.name AS table_name, 
                                       dc.name AS constraint_name, col.name AS column_name,
                                       dc.definition
                                FROM sys.default_constraints dc
                                INNER JOIN sys.tables t ON dc.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                INNER JOIN sys.columns col ON dc.parent_object_id = col.object_id 
                                    AND dc.parent_column_id = col.column_id
                                ORDER BY s.name, t.name, dc.name
                            """)
                            src_defaults_data = []
                            for row in src_cur.fetchall():
                                src_defaults_data.append({
                                    'schema': row[0],
                                    'table': row[1],
                                    'name': row[2],
                                    'column': row[3],
                                    'definition': row[4] or ''
                                })
                        
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(src_fks_data)} FK(s) + {len(src_other_constraints)} other constraint(s) in source\n")
                        self.validation_log.see(tk.END)
                        
                        self.validation_log.insert(tk.END, "  Querying destination database for constraints...\n")
                        self.validation_log.see(tk.END)
                        
                        # Fetch destination FKs and other constraints
                        dest_fks_data = []
                        dest_other_constraints = {}
                        dest_defaults_data = []  # Default constraints with column info
                        
                        if dest_db_type == 'db2':
                            # DB2 doesn't have named default constraints
                            dest_defaults_data = []
                            schema_filter = dest_schema_filter or ''
                            if schema_filter:
                                dest_cur.execute("""
                                    SELECT r.TABSCHEMA, r.TABNAME, r.CONSTNAME,
                                           fk.COLNAME, r.REFTABSCHEMA, r.REFTABNAME, pk.COLNAME AS REFCOLNAME,
                                           r.DELETERULE, r.UPDATERULE
                                    FROM SYSCAT.REFERENCES r
                                    LEFT JOIN SYSCAT.KEYCOLUSE fk ON r.TABSCHEMA = fk.TABSCHEMA 
                                        AND r.TABNAME = fk.TABNAME AND r.CONSTNAME = fk.CONSTNAME
                                    LEFT JOIN SYSCAT.KEYCOLUSE pk ON r.REFTABSCHEMA = pk.TABSCHEMA 
                                        AND r.REFTABNAME = pk.TABNAME AND r.REFKEYNAME = pk.CONSTNAME
                                        AND fk.COLSEQ = pk.COLSEQ
                                    WHERE r.TABSCHEMA = ?
                                    ORDER BY r.TABSCHEMA, r.TABNAME, r.CONSTNAME, fk.COLSEQ
                                """, [schema_filter])
                            else:
                                dest_cur.execute("""
                                    SELECT r.TABSCHEMA, r.TABNAME, r.CONSTNAME,
                                           fk.COLNAME, r.REFTABSCHEMA, r.REFTABNAME, pk.COLNAME AS REFCOLNAME,
                                           r.DELETERULE, r.UPDATERULE
                                    FROM SYSCAT.REFERENCES r
                                    LEFT JOIN SYSCAT.KEYCOLUSE fk ON r.TABSCHEMA = fk.TABSCHEMA 
                                        AND r.TABNAME = fk.TABNAME AND r.CONSTNAME = fk.CONSTNAME
                                    LEFT JOIN SYSCAT.KEYCOLUSE pk ON r.REFTABSCHEMA = pk.TABSCHEMA 
                                        AND r.REFTABNAME = pk.TABNAME AND r.REFKEYNAME = pk.CONSTNAME
                                        AND fk.COLSEQ = pk.COLSEQ
                                    WHERE r.TABSCHEMA NOT LIKE 'SYS%'
                                    ORDER BY r.TABSCHEMA, r.TABNAME, r.CONSTNAME, fk.COLSEQ
                                """)
                            
                            fk_data = {}
                            for row in dest_cur.fetchall():
                                key = f"{to_str(row[0])}.{to_str(row[1])}.{to_str(row[2])}"
                                if key not in fk_data:
                                    fk_data[key] = {
                                        'name': to_str(row[2]),
                                        'table': to_str(row[1]),
                                        'schema': to_str(row[0]),
                                        'ref_table': to_str(row[5]),
                                        'ref_schema': to_str(row[4]),
                                        'columns': [],
                                        'ref_columns': [],
                                        'on_delete': to_str(row[7]),
                                        'on_update': to_str(row[8])
                                    }
                                col = to_str(row[3])
                                ref_col = to_str(row[6])
                                if col and col not in fk_data[key]['columns']:
                                    fk_data[key]['columns'].append(col)
                                if ref_col and ref_col not in fk_data[key]['ref_columns']:
                                    fk_data[key]['ref_columns'].append(ref_col)
                            dest_fks_data = list(fk_data.values())
                            
                            if schema_filter:
                                dest_cur.execute("""
                                    SELECT TABSCHEMA, TABNAME, CONSTNAME, TYPE
                                    FROM SYSCAT.TABCONST
                                    WHERE TABSCHEMA = ? AND TYPE NOT IN ('P', 'F')
                                    ORDER BY TABSCHEMA, TABNAME, CONSTNAME
                                """, [schema_filter])
                            else:
                                dest_cur.execute("""
                                    SELECT TABSCHEMA, TABNAME, CONSTNAME, TYPE
                                    FROM SYSCAT.TABCONST
                                    WHERE TABSCHEMA NOT LIKE 'SYS%' AND TYPE NOT IN ('P', 'F')
                                    ORDER BY TABSCHEMA, TABNAME, CONSTNAME
                                """)
                            for row in dest_cur.fetchall():
                                key = f"{to_str(row[0])}.{to_str(row[1])}.{to_str(row[2])}"
                                dest_other_constraints[key] = to_str(row[3])
                        else:
                            # SQL Server
                            dest_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    fk.name AS FK_Name,
                                    c.name AS Column_Name,
                                    rs.name AS Ref_Schema,
                                    rt.name AS Ref_Table,
                                    rc.name AS Ref_Column,
                                    fk.delete_referential_action_desc,
                                    fk.update_referential_action_desc
                                FROM sys.foreign_keys fk
                                INNER JOIN sys.tables t ON fk.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                INNER JOIN sys.tables rt ON fk.referenced_object_id = rt.object_id
                                INNER JOIN sys.schemas rs ON rt.schema_id = rs.schema_id
                                LEFT JOIN sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
                                LEFT JOIN sys.columns c ON fkc.parent_object_id = c.object_id AND fkc.parent_column_id = c.column_id
                                LEFT JOIN sys.columns rc ON fkc.referenced_object_id = rc.object_id AND fkc.referenced_column_id = rc.column_id
                                ORDER BY s.name, t.name, fk.name, fkc.constraint_column_id
                            """)
                            
                            fk_data = {}
                            for row in dest_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                if key not in fk_data:
                                    fk_data[key] = {
                                        'name': row[2],
                                        'table': row[1],
                                        'schema': row[0],
                                        'ref_table': row[5],
                                        'ref_schema': row[4],
                                        'columns': [],
                                        'ref_columns': [],
                                        'on_delete': row[7] or '',
                                        'on_update': row[8] or ''
                                    }
                                if row[3] and row[3] not in fk_data[key]['columns']:
                                    fk_data[key]['columns'].append(row[3])
                                if row[6] and row[6] not in fk_data[key]['ref_columns']:
                                    fk_data[key]['ref_columns'].append(row[6])
                            dest_fks_data = list(fk_data.values())
                            
                            # Fetch other constraints (CHECK, UNIQUE - not DEFAULT)
                            dest_cur.execute("""
                                SELECT s.name, t.name, c.name, c.type_desc
                                FROM sys.objects c
                                INNER JOIN sys.tables t ON c.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                WHERE c.type IN ('C', 'UQ')
                                ORDER BY s.name, t.name, c.name
                            """)
                            for row in dest_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                dest_other_constraints[key] = row[3]
                            
                            # Fetch DEFAULT constraints separately with column info
                            dest_cur.execute("""
                                SELECT s.name AS schema_name, t.name AS table_name, 
                                       dc.name AS constraint_name, col.name AS column_name,
                                       dc.definition
                                FROM sys.default_constraints dc
                                INNER JOIN sys.tables t ON dc.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                INNER JOIN sys.columns col ON dc.parent_object_id = col.object_id 
                                    AND dc.parent_column_id = col.column_id
                                ORDER BY s.name, t.name, dc.name
                            """)
                            dest_defaults_data = []
                            for row in dest_cur.fetchall():
                                dest_defaults_data.append({
                                    'schema': row[0],
                                    'table': row[1],
                                    'name': row[2],
                                    'column': row[3],
                                    'definition': row[4] or ''
                                })
                        
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(dest_fks_data)} FK(s) + {len(dest_other_constraints)} other constraint(s) in destination\n")
                        self.validation_log.insert(tk.END, "  Comparing constraints (using FK signature matching)...\n")
                        self.validation_log.see(tk.END)
                        
                        db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
                        
                        # Use signature-based matching for FKs if available
                        if _SIGNATURE_MATCHING_AVAILABLE and compare_foreign_keys_with_signatures:
                            fk_results = compare_foreign_keys_with_signatures(src_fks_data, dest_fks_data)
                            
                            summary = fk_results['summary']
                            self.validation_log.insert(tk.END, f"  FK signature matching: {summary['matched']} matched ({summary['matched_renamed']} renamed), {summary['missing']} missing, {summary['extra']} extra\n")
                            self.validation_log.see(tk.END)
                            
                            # Show matched by signature (renamed - NOT an error!)
                            for match_info in fk_results.get('matched', []):
                                if match_info.get('match_type') == 'SIGNATURE':
                                    src_fk = match_info['source']
                                    dest_fk = match_info['dest']
                                    src_name = f"{src_fk.get('schema', '')}.{src_fk.get('table', '')}.{src_fk.get('name', '')}"
                                    dest_name = dest_fk.get('name', '')
                                    cols = ', '.join(src_fk.get('columns', [])[:2])
                                    ref_table = src_fk.get('ref_table', '')
                                    item = self.results_tree.insert("", tk.END, text=src_name,
                                                           values=(db_name, src_fk.get('name', ''), dest_name, "Renamed OK", f"FK on ({cols}) -> {ref_table}, auto-generated names"))
                                    self.all_tree_items.append(item)
                            
                            # Show action warnings (same structure but different ON DELETE/UPDATE)
                            for warn_info in fk_results.get('action_mismatches', []):
                                src_fk = warn_info['source']
                                dest_fk = warn_info['dest']
                                fk_name = f"{src_fk.get('schema', '')}.{src_fk.get('table', '')}.{src_fk.get('name', '')}"
                                item = self.results_tree.insert("", tk.END, text=fk_name,
                                                       values=(db_name, f"DEL:{src_fk.get('on_delete')}", f"DEL:{dest_fk.get('on_delete')}", "Action Diff", "FK actions differ"))
                                self.all_tree_items.append(item)
                            
                            # Show actually missing FKs
                            for fk_info in fk_results.get('missing_in_dest', []):
                                fk_name = f"{fk_info.get('schema', '')}.{fk_info.get('table', '')}.{fk_info.get('name', '')}"
                                cols = ', '.join(fk_info.get('columns', [])[:2])
                                ref_table = fk_info.get('ref_table', '')
                                item = self.results_tree.insert("", tk.END, text=fk_name,
                                                       values=(db_name, "Exists", "Missing", "Missing", f"FK on ({cols}) -> {ref_table}"))
                                self.all_tree_items.append(item)
                            
                            # Show extra FKs in destination
                            for fk_info in fk_results.get('extra_in_dest', []):
                                fk_name = f"{fk_info.get('schema', '')}.{fk_info.get('table', '')}.{fk_info.get('name', '')}"
                                cols = ', '.join(fk_info.get('columns', [])[:2])
                                ref_table = fk_info.get('ref_table', '')
                                item = self.results_tree.insert("", tk.END, text=fk_name,
                                                       values=(db_name, "Missing", "Exists", "Extra only", f"FK on ({cols}) -> {ref_table}"))
                                self.all_tree_items.append(item)
                        else:
                            # Fallback to name-based comparison for FKs
                            self.validation_log.insert(tk.END, "  (Using name-based FK comparison - signature matching not available)\n")
                            src_fk_names = {f"{fk['schema']}.{fk['table']}.{fk['name']}": fk for fk in src_fks_data}
                            dest_fk_names = {f"{fk['schema']}.{fk['table']}.{fk['name']}": fk for fk in dest_fks_data}
                            
                            for fk in set(src_fk_names.keys()) - set(dest_fk_names.keys()):
                                item = self.results_tree.insert("", tk.END, text=fk,
                                                       values=(db_name, "Exists", "Missing", "Missing", "FK not in destination"))
                                self.all_tree_items.append(item)
                            for fk in set(dest_fk_names.keys()) - set(src_fk_names.keys()):
                                item = self.results_tree.insert("", tk.END, text=fk,
                                                       values=(db_name, "Missing", "Exists", "Extra only", "FK not in source"))
                                self.all_tree_items.append(item)
                        
                        # Compare other constraints (check, unique) by name
                        missing_other = set(src_other_constraints.keys()) - set(dest_other_constraints.keys())
                        extra_other = set(dest_other_constraints.keys()) - set(src_other_constraints.keys())
                        
                        self.validation_log.insert(tk.END, f"  CHECK/UNIQUE constraints: {len(missing_other)} missing, {len(extra_other)} extra\n")
                        
                        for const in sorted(missing_other):
                            const_type = src_other_constraints[const]
                            item = self.results_tree.insert("", tk.END, text=const,
                                                   values=(db_name, "Exists", "Missing", "Missing", f"{const_type} constraint not in destination"))
                            self.all_tree_items.append(item)
                        
                        for const in sorted(extra_other):
                            const_type = dest_other_constraints[const]
                            item = self.results_tree.insert("", tk.END, text=const,
                                                   values=(db_name, "Missing", "Exists", "Extra only", f"{const_type} constraint not in source"))
                            self.all_tree_items.append(item)
                        
                        # Compare DEFAULT constraints using signature-based matching
                        # Determine if this is a cross-database comparison (DB2 -> SQL Server)
                        is_cross_db = src_db_type != dest_db_type
                        
                        if _SIGNATURE_MATCHING_AVAILABLE and compare_default_constraints_with_signatures:
                            default_results = compare_default_constraints_with_signatures(
                                src_defaults_data, dest_defaults_data, cross_database=is_cross_db
                            )
                            
                            summary = default_results['summary']
                            self.validation_log.insert(tk.END, f"  DEFAULT constraints: {summary['matched']} matched ({summary['matched_renamed']} renamed), ")
                            self.validation_log.insert(tk.END, f"{summary['missing']} missing, {summary['extra']} extra")
                            if summary['auto_generated_ignored'] > 0:
                                self.validation_log.insert(tk.END, f", {summary['auto_generated_ignored']} auto-generated (ignored)")
                            self.validation_log.insert(tk.END, "\n")
                            self.validation_log.see(tk.END)
                            
                            # Show matched by signature (renamed - NOT an error, just info)
                            for match_info in default_results.get('matched', []):
                                if match_info.get('match_type') == 'SIGNATURE':
                                    src_dc = match_info['source']
                                    dest_dc = match_info['dest']
                                    src_name = f"{src_dc.get('schema', '')}.{src_dc.get('table', '')}.{src_dc.get('name', '')}"
                                    dest_name = dest_dc.get('name', '')
                                    col_name = src_dc.get('column', '')
                                    item = self.results_tree.insert("", tk.END, text=src_name,
                                                           values=(db_name, src_dc.get('name', ''), dest_name, "Renamed OK", f"Default on {col_name}, auto-generated names"))
                                    self.all_tree_items.append(item)
                            
                            # Show actually missing defaults (real errors)
                            for dc_info in default_results.get('missing_in_dest', []):
                                dc_name = f"{dc_info.get('schema', '')}.{dc_info.get('table', '')}.{dc_info.get('name', '')}"
                                col_name = dc_info.get('column', '')
                                definition = dc_info.get('definition', '')[:30]
                                item = self.results_tree.insert("", tk.END, text=dc_name,
                                                       values=(db_name, "Exists", "Missing", "Missing", f"Default on {col_name}: {definition}"))
                                self.all_tree_items.append(item)
                            
                            # Show extra defaults that are NOT auto-generated (warnings)
                            for dc_info in default_results.get('extra_in_dest', []):
                                dc_name = f"{dc_info.get('schema', '')}.{dc_info.get('table', '')}.{dc_info.get('name', '')}"
                                col_name = dc_info.get('column', '')
                                definition = dc_info.get('definition', '')[:30]
                                item = self.results_tree.insert("", tk.END, text=dc_name,
                                                       values=(db_name, "Missing", "Exists", "Extra only", f"Default on {col_name}: {definition}"))
                                self.all_tree_items.append(item)
                            
                            # Show auto-generated defaults (ignorable info)
                            for dc_info in default_results.get('auto_generated_in_dest', []):
                                dc_name = f"{dc_info.get('schema', '')}.{dc_info.get('table', '')}.{dc_info.get('name', '')}"
                                col_name = dc_info.get('column', '')
                                item = self.results_tree.insert("", tk.END, text=dc_name,
                                                       values=(db_name, "N/A", "Exists", "Auto-Gen", f"Auto-generated default on {col_name} (safe to ignore)"))
                                self.all_tree_items.append(item)
                        else:
                            # Fallback to simple name-based comparison if signature matching not available
                            src_default_names = {f"{d['schema']}.{d['table']}.{d['name']}": d for d in src_defaults_data}
                            dest_default_names = {f"{d['schema']}.{d['table']}.{d['name']}": d for d in dest_defaults_data}
                            
                            missing_defaults = set(src_default_names.keys()) - set(dest_default_names.keys())
                            extra_defaults = set(dest_default_names.keys()) - set(src_default_names.keys())
                            
                            self.validation_log.insert(tk.END, f"  DEFAULT constraints: {len(missing_defaults)} missing, {len(extra_defaults)} extra\n")
                            
                            for dc in sorted(missing_defaults):
                                item = self.results_tree.insert("", tk.END, text=dc,
                                                       values=(db_name, "Exists", "Missing", "Missing", "DEFAULT_CONSTRAINT not in destination"))
                                self.all_tree_items.append(item)
                            
                            for dc in sorted(extra_defaults):
                                dc_info = dest_default_names[dc]
                                name = dc_info.get('name', '')
                                # Check if auto-generated
                                if is_cross_db and is_auto_generated_constraint_name and is_auto_generated_constraint_name(name):
                                    item = self.results_tree.insert("", tk.END, text=dc,
                                                           values=(db_name, "N/A", "Exists", "Auto-Gen", "Auto-generated (safe to ignore)"))
                                else:
                                    item = self.results_tree.insert("", tk.END, text=dc,
                                                           values=(db_name, "Missing", "Exists", "Extra only", "DEFAULT_CONSTRAINT not in source"))
                                self.all_tree_items.append(item)
                    
                    # Validate all programmables and other objects
                    if self.validate_programmables_var.get():
                        self.validation_log.insert(tk.END, "\nStep 6: Validating programmables and other objects...\n")
                        self.validation_log.insert(tk.END, "  Querying source database for programmables...\n")
                        self.validation_log.see(tk.END)
                        
                        # Use DB2-specific query if source is DB2
                        if src_db_type == 'db2':
                            schema_filter = src_schema_filter or ''
                            if schema_filter:
                                src_cur.execute("""
                                    SELECT ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                    FROM SYSCAT.ROUTINES
                                    WHERE ROUTINESCHEMA = ?
                                    ORDER BY ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                """, [schema_filter])
                            else:
                                src_cur.execute("""
                                    SELECT ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                    FROM SYSCAT.ROUTINES
                                    WHERE ROUTINESCHEMA NOT LIKE 'SYS%'
                                    ORDER BY ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                """)
                        else:
                            src_cur.execute("""
                                SELECT  o.type_desc AS Object_Type,
                                       s.name AS Schema_Name,
                                       o.name AS Object_Name
                                FROM  sys.objects o 
                                JOIN  sys.schemas s
                                  ON  s.schema_id = o.schema_id
                               WHERE  o.type NOT IN ('S', 'U', 'PK', 'D', 'C', 'F', 'IT', 'SQ', 'TR', 'UQ')
                               ORDER BY  Object_Type, Schema_Name, Object_Name
                            """)
                        
                        src_objects_list = src_cur.fetchall()
                        src_objects = {}
                        for row in src_objects_list:
                            obj_type = row[0].strip() if isinstance(row[0], str) else row[0]
                            schema_name = row[1].strip() if isinstance(row[1], str) else row[1]
                            obj_name = row[2].strip() if isinstance(row[2], str) else row[2]
                            key = f"{obj_type}.{schema_name}.{obj_name}"
                            src_objects[key] = obj_type
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(src_objects)} programmable(s) in source\n")
                        self.validation_log.see(tk.END)
                        
                        self.validation_log.insert(tk.END, "  Querying destination database for programmables...\n")
                        self.validation_log.see(tk.END)
                        
                        # Use DB2-specific query if destination is DB2
                        if dest_db_type == 'db2':
                            schema_filter = dest_schema_filter or ''
                            if schema_filter:
                                dest_cur.execute("""
                                    SELECT ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                    FROM SYSCAT.ROUTINES
                                    WHERE ROUTINESCHEMA = ?
                                    ORDER BY ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                """, [schema_filter])
                            else:
                                dest_cur.execute("""
                                    SELECT ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                    FROM SYSCAT.ROUTINES
                                    WHERE ROUTINESCHEMA NOT LIKE 'SYS%'
                                    ORDER BY ROUTINETYPE, ROUTINESCHEMA, ROUTINENAME
                                """)
                        else:
                            dest_cur.execute("""
                                SELECT  o.type_desc AS Object_Type,
                                       s.name AS Schema_Name,
                                       o.name AS Object_Name
                                FROM  sys.objects o 
                                JOIN  sys.schemas s
                                  ON  s.schema_id = o.schema_id
                               WHERE  o.type NOT IN ('S', 'U', 'PK', 'D', 'C', 'F', 'IT', 'SQ', 'TR', 'UQ')
                               ORDER BY  Object_Type, Schema_Name, Object_Name
                            """)
                        
                        dest_objects_list = dest_cur.fetchall()
                        dest_objects = {}
                        for row in dest_objects_list:
                            obj_type = row[0].strip() if isinstance(row[0], str) else row[0]
                            schema_name = row[1].strip() if isinstance(row[1], str) else row[1]
                            obj_name = row[2].strip() if isinstance(row[2], str) else row[2]
                            key = f"{obj_type}.{schema_name}.{obj_name}"
                            dest_objects[key] = obj_type
                        self.validation_log.insert(tk.END, f"  [OK] Found {len(dest_objects)} programmable(s) in destination\n")
                        self.validation_log.insert(tk.END, "  Comparing programmables...\n")
                        self.validation_log.see(tk.END)
                        
                        db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
                        missing_in_dest = set(src_objects.keys()) - set(dest_objects.keys())
                        extra_in_dest = set(dest_objects.keys()) - set(src_objects.keys())
                        common_objects = set(src_objects.keys()) & set(dest_objects.keys())
                        self.validation_log.insert(tk.END, f"  Results: {len(common_objects)} match, {len(missing_in_dest)} missing in dest, {len(extra_in_dest)} extra in dest\n")
                        self.validation_log.see(tk.END)
                        
                        for obj in sorted(common_objects):
                            if src_objects[obj] != dest_objects[obj]:
                                item = self.results_tree.insert("", tk.END, text=obj,
                                                       values=(db_name, "Different", "Different", "Mismatch", "Object type differs"))
                                self.all_tree_items.append(item)
                            else:
                                item = self.results_tree.insert("", tk.END, text=obj,
                                                       values=(db_name, "Exists", "Exists", "Match OK", ""))
                                self.all_tree_items.append(item)
                        
                        for obj in sorted(missing_in_dest):
                            item = self.results_tree.insert("", tk.END, text=obj,
                                                   values=(db_name, "Exists", "Missing", "Missing", f"{src_objects[obj]} not in destination"))
                            self.all_tree_items.append(item)
                        
                        for obj in sorted(extra_in_dest):
                            item = self.results_tree.insert("", tk.END, text=obj,
                                                   values=(db_name, "Missing", "Exists", "Extra only", f"{dest_objects[obj]} not in source"))
                            self.all_tree_items.append(item)
                    
                    self.validation_log.insert(tk.END, "\n[OK] Schema validation completed!\n")
                    self.export_btn.config(state=tk.NORMAL)
                    self.fix_missing_btn.config(state=tk.NORMAL)
                    messagebox.showinfo("Success", "Schema validation completed!")
                    
                finally:
                    # Close cursors and connections
                    try:
                        if 'src_cur' in locals():
                            src_cur.close()
                    except:
                        pass
                    try:
                        if 'dest_cur' in locals():
                            dest_cur.close()
                    except:
                        pass
                    try:
                        if 'src_conn' in locals() and src_conn:
                            src_conn.close()
                    except:
                        pass
                    try:
                        if 'dest_conn' in locals() and dest_conn:
                            dest_conn.close()
                    except:
                        pass
                    
            except Exception as e:
                error_str = str(e)
                self.validation_log.insert(tk.END, f"\n[X] Error: {error_str}\n")
                
                # Check if this is a driver-related error
                if is_driver_missing_error(error_str):
                    self._handle_driver_missing_error(error_str)
                else:
                    messagebox.showerror("Error", f"Validation failed: {error_str}")
            finally:
                self.validate_btn.config(state=tk.NORMAL)
                if len(self.results_tree.get_children()) > 0:
                    self.fix_missing_btn.config(state=tk.NORMAL)
                
        threading.Thread(target=run_validation, daemon=True).start()
    
    def _handle_driver_missing_error(self, error_str: str):
        """Handle ODBC driver missing error - offer installation."""
        if not DRIVER_UTILS_AVAILABLE:
            messagebox.showerror("ODBC Driver Not Found", 
                f"SQL Server ODBC Driver is not installed.\n\n"
                f"Please install Microsoft ODBC Driver 17 or 18 for SQL Server.\n\n"
                f"Download from:\n"
                f"https://learn.microsoft.com/en-us/sql/connect/odbc/download-odbc-driver-for-sql-server\n\n"
                f"Error: {error_str}")
            return
        
        # Check current driver status
        driver_ok, driver_name = check_sql_server_odbc_driver()
        
        if driver_ok:
            # Driver exists but there's still an error - might be a different issue
            messagebox.showerror("Connection Error", 
                f"ODBC Driver is installed ({driver_name}) but connection failed.\n\n"
                f"This might be a configuration issue.\n\n"
                f"Error: {error_str}")
            return
        
        # Offer to install driver
        response = messagebox.askyesnocancel(
            "ODBC Driver Not Found",
            "SQL Server ODBC Driver is not installed on this machine.\n\n"
            "This driver is required for database connections.\n\n"
            "Would you like to install it now?\n\n"
            "Yes: Install automatically (requires admin rights)\n"
            "No: Show manual installation instructions\n"
            "Cancel: Close this dialog"
        )
        
        if response is True:
            # Yes - Try automatic installation
            self._install_odbc_driver()
        elif response is False:
            # No - Show manual instructions
            self._show_manual_install_instructions()
    
    def _install_odbc_driver(self):
        """Install ODBC driver with progress dialog."""
        # Create progress dialog
        progress_win = tk.Toplevel(self.frame)
        progress_win.title("Installing ODBC Driver")
        progress_win.geometry("400x150")
        progress_win.resizable(False, False)
        progress_win.transient(self.frame.winfo_toplevel())
        progress_win.grab_set()
        
        # Center on parent
        progress_win.update_idletasks()
        parent = self.frame.winfo_toplevel()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - 200
        y = parent.winfo_y() + (parent.winfo_height() // 2) - 75
        progress_win.geometry(f"+{x}+{y}")
        
        tk.Label(progress_win, text="Installing ODBC Driver 18 for SQL Server...", 
                 font=("Segoe UI", 10)).pack(pady=20)
        tk.Label(progress_win, text="This may take a few minutes.\nA UAC prompt may appear.", 
                 font=("Segoe UI", 9)).pack()
        
        progress = ttk.Progressbar(progress_win, mode='indeterminate', length=300)
        progress.pack(pady=20)
        progress.start(10)
        
        def run_install():
            try:
                success, message = install_odbc_via_powershell()
                self.frame.after(0, lambda: self._finish_driver_install(progress_win, success, message))
            except Exception as e:
                self.frame.after(0, lambda: self._finish_driver_install(progress_win, False, str(e)))
        
        threading.Thread(target=run_install, daemon=True).start()
    
    def _finish_driver_install(self, progress_win, success, message):
        """Finish driver installation and show result."""
        try:
            progress_win.destroy()
        except:
            pass
        
        if success:
            # Verify installation
            driver_ok, driver_name = check_sql_server_odbc_driver()
            if driver_ok:
                messagebox.showinfo("Success", 
                    f"ODBC Driver installed successfully!\n\n"
                    f"Driver: {driver_name}\n\n"
                    f"You can now retry the validation.")
                self.validation_log.insert(tk.END, f"\n[OK] ODBC Driver installed: {driver_name}\n")
            else:
                messagebox.showinfo("Installation Complete", 
                    message + "\n\nPlease restart the application to use the new driver.")
        else:
            messagebox.showerror("Installation Failed", 
                f"Failed to install ODBC Driver:\n\n{message}\n\n"
                "Please try manual installation.")
            self._show_manual_install_instructions()
    
    def _show_manual_install_instructions(self):
        """Show manual installation instructions for ODBC driver."""
        if DRIVER_UTILS_AVAILABLE:
            instructions = get_manual_install_instructions()
        else:
            import sys
            is_64bit = sys.maxsize > 2**32
            arch = "64-bit" if is_64bit else "32-bit"
            instructions = f"""
ODBC Driver Installation Instructions
=====================================

Your Python installation is {arch}, so you need the matching ODBC driver.

Option 1: Download from Microsoft
---------------------------------
1. Go to: https://learn.microsoft.com/en-us/sql/connect/odbc/download-odbc-driver-for-sql-server
2. Download "ODBC Driver 18 for SQL Server" ({arch})
3. Run the installer and follow the prompts
4. Restart this application

Option 2: Using winget (Windows 11/10)
--------------------------------------
Open Command Prompt as Administrator and run:
    winget install Microsoft.msodbcsql18

After installation, restart this application.
"""
        
        # Show in a scrollable dialog
        dialog = tk.Toplevel(self.frame)
        dialog.title("ODBC Driver Installation Instructions")
        dialog.geometry("600x400")
        dialog.transient(self.frame.winfo_toplevel())
        
        text = tk.Text(dialog, wrap=tk.WORD, font=("Consolas", 10))
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text.insert(tk.END, instructions)
        text.config(state=tk.DISABLED)
        
        tk.Button(dialog, text="Close", command=dialog.destroy).pack(pady=10)
        
    def _export_report(self):
        """Export validation report to Excel with enhanced details."""
        if not self.validation_results and len(self.results_tree.get_children()) == 0:
            messagebox.showwarning("Warning", "No validation results to export!")
            return
        
        # Ask user if they want to export only differences
        export_only_diffs = messagebox.askyesno(
            "Export Options",
            "Export only differences (exclude matched objects)?\n\n"
            "Yes = Only differences (smaller file)\n"
            "No = All objects (including matches)"
        )
            
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if not filename:
            return
            
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill
            from datetime import datetime
            
            # Collect data from treeview
            all_data = []
            table_data = []
            column_data = []
            type_mapping_data = []
            
            for item in self.results_tree.get_children():
                values = self.results_tree.item(item)
                obj = values['text']
                cols = values['values']
                
                if len(cols) >= 5:
                    status = str(cols[3]) if len(cols) > 3 else ""
                    details = str(cols[4]) if len(cols) > 4 else ""
                    
                    # Skip matched if filtering
                    st = str(status)
                    is_matched = (
                        'Match OK' in st or 'Mapped OK' in st or 'Renamed OK' in st
                        or (('Match' in st or '[OK]' in st) and 'Mismatch' not in st)
                    )
                    if export_only_diffs and is_matched:
                        continue
                    
                    row = {
                        "Database": cols[0] if cols[0] else f"{self.src_db_var.get()} vs {self.dest_db_var.get()}",
                        "Object": obj,
                        "Source": cols[1],
                        "Destination": cols[2],
                        "Status": status,
                        "Details": details,
                        "Action Required": self._get_action_required(status)
                    }
                    all_data.append(row)
                    
                    # Categorize by object type
                    if '.' in obj:
                        parts = obj.split('.')
                        if len(parts) == 2:
                            # Table level: Schema.Table
                            table_data.append(row)
                        elif len(parts) >= 3:
                            # Column level: Schema.Table.Column
                            column_data.append({
                                "Table": '.'.join(parts[:-1]),
                                "Column": parts[-1],
                                "Source Type": cols[1] if len(cols) > 1 else "",
                                "Dest Type": cols[2] if len(cols) > 2 else "",
                                "Status": status,
                                "Details": details
                            })
                            
                            # Type mapping details for DB2 to SQL Server
                            src_db_type = self.src_db_type_var.get() if hasattr(self, 'src_db_type_var') else 'sqlserver'
                            dest_db_type = self.dest_db_type_var.get() if hasattr(self, 'dest_db_type_var') else 'sqlserver'
                            
                            if src_db_type == 'db2' and dest_db_type == 'sqlserver':
                                src_type = str(cols[1]) if len(cols) > 1 else ""
                                dest_type = str(cols[2]) if len(cols) > 2 else ""
                                
                                if src_type and src_type != 'Missing':
                                    try:
                                        from gui.utils.db2_type_mapping import get_expected_sql_type
                                        base_type = src_type.split('(')[0] if '(' in src_type else src_type
                                        expected_type = get_expected_sql_type(base_type)
                                        
                                        type_mapping_data.append({
                                            "Table.Column": obj,
                                            "DB2 Type": src_type,
                                            "Expected SQL Type": expected_type,
                                            "Actual SQL Type": dest_type,
                                            "Validation Status": (
                                                'SUCCESS' if ('Match OK' in str(status) or 'Mapped OK' in str(status) or 'Renamed OK' in str(status) or ('[OK]' in str(status) and 'Mismatch' not in str(status)))
                                                else ('WARNING' if ('Extra only' in str(status) or 'Warning' in str(status) or '[WARN]' in str(status)) else 'ERROR')
                                            ),
                                            "Message": details
                                        })
                                    except ImportError:
                                        pass
                    else:
                        table_data.append(row)
            
            def _export_st_ok(d):
                t = str(d.get('Status', ''))
                return 'Match OK' in t or 'Mapped OK' in t or 'Renamed OK' in t or ('[OK]' in t and 'Mismatch' not in t)
            def _export_st_warn(d):
                t = str(d.get('Status', ''))
                return 'Extra only' in t or 'Warning' in t or '[WARN]' in t
            def _export_st_err(d):
                t = str(d.get('Status', ''))
                return 'Missing' in t or 'Mismatch' in t or '[X]' in t or t.strip() == 'Error'

            if filename.endswith('.xlsx'):
                # Export to Excel with multiple sheets
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Summary statistics
                    total = len(all_data)
                    matched = len([d for d in all_data if _export_st_ok(d)])
                    warnings = len([d for d in all_data if _export_st_warn(d)])
                    errors = len([d for d in all_data if _export_st_err(d)])
                    match_rate = (matched / total * 100) if total > 0 else 0
                    
                    summary_data = [
                        {'Metric': 'Total Objects Validated', 'Value': total},
                        {'Metric': 'Matched (OK)', 'Value': matched},
                        {'Metric': 'Warnings', 'Value': warnings},
                        {'Metric': 'Errors', 'Value': errors},
                        {'Metric': 'Match Rate', 'Value': f"{match_rate:.1f}%"},
                        {'Metric': '', 'Value': ''},
                        {'Metric': 'Tables Checked', 'Value': len(table_data)},
                        {'Metric': 'Columns Checked', 'Value': len(column_data)},
                        {'Metric': 'Type Mappings', 'Value': len(type_mapping_data)},
                        {'Metric': '', 'Value': ''},
                        {'Metric': 'Export Mode', 'Value': 'Differences Only' if export_only_diffs else 'All Objects'}
                    ]
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, index=False, sheet_name="Summary")
                    
                    # All validation results
                    if all_data:
                        df = pd.DataFrame(all_data)
                        df.to_excel(writer, index=False, sheet_name="All Results")
                    
                    # Column details sheet
                    if column_data:
                        col_df = pd.DataFrame(column_data)
                        col_df.to_excel(writer, index=False, sheet_name="Column Details")
                    
                    # Type mapping validation sheet
                    if type_mapping_data:
                        type_df = pd.DataFrame(type_mapping_data)
                        type_df.to_excel(writer, index=False, sheet_name="Type Mapping")
                    
                    # Connection info
                    info_data = [
                        {'Property': 'Export Date', 'Value': datetime.now().isoformat()},
                        {'Property': 'Source Server', 'Value': self.src_server_var.get()},
                        {'Property': 'Source Database', 'Value': self.src_db_var.get()},
                        {'Property': 'Source Type', 'Value': self.src_db_type_var.get() if hasattr(self, 'src_db_type_var') else 'sqlserver'},
                        {'Property': 'Source Schema', 'Value': self.src_schema_var.get() if hasattr(self, 'src_schema_var') else 'All'},
                        {'Property': 'Dest Server', 'Value': self.dest_server_var.get()},
                        {'Property': 'Dest Database', 'Value': self.dest_db_var.get()},
                        {'Property': 'Dest Type', 'Value': self.dest_db_type_var.get() if hasattr(self, 'dest_db_type_var') else 'sqlserver'},
                        {'Property': 'Dest Schema', 'Value': self.dest_schema_var.get() if hasattr(self, 'dest_schema_var') else 'All'}
                    ]
                    info_df = pd.DataFrame(info_data)
                    info_df.to_excel(writer, index=False, sheet_name="Connection Info")
                    
                    # Apply color coding to all sheets
                    workbook = writer.book
                    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        # Find Status column
                        status_col = None
                        for col_idx, cell in enumerate(sheet[1], 1):
                            if cell.value and 'Status' in str(cell.value):
                                status_col = col_idx
                                break
                        
                        if status_col:
                            for row_idx in range(2, sheet.max_row + 1):
                                cell = sheet.cell(row=row_idx, column=status_col)
                                status_val = str(cell.value).upper() if cell.value else ''
                                if 'SUCCESS' in status_val or ('MATCH' in status_val and 'MISMATCH' not in status_val) or 'MAPPED' in status_val:
                                    cell.fill = green_fill
                                elif 'WARNING' in status_val or 'EXTRA' in status_val:
                                    cell.fill = yellow_fill
                                elif 'MISSING' in status_val or 'MISMATCH' in status_val or 'ERROR' in status_val:
                                    cell.fill = red_fill
                        
                        # Auto-adjust column widths
                        for column in sheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            sheet.column_dimensions[column_letter].width = adjusted_width
                
                messagebox.showinfo("Success", f"Report exported to Excel:\n{filename}\n\nSheets: Summary, All Results, Column Details, Type Mapping, Connection Info")
            else:
                # Export to JSON
                report = {
                    "timestamp": datetime.now().isoformat(),
                    "source": {
                        "server": self.src_server_var.get(),
                        "database": self.src_db_var.get(),
                        "type": self.src_db_type_var.get() if hasattr(self, 'src_db_type_var') else 'sqlserver'
                    },
                    "destination": {
                        "server": self.dest_server_var.get(),
                        "database": self.dest_db_var.get(),
                        "type": self.dest_db_type_var.get() if hasattr(self, 'dest_db_type_var') else 'sqlserver'
                    },
                    "summary": {
                        "total": len(all_data),
                        "matched": len([d for d in all_data if _export_st_ok(d)]),
                        "warnings": len([d for d in all_data if _export_st_warn(d)]),
                        "errors": len([d for d in all_data if _export_st_err(d)])
                    },
                    "results": all_data,
                    "column_details": column_data,
                    "type_mappings": type_mapping_data
                }
                
                with open(filename, 'w') as f:
                    json.dump(report, f, indent=2)
                    
                messagebox.showinfo("Success", f"Report exported to JSON:\n{filename}")
        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Failed to export report: {str(e)}\n\n{traceback.format_exc()}")
    
    def _get_action_required(self, status):
        """Determine action required based on status."""
        status_str = str(status).upper()
        if 'MATCH' in status_str and 'MISMATCH' not in status_str:
            return 'No Action'
        elif 'MAPPED' in status_str:
            return 'No Action'
        elif 'MISSING' in status_str:
            return 'Create Object'
        elif 'EXTRA' in status_str:
            return 'Review'
        elif 'MISMATCH' in status_str:
            return 'Alter Object'
        elif 'WARNING' in status_str:
            return 'Review'
        else:
            return 'Review'
            
    def _sort_treeview(self, col):
        """Sort treeview by column."""
        # Get all items with their values
        items = []
        for item in self.results_tree.get_children(''):
            if col == "#0":
                # For tree column, use the text
                val = self.results_tree.item(item, "text")
            else:
                # For other columns, use the column value
                val = self.results_tree.set(item, col)
            items.append((val, item))
        
        # Determine sort direction
        reverse = self.treeview_sort_reverse.get(col, False)
        self.treeview_sort_reverse[col] = not reverse
        
        # Sort items
        try:
            # Try numeric sort first (for numeric columns)
            items.sort(key=lambda t: int(str(t[0])) if str(t[0]).isdigit() else (float(str(t[0])) if str(t[0]).replace('.', '').replace('-', '').isdigit() else float('inf')), reverse=reverse)
        except:
            # Fall back to string sort
            items.sort(key=lambda t: str(t[0]).lower() if t[0] else "", reverse=reverse)
        
        # Rearrange items
        for index, (val, item) in enumerate(items):
            self.results_tree.move(item, '', index)
            
    def _download_template(self, template_type: str):
        """Download sample Excel template."""
        try:
            file_path = create_sample_excel(template_type)
            if file_path:
                messagebox.showinfo("Success", f"Template saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template: {str(e)}")
            
    def _upload_excel(self):
        """Upload Excel file for bulk validation."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not file_path:
            return
            
        try:
            configs = read_excel_file(
                file_path,
                required_columns=["src_server", "src_db", "dest_server", "dest_db"],
                default_user=self.src_user_var.get() or None
            )
            
            self.excel_configs = configs
            self.excel_file_var.set(f"Loaded {len(configs)} configuration(s) from {Path(file_path).name}")
            self.bulk_validate_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Loaded {len(configs)} configuration(s) from Excel file!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {str(e)}")
            
    def _start_bulk_validation(self):
        """Start bulk validation from Excel configurations."""
        if not self.excel_configs:
            messagebox.showwarning("Warning", "No configurations loaded! Please upload an Excel file first.")
            return
            
        self.bulk_validate_btn.config(state=tk.DISABLED)
        self.results_tree.delete(*self.results_tree.get_children())
        self.all_tree_items = []  # Clear filter items list
        self.validation_log.delete("1.0", tk.END)
        self.validation_log.insert(tk.END, f"Starting bulk validation for {len(self.excel_configs)} configuration(s)...\n")
        self.validation_results = {}
        
        def run_bulk():
            total_success = 0
            total_fail = 0
            
            for idx, cfg in enumerate(self.excel_configs, 1):
                self.validation_log.insert(tk.END, f"\n{'='*60}\n")
                self.validation_log.insert(tk.END, f"[{idx}/{len(self.excel_configs)}] Validating {cfg.get('src_db')} vs {cfg.get('dest_db')}...\n")
                self.validation_log.see(tk.END)
                
                try:
                    src_auth = cfg.get("src_auth", self.src_auth_var.get()) or self.src_auth_var.get()
                    dest_auth = cfg.get("dest_auth", self.dest_auth_var.get()) or self.dest_auth_var.get()
                    src_user = cfg.get("src_user", cfg.get("user", self.src_user_var.get())) or self.src_user_var.get()
                    dest_user = cfg.get("dest_user", cfg.get("user", self.dest_user_var.get())) or self.dest_user_var.get()
                    src_conn = connect_to_any_database(
                        server=cfg.get("src_server"),
                        database=cfg.get("src_db"),
                        auth=src_auth,
                        user=src_user,
                        password=cfg.get("src_password", self.src_password_var.get()) or None,
                        db_type=cfg.get("src_db_type", self.src_db_type_var.get()),
                        port=int(cfg.get("src_port", self.src_port_var.get()) or 50000),
                        timeout=30
                    )
                    dest_conn = connect_to_any_database(
                        server=cfg.get("dest_server"),
                        database=cfg.get("dest_db"),
                        auth=dest_auth,
                        user=dest_user,
                        password=cfg.get("dest_password", self.dest_password_var.get()) or None,
                        db_type=cfg.get("dest_db_type", self.dest_db_type_var.get()),
                        port=int(cfg.get("dest_port", self.dest_port_var.get()) or 50000),
                        timeout=30
                    )
                    
                    try:
                        with src_conn, dest_conn:
                            src_cur = src_conn.cursor()
                            dest_cur = dest_conn.cursor()
                        
                        # Validate tables
                        # Validate tables
                        if self.validate_tables_var.get():
                            self.validation_log.insert(tk.END, f"[{cfg.get('src_db')}] Validating tables...\n")
                            self.validation_log.see(tk.END)
                            
                            src_cur.execute("""
                                SELECT TABLE_SCHEMA, TABLE_NAME
                                FROM INFORMATION_SCHEMA.TABLES
                                WHERE TABLE_TYPE = 'BASE TABLE'
                                ORDER BY TABLE_SCHEMA, TABLE_NAME
                            """)
                            src_rows_tbl = src_cur.fetchall()
                            src_tables_list = [
                                (str(r[0]).strip(), str(r[1]).strip()) for r in src_rows_tbl
                            ]
                            src_tables = {f"{a}.{b}" for a, b in src_tables_list}
                            
                            dest_cur.execute("""
                                SELECT TABLE_SCHEMA, TABLE_NAME
                                FROM INFORMATION_SCHEMA.TABLES
                                WHERE TABLE_TYPE = 'BASE TABLE'
                                ORDER BY TABLE_SCHEMA, TABLE_NAME
                            """)
                            dest_rows_tbl = dest_cur.fetchall()
                            dest_tables_list = [
                                (str(r[0]).strip(), str(r[1]).strip()) for r in dest_rows_tbl
                            ]
                            dest_tables = {f"{a}.{b}" for a, b in dest_tables_list}
                            
                            src_dt_bulk = (
                                cfg.get("src_db_type", self.src_db_type_var.get()) or "sqlserver"
                            ).lower()
                            dest_dt_bulk = (
                                cfg.get("dest_db_type", self.dest_db_type_var.get()) or "sqlserver"
                            ).lower()
                            cross_db_bulk = src_dt_bulk != dest_dt_bulk
                            bulk_table_map = {}
                            if cross_db_bulk:
                                missing_in_dest = src_tables - dest_tables
                                extra_in_dest = dest_tables - src_tables
                                common_tables = src_tables & dest_tables
                            elif (
                                self.schema_remap_enabled_var.get()
                                and self.schema_remap_from_var.get().strip()
                                and self.schema_remap_to_var.get().strip()
                            ):
                                try:
                                    (
                                        common_tables,
                                        missing_in_dest,
                                        extra_in_dest,
                                        bulk_table_map,
                                    ) = pair_tables_for_schema_remap(
                                        src_tables_list,
                                        dest_tables_list,
                                        self.schema_remap_from_var.get(),
                                        self.schema_remap_to_var.get(),
                                    )
                                except ValueError:
                                    missing_in_dest = src_tables - dest_tables
                                    extra_in_dest = dest_tables - src_tables
                                    common_tables = src_tables & dest_tables
                                    bulk_table_map = {}
                            else:
                                missing_in_dest = src_tables - dest_tables
                                extra_in_dest = dest_tables - src_tables
                                common_tables = src_tables & dest_tables
                            
                            db_name = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                            
                            for table in sorted(common_tables):
                                item = self.results_tree.insert("", tk.END, text=table,
                                                       values=(db_name, "Exists", "Exists", "Match OK", ""))
                                self.all_tree_items.append(item)
                            
                            for table in sorted(missing_in_dest):
                                item = self.results_tree.insert("", tk.END, text=table,
                                                       values=(db_name, "Exists", "Missing", "Missing", "Table not in destination"))
                                self.all_tree_items.append(item)
                            
                            for table in sorted(extra_in_dest):
                                item = self.results_tree.insert("", tk.END, text=table,
                                                       values=(db_name, "Missing", "Exists", "Extra only", "Table not in source"))
                                self.all_tree_items.append(item)
                        
                        # Validate columns
                        if self.validate_columns_var.get() and self.validate_tables_var.get():
                            self.validation_log.insert(tk.END, f"[{cfg.get('src_db')}] Validating columns...\n")
                            self.validation_log.see(tk.END)
                            
                            for table in sorted(common_tables):
                                schema, name = table.split(".", 1)
                                dest_schema, dest_name = schema, name
                                if table in bulk_table_map:
                                    d_full = bulk_table_map[table]
                                    dest_schema, dest_name = d_full.split(".", 1)
                                
                                src_cur.execute("""
                                    SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE
                                    FROM INFORMATION_SCHEMA.COLUMNS
                                    WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                                    ORDER BY ORDINAL_POSITION
                                """, schema, name)
                                src_columns = {row[0]: row[1:] for row in src_cur.fetchall()}
                                
                                dest_cur.execute("""
                                    SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE
                                    FROM INFORMATION_SCHEMA.COLUMNS
                                    WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                                    ORDER BY ORDINAL_POSITION
                                """, dest_schema, dest_name)
                                dest_columns = {row[0]: row[1:] for row in dest_cur.fetchall()}
                                
                                db_name = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                                
                                for col in src_columns:
                                    if col not in dest_columns:
                                        item = self.results_tree.insert("", tk.END, text=f"{table}.{col}",
                                                               values=(db_name, "Exists", "Missing", "Missing", "Column not in destination"))
                                        self.all_tree_items.append(item)
                                    elif src_columns[col] != dest_columns[col]:
                                        item = self.results_tree.insert("", tk.END, text=f"{table}.{col}",
                                                               values=(db_name, "Different", "Different", "Mismatch", "Column definition differs"))
                                        self.all_tree_items.append(item)
                        
                        # Validate indexes
                        if self.validate_indexes_var.get():
                            self.validation_log.insert(tk.END, f"[{cfg.get('src_db')}] Validating indexes...\n")
                            self.validation_log.see(tk.END)
                            
                            src_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    i.name AS Index_Name,
                                    i.type_desc AS Index_Type,
                                    i.is_unique,
                                    i.is_primary_key
                                FROM sys.indexes i
                                INNER JOIN sys.tables t ON i.object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                WHERE i.type > 0
                                    AND i.is_primary_key = 0
                                    AND i.is_unique_constraint = 0
                                ORDER BY s.name, t.name, i.name
                            """)
                            src_indexes = {}
                            for row in src_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                src_indexes[key] = (row[3], row[4], row[5])
                            
                            dest_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    i.name AS Index_Name,
                                    i.type_desc AS Index_Type,
                                    i.is_unique,
                                    i.is_primary_key
                                FROM sys.indexes i
                                INNER JOIN sys.tables t ON i.object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                WHERE i.type > 0
                                    AND i.is_primary_key = 0
                                    AND i.is_unique_constraint = 0
                                ORDER BY s.name, t.name, i.name
                            """)
                            dest_indexes = {}
                            for row in dest_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                dest_indexes[key] = (row[3], row[4], row[5])
                            
                            db_name = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                            missing_in_dest = set(src_indexes.keys()) - set(dest_indexes.keys())
                            extra_in_dest = set(dest_indexes.keys()) - set(src_indexes.keys())
                            common_indexes = set(src_indexes.keys()) & set(dest_indexes.keys())
                            
                            for idx in sorted(common_indexes):
                                if src_indexes[idx] != dest_indexes[idx]:
                                    item = self.results_tree.insert("", tk.END, text=idx,
                                                           values=(db_name, "Different", "Different", "Mismatch", "Index definition differs"))
                                    self.all_tree_items.append(item)
                            
                            for idx in sorted(missing_in_dest):
                                item = self.results_tree.insert("", tk.END, text=idx,
                                                       values=(db_name, "Exists", "Missing", "Missing", "Index not in destination"))
                                self.all_tree_items.append(item)
                            
                            for idx in sorted(extra_in_dest):
                                item = self.results_tree.insert("", tk.END, text=idx,
                                                       values=(db_name, "Missing", "Exists", "Extra only", "Index not in source"))
                                self.all_tree_items.append(item)
                        
                        # Validate primary keys (CLUSTERED vs NONCLUSTERED, columns) for exact replication
                        if self.validate_indexes_var.get():
                            self.validation_log.insert(tk.END, f"[{cfg.get('src_db')}] Validating primary keys...\n")
                            self.validation_log.see(tk.END)
                            # Get PK list with type (SQL Server 2016 compatible - no STRING_AGG)
                            src_cur.execute("""
                                SELECT s.name AS Schema_Name, t.name AS Table_Name, kc.name AS PK_Name, i.type_desc AS Index_Type
                                FROM sys.key_constraints kc
                                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                                JOIN sys.schemas s ON s.schema_id = t.schema_id
                                JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                                WHERE kc.type = 'PK' AND t.is_ms_shipped = 0
                            """)
                            src_pk_rows = src_cur.fetchall()
                            src_cur.execute("""
                                SELECT s.name, t.name, kc.name, c.name AS col_name, ic.key_ordinal
                                FROM sys.key_constraints kc
                                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                                JOIN sys.schemas s ON s.schema_id = t.schema_id
                                JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                                JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id AND ic.is_included_column = 0
                                JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                WHERE kc.type = 'PK' AND t.is_ms_shipped = 0
                                ORDER BY s.name, t.name, kc.name, ic.key_ordinal
                            """)
                            src_pk_cols = {}
                            for row in src_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                if key not in src_pk_cols:
                                    src_pk_cols[key] = []
                                src_pk_cols[key].append(row[3])
                            src_pks = {}
                            for row in src_pk_rows:
                                tbl_key = f"{row[0]}.{row[1]}"
                                obj_name = f"{row[0]}.{row[1]}.{row[2]}"
                                col_list = ",".join(src_pk_cols.get(obj_name, []))
                                src_pks[tbl_key] = (obj_name, (row[3] or "").strip().upper(), col_list)
                            dest_cur.execute("""
                                SELECT s.name AS Schema_Name, t.name AS Table_Name, kc.name AS PK_Name, i.type_desc AS Index_Type
                                FROM sys.key_constraints kc
                                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                                JOIN sys.schemas s ON s.schema_id = t.schema_id
                                JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                                WHERE kc.type = 'PK' AND t.is_ms_shipped = 0
                            """)
                            dest_pk_rows = dest_cur.fetchall()
                            dest_cur.execute("""
                                SELECT s.name, t.name, kc.name, c.name AS col_name, ic.key_ordinal
                                FROM sys.key_constraints kc
                                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                                JOIN sys.schemas s ON s.schema_id = t.schema_id
                                JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                                JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id AND ic.is_included_column = 0
                                JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                WHERE kc.type = 'PK' AND t.is_ms_shipped = 0
                                ORDER BY s.name, t.name, kc.name, ic.key_ordinal
                            """)
                            dest_pk_cols = {}
                            for row in dest_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                if key not in dest_pk_cols:
                                    dest_pk_cols[key] = []
                                dest_pk_cols[key].append(row[3])
                            dest_pks = {}
                            for row in dest_pk_rows:
                                tbl_key = f"{row[0]}.{row[1]}"
                                obj_name = f"{row[0]}.{row[1]}.{row[2]}"
                                col_list = ",".join(dest_pk_cols.get(obj_name, []))
                                dest_pks[tbl_key] = (obj_name, (row[3] or "").strip().upper(), col_list)
                            db_name = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                            all_tables = set(src_pks.keys()) | set(dest_pks.keys())
                            for tbl in sorted(all_tables):
                                src_info = src_pks.get(tbl)
                                dest_info = dest_pks.get(tbl)
                                if not src_info:
                                    obj_name = dest_info[0]
                                    item = self.results_tree.insert("", tk.END, text=obj_name,
                                                           values=(db_name, "Missing", "Exists", "Extra only", "Primary key not in source"))
                                    self.all_tree_items.append(item)
                                elif not dest_info:
                                    obj_name = src_info[0]
                                    item = self.results_tree.insert("", tk.END, text=obj_name,
                                                           values=(db_name, "Exists", "Missing", "Missing", "Primary key not in destination"))
                                    self.all_tree_items.append(item)
                                else:
                                    obj_name = src_info[0]
                                    src_type, src_cols = src_info[1], src_info[2]
                                    dest_type, dest_cols = dest_info[1], dest_info[2]
                                    if src_type != dest_type or src_cols != dest_cols:
                                        detail = "Primary key type or columns differ"
                                        if src_type != dest_type:
                                            detail = f"Primary key type differs ({src_type} vs {dest_type})"
                                        elif src_cols != dest_cols:
                                            detail = "Primary key columns differ"
                                        item = self.results_tree.insert("", tk.END, text=obj_name,
                                                               values=(db_name, "Different", "Different", "Mismatch", detail))
                                        self.all_tree_items.append(item)
                        
                        # Validate constraints
                        if self.validate_constraints_var.get():
                            self.validation_log.insert(tk.END, f"[{cfg.get('src_db')}] Validating constraints...\n")
                            self.validation_log.see(tk.END)
                            
                            src_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    c.name AS Constraint_Name,
                                    c.type_desc AS Constraint_Type
                                FROM sys.objects c
                                INNER JOIN sys.tables t ON c.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                WHERE c.type IN ('F', 'C', 'UQ', 'D')
                                ORDER BY s.name, t.name, c.name
                            """)
                            src_constraints = {f"{row[0]}.{row[1]}.{row[2]}": row[3] for row in src_cur.fetchall()}
                            
                            dest_cur.execute("""
                                SELECT 
                                    s.name AS Schema_Name,
                                    t.name AS Table_Name,
                                    c.name AS Constraint_Name,
                                    c.type_desc AS Constraint_Type
                                FROM sys.objects c
                                INNER JOIN sys.tables t ON c.parent_object_id = t.object_id
                                INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
                                WHERE c.type IN ('F', 'C', 'UQ', 'D')
                                ORDER BY s.name, t.name, c.name
                            """)
                            dest_constraints = {f"{row[0]}.{row[1]}.{row[2]}": row[3] for row in dest_cur.fetchall()}
                            
                            db_name = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                            missing_in_dest = set(src_constraints.keys()) - set(dest_constraints.keys())
                            extra_in_dest = set(dest_constraints.keys()) - set(src_constraints.keys())
                            common_constraints = set(src_constraints.keys()) & set(dest_constraints.keys())
                            
                            for const in sorted(common_constraints):
                                if src_constraints[const] != dest_constraints[const]:
                                    item = self.results_tree.insert("", tk.END, text=const,
                                                           values=(db_name, "Different", "Different", "Mismatch", "Constraint type differs"))
                                    self.all_tree_items.append(item)
                            
                            for const in sorted(missing_in_dest):
                                item = self.results_tree.insert("", tk.END, text=const,
                                                       values=(db_name, "Exists", "Missing", "Missing", "Constraint not in destination"))
                                self.all_tree_items.append(item)
                            
                            for const in sorted(extra_in_dest):
                                item = self.results_tree.insert("", tk.END, text=const,
                                                       values=(db_name, "Missing", "Exists", "Extra only", "Constraint not in source"))
                                self.all_tree_items.append(item)
                        
                        # Validate all programmables and other objects
                        if self.validate_programmables_var.get():
                            self.validation_log.insert(tk.END, f"[{cfg.get('src_db')}] Validating programmables and other objects...\n")
                            self.validation_log.see(tk.END)
                            
                            src_cur.execute("""
                                SELECT  o.type_desc AS Object_Type,
                                       s.name AS Schema_Name,
                                       o.name AS Object_Name
                                FROM  sys.objects o 
                                JOIN  sys.schemas s
                                  ON  s.schema_id = o.schema_id
                               WHERE  o.type NOT IN ('S', 'PK', 'D', 'C', 'F', 'IT', 'SQ', 'TR', 'UQ')
                               ORDER BY  Object_Type, Schema_Name, Object_Name
                            """)
                            src_objects = {}
                            for row in src_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                src_objects[key] = row[0]
                            
                            dest_cur.execute("""
                                SELECT  o.type_desc AS Object_Type,
                                       s.name AS Schema_Name,
                                       o.name AS Object_Name
                                FROM  sys.objects o 
                                JOIN  sys.schemas s
                                  ON  s.schema_id = o.schema_id
                               WHERE  o.type NOT IN ('S', 'PK', 'D', 'C', 'F', 'IT', 'SQ', 'TR', 'UQ')
                               ORDER BY  Object_Type, Schema_Name, Object_Name
                            """)
                            dest_objects = {}
                            for row in dest_cur.fetchall():
                                key = f"{row[0]}.{row[1]}.{row[2]}"
                                dest_objects[key] = row[0]
                            
                            db_name = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                            missing_in_dest = set(src_objects.keys()) - set(dest_objects.keys())
                            extra_in_dest = set(dest_objects.keys()) - set(src_objects.keys())
                            common_objects = set(src_objects.keys()) & set(dest_objects.keys())
                            
                            for obj in sorted(common_objects):
                                if src_objects[obj] != dest_objects[obj]:
                                    item = self.results_tree.insert("", tk.END, text=obj,
                                                           values=(db_name, "Different", "Different", "Mismatch", "Object type differs"))
                                    self.all_tree_items.append(item)
                                else:
                                    item = self.results_tree.insert("", tk.END, text=obj,
                                                           values=(db_name, "Exists", "Exists", "Match OK", ""))
                                    self.all_tree_items.append(item)
                            
                            for obj in sorted(missing_in_dest):
                                item = self.results_tree.insert("", tk.END, text=obj,
                                                       values=(db_name, "Exists", "Missing", "Missing", f"{src_objects[obj]} not in destination"))
                                self.all_tree_items.append(item)
                            
                            for obj in sorted(extra_in_dest):
                                item = self.results_tree.insert("", tk.END, text=obj,
                                                       values=(db_name, "Missing", "Exists", "Extra only", f"{dest_objects[obj]} not in source"))
                                self.all_tree_items.append(item)
                        
                        total_success += 1
                    except Exception as e:
                        self.validation_log.insert(tk.END, f"[X] Error validating {cfg.get('src_db')}: {str(e)}\n")
                        self.validation_log.see(tk.END)
                        if src_conn:
                            try:
                                src_conn.close()
                            except:
                                pass
                        if dest_conn:
                            try:
                                dest_conn.close()
                            except:
                                pass
                        total_fail += 1
                        
                except Exception as e:
                    self.validation_log.insert(tk.END, f"[X] Error: {str(e)}\n")
                    total_fail += 1
                    
            self.validation_log.insert(tk.END, f"\n{'='*60}\n")
            self.validation_log.insert(tk.END, f"Bulk validation completed: {total_success} succeeded, {total_fail} failed\n")
            self.export_btn.config(state=tk.NORMAL)
            if len(self.results_tree.get_children()) > 0:
                self.fix_missing_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Bulk Validation Complete", f"Completed: {total_success} succeeded, {total_fail} failed")
            self.bulk_validate_btn.config(state=tk.NORMAL)
            
        threading.Thread(target=run_bulk, daemon=True).start()
        
    def _get_object_code(self, cur, obj_name):
        """
        Get SQL code for an object from database.
        
        Returns:
            tuple: (source_code, None) or (None, None) if error
        """
        import sys
        from pathlib import Path
        parent_dir = Path(__file__).parent.parent.parent.parent
        sys.path.insert(0, str(parent_dir))
        
        try:
            try:
                from src.backup.exporters import (
                    fetch_columns, fetch_primary_key, build_create_table_sql,
                    object_definition, wrap_create_or_alter, qident, type_sql, parse_int_or_default
                )
            except ImportError:
                try:
                    from schema_backup import (
                        fetch_columns, fetch_primary_key, build_create_table_sql,
                        object_definition, wrap_create_or_alter, qident, type_sql, parse_int_or_default
                    )
                except ImportError:
                    from sechma_backup import (
                        fetch_columns, fetch_primary_key, build_create_table_sql,
                        object_definition, wrap_create_or_alter, qident, type_sql, parse_int_or_default
                    )
        except ImportError:
            return None, None
        
        parts = obj_name.split('.')
        code = None
        
        try:
            if len(parts) == 2:
                # Table: Schema.Table
                schema, table = parts
                cols = fetch_columns(cur, schema, table)
                pk_rows = fetch_primary_key(cur, schema, table)
                code = build_create_table_sql(schema, table, cols, pk_rows)
                
            elif len(parts) == 3:
                # Could be: ObjectType.Schema.Object, Schema.Table.Column, Schema.Table.Index, etc.
                part1, part2, part3 = parts
                
                # Check if it's a programmable (ObjectType.Schema.Object)
                type_map = {
                    'VIEW': 'V', 'SQL_STORED_PROCEDURE': 'P', 'SQL_STOR': 'P',
                    'SQL_SCALAR_FUNCTION': 'FN', 'SQL_TABLE_VALUED_FUNCTION': 'TF',
                    'SQL_INLINE_TABLE_VALUED_FUNCTION': 'IF', 'SQL_FUNCTION': 'FN',
                    'CLR_STORED_PROCEDURE': 'PC', 'CLR_SCALAR_FUNCTION': 'FS',
                    'CLR_TABLE_VALUED_FUNCTION': 'FT', 'CLR_AGGREGATE_FUNCTION': 'AF',
                    'SYNONYM': 'SN', 'SEQUENCE': 'SO'
                }
                
                if part1 in type_map:
                    # It's a programmable
                    obj_type, schema, obj_name_only = parts
                    obj_type_code = type_map.get(obj_type, 'P')
                    
                    cur.execute("""
                        SELECT o.object_id
                        FROM sys.objects o
                        JOIN sys.schemas s ON s.schema_id = o.schema_id
                        WHERE s.name = ? AND o.name = ? AND o.type = ?
                    """, schema, obj_name_only, obj_type_code)
                    
                    row = cur.fetchone()
                    if row:
                        obj_id = row[0]
                        definition = object_definition(cur, obj_id)
                        if definition:
                            code = wrap_create_or_alter(schema, obj_name_only, definition, obj_type)
                else:
                    # Could be Schema.Table.Column, Schema.Table.Index, Schema.Table.Constraint, etc.
                    schema, table, obj_name_only = parts
                    
                    # Check if it's a unique constraint
                    cur.execute("""
                        SELECT 
                            kc.name AS constraint_name,
                            kc.type_desc,
                            kc.is_system_named
                        FROM sys.key_constraints kc
                        WHERE kc.parent_object_id = OBJECT_ID(?, 'U') AND kc.name = ? AND kc.type = 'UQ'
                    """, f"{schema}.{table}", obj_name_only)
                    unique_row = cur.fetchone()
                    if unique_row:
                        # Get constraint columns
                        cur.execute("""
                            SELECT c.name AS column_name, ic.is_descending_key
                            FROM sys.index_columns ic
                            JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                            JOIN sys.key_constraints kc ON kc.unique_index_id = ic.index_id AND kc.parent_object_id = ic.object_id
                            WHERE kc.parent_object_id = OBJECT_ID(?, 'U') AND kc.name = ?
                            ORDER BY ic.key_ordinal
                        """, f"{schema}.{table}", obj_name_only)
                        uq_cols = cur.fetchall()
                        if uq_cols:
                            cols_list = ', '.join([f"{qident(c.column_name)}" for c in uq_cols])
                            code = f"-- Unique Constraint: {obj_name}\nALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(unique_row.constraint_name)} UNIQUE ({cols_list});"
                    
                    # Check if it's a foreign key
                    if not code:
                        cur.execute("""
                            SELECT
                                fk.name AS fk_name,
                                s2.name AS ref_schema_name,
                                t2.name AS ref_table_name,
                                fk.delete_referential_action_desc,
                                fk.update_referential_action_desc
                            FROM sys.foreign_keys fk
                            JOIN sys.tables t1 ON t1.object_id = fk.parent_object_id
                            JOIN sys.schemas s1 ON s1.schema_id = t1.schema_id
                            JOIN sys.tables t2 ON t2.object_id = fk.referenced_object_id
                            JOIN sys.schemas s2 ON s2.schema_id = t2.schema_id
                            WHERE s1.name = ? AND t1.name = ? AND fk.name = ?
                        """, schema, table, obj_name_only)
                        fk_row = cur.fetchone()
                        if fk_row:
                            # Get FK columns
                            cur.execute("""
                                SELECT
                                    pc.name AS parent_column,
                                    rc.name AS ref_column,
                                    fkc.constraint_column_id
                                FROM sys.foreign_key_columns fkc
                                JOIN sys.columns pc
                                    ON pc.object_id = fkc.parent_object_id AND pc.column_id = fkc.parent_column_id
                                JOIN sys.columns rc
                                    ON rc.object_id = fkc.referenced_object_id AND rc.column_id = fkc.referenced_column_id
                                WHERE fkc.constraint_object_id = OBJECT_ID(?, 'F')
                                ORDER BY fkc.constraint_column_id;
                            """, obj_name_only)
                            fk_cols = cur.fetchall()
                            if fk_cols:
                                parent_cols = ", ".join(qident(c.parent_column) for c in fk_cols)
                                ref_cols = ", ".join(qident(c.ref_column) for c in fk_cols)
                                parent = f"{qident(schema)}.{qident(table)}"
                                ref = f"{qident(fk_row.ref_schema_name)}.{qident(fk_row.ref_table_name)}"
                                actions = []
                                if fk_row.delete_referential_action_desc and fk_row.delete_referential_action_desc.upper() != "NO_ACTION":
                                    actions.append(f"ON DELETE {fk_row.delete_referential_action_desc.replace('_', ' ')}")
                                if fk_row.update_referential_action_desc and fk_row.update_referential_action_desc.upper() != "NO_ACTION":
                                    actions.append(f"ON UPDATE {fk_row.update_referential_action_desc.replace('_', ' ')}")
                                code = f"-- Foreign Key: {obj_name}\nALTER TABLE {parent} WITH CHECK ADD CONSTRAINT {qident(fk_row.fk_name)} FOREIGN KEY ({parent_cols}) REFERENCES {ref} ({ref_cols})"
                                if actions:
                                    code += " " + " ".join(actions)
                                code += ";"
                    
                    # Check if it's a check constraint
                    if not code:
                        cur.execute("""
                            SELECT 
                                cc.name AS constraint_name,
                                cc.definition
                            FROM sys.check_constraints cc
                            WHERE cc.parent_object_id = OBJECT_ID(?, 'U') AND cc.name = ?
                        """, f"{schema}.{table}", obj_name_only)
                        check_row = cur.fetchone()
                        if check_row:
                            code = f"-- Check Constraint: {obj_name}\nALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(check_row.constraint_name)} CHECK {check_row.definition};"
                    
                    # Check if it's a default constraint
                    if not code:
                        cur.execute("""
                            SELECT 
                                dc.name AS constraint_name,
                                c.name AS column_name,
                                dc.definition
                            FROM sys.default_constraints dc
                            JOIN sys.columns c ON c.object_id = dc.parent_object_id AND c.column_id = dc.parent_column_id
                            WHERE dc.parent_object_id = OBJECT_ID(?, 'U') AND dc.name = ?
                        """, f"{schema}.{table}", obj_name_only)
                        default_row = cur.fetchone()
                        if default_row:
                            code = f"-- Default Constraint: {obj_name}\nALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(default_row.constraint_name)} DEFAULT {default_row.definition} FOR {qident(default_row.column_name)};"
                    
                    # Check if it's a primary key constraint (for exact replication: CLUSTERED vs NONCLUSTERED)
                    if not code:
                        cur.execute("""
                            SELECT kc.name, i.type_desc
                            FROM sys.key_constraints kc
                            JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                            WHERE kc.parent_object_id = OBJECT_ID(?, 'U') AND kc.name = ? AND kc.type = 'PK'
                        """, f"{schema}.{table}", obj_name_only)
                        pk_row = cur.fetchone()
                        if pk_row:
                            cur.execute("""
                                SELECT c.name, ic.is_descending_key, ic.key_ordinal
                                FROM sys.key_constraints kc
                                JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                                JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id AND ic.is_included_column = 0
                                JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                WHERE kc.parent_object_id = OBJECT_ID(?, 'U') AND kc.name = ?
                                ORDER BY ic.key_ordinal
                            """, f"{schema}.{table}", obj_name_only)
                            pk_cols = cur.fetchall()
                            if pk_cols:
                                type_desc_val = (pk_row.type_desc or "").strip().upper()
                                pk_typ = "CLUSTERED" if type_desc_val == "CLUSTERED" else "NONCLUSTERED"
                                col_list = ", ".join(qident(c.name) + (" DESC" if c.is_descending_key else " ASC") for c in pk_cols)
                                code = f"-- Primary Key: {obj_name}\nALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(pk_row.name)} PRIMARY KEY {pk_typ} (\n\t{col_list}\n);"
                    
                    # Check if it's an index
                    if not code:
                        cur.execute("""
                            SELECT 
                                i.name AS index_name,
                                i.is_unique,
                                i.type_desc,
                                i.has_filter,
                                CAST(i.filter_definition AS NVARCHAR(MAX)) AS filter_definition,
                                i.ignore_dup_key,
                                i.allow_row_locks,
                                i.allow_page_locks,
                                i.fill_factor,
                                i.is_padded,
                                ds.name AS filegroup_name
                            FROM sys.indexes i
                            JOIN sys.tables t ON t.object_id = i.object_id
                            JOIN sys.schemas s ON s.schema_id = t.schema_id
                            LEFT JOIN sys.data_spaces ds ON ds.data_space_id = i.data_space_id
                            WHERE s.name = ? AND t.name = ? AND i.name = ?
                              AND i.is_primary_key = 0
                              AND i.is_unique_constraint = 0
                        """, schema, table, obj_name_only)
                        idx_row = cur.fetchone()
                        if idx_row:
                            # Get index columns
                            cur.execute("""
                                SELECT
                                    c.name AS col_name,
                                    ic.is_descending_key,
                                    ic.is_included_column,
                                    ic.key_ordinal
                                FROM sys.index_columns ic
                                JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                WHERE ic.object_id = OBJECT_ID(?, 'U')
                                  AND ic.index_id = (SELECT index_id FROM sys.indexes WHERE object_id = OBJECT_ID(?, 'U') AND name = ?)
                                ORDER BY ic.is_included_column, ic.key_ordinal;
                            """, f"{schema}.{table}", f"{schema}.{table}", obj_name_only)
                            idx_cols = cur.fetchall()
                            key_cols = []
                            inc_cols = []
                            for c in idx_cols:
                                if c.is_included_column:
                                    inc_cols.append(qident(c.col_name))
                                else:
                                    direction = " DESC" if c.is_descending_key else " ASC"
                                    key_cols.append(qident(c.col_name) + direction)
                            
                            if key_cols:
                                table_name = f"{qident(schema)}.{qident(table)}"
                                unique = "UNIQUE " if idx_row.is_unique else ""
                                type_desc_value = (idx_row.type_desc or "").strip().upper()
                                typ = "CLUSTERED" if type_desc_value == "CLUSTERED" else "NONCLUSTERED"
                                
                                key_cols_formatted = ',\n\t'.join(key_cols)
                                code = f"-- Index: {obj_name}\nCREATE {unique}{typ} INDEX {qident(obj_name_only)} ON {table_name}\n(\n\t{key_cols_formatted}\n)"
                                if inc_cols:
                                    inc_cols_formatted = ',\n\t'.join(inc_cols)
                                    code += f"\nINCLUDE (\n\t{inc_cols_formatted}\n)"
                                
                                if idx_row.has_filter and idx_row.filter_definition:
                                    filter_def = str(idx_row.filter_definition).strip()
                                    if filter_def:
                                        if not filter_def.upper().startswith('WHERE'):
                                            code += f"\nWHERE {filter_def}"
                                        else:
                                            code += f"\n{filter_def}"
                                
                                with_options = []
                                with_options.append("STATISTICS_NORECOMPUTE = OFF")
                                if idx_row.is_unique and idx_row.ignore_dup_key is not None:
                                    with_options.append(f"IGNORE_DUP_KEY = {'ON' if idx_row.ignore_dup_key else 'OFF'}")
                                with_options.append("DROP_EXISTING = OFF")
                                with_options.append("ONLINE = OFF")
                                if idx_row.allow_row_locks is not None:
                                    with_options.append(f"ALLOW_ROW_LOCKS = {'ON' if idx_row.allow_row_locks else 'OFF'}")
                                if idx_row.allow_page_locks is not None:
                                    with_options.append(f"ALLOW_PAGE_LOCKS = {'ON' if idx_row.allow_page_locks else 'OFF'}")
                                if idx_row.fill_factor is not None and idx_row.fill_factor > 0 and idx_row.fill_factor < 100:
                                    with_options.append(f"FILLFACTOR = {idx_row.fill_factor}")
                                    if idx_row.is_padded is not None:
                                        with_options.append(f"PAD_INDEX = {'ON' if idx_row.is_padded else 'OFF'}")
                                
                                if with_options:
                                    options_str = ', \n    '.join(with_options)
                                    code += f"\nWITH (\n    {options_str}\n)"
                                
                                filegroup_name = getattr(idx_row, 'filegroup_name', None) if idx_row else None
                                if filegroup_name and str(filegroup_name).upper() != 'PRIMARY':
                                    code += f" ON [{filegroup_name}]"
                                else:
                                    code += " ON [PRIMARY]"
                                code += ";"
                    
                    # Check if it's a column
                    if not code:
                        cur.execute("""
                            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES 
                            WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                        """, schema, table)
                        if cur.fetchone()[0] > 0:
                            # It's a column
                            cols = fetch_columns(cur, schema, table)
                            col_def = None
                            for col in cols:
                                if col.column_name == obj_name_only:
                                    col_def = col
                                    break
                            
                            if col_def:
                                type_str = type_sql(col_def.type_name, col_def.max_length, col_def.precision, col_def.scale)
                                nullable = "NULL" if col_def.is_nullable else "NOT NULL"
                                code = f"-- Column: {obj_name}\nALTER TABLE {qident(schema)}.{qident(table)} ADD {qident(obj_name_only)} {type_str} {nullable}"
                                if col_def.is_identity:
                                    seed = parse_int_or_default(col_def.seed_value_str, 1)
                                    inc = parse_int_or_default(col_def.increment_value_str, 1)
                                    code = f"-- Column: {obj_name}\nALTER TABLE {qident(schema)}.{qident(table)} ADD {qident(obj_name_only)} {type_str} IDENTITY({seed},{inc}) {nullable}"
                    
                    # If still no code found, object doesn't exist - return None (not error message)
                    # This is expected for missing objects in destination
                    if not code:
                        code = None
        except Exception as e:
            # Return None on error - caller will handle it
            return None, None
        
        return code, None

    def _get_drop_statement(self, cur, obj_name):
        """
        Generate DROP statement for an object on destination so it can be recreated from source.
        Used when fixing Mismatch (definition differs): DROP then CREATE from source.
        Returns DROP SQL string or None if not applicable.
        """
        try:
            try:
                from src.backup.exporters import qident
            except ImportError:
                try:
                    from backup.exporters import qident
                except ImportError:
                    from utils.paths import qident
        except ImportError:
            def qident(n):
                return n if n is None else f"[{str(n).replace(']', ']]')}]"
        try:
            parts = obj_name.split(".")
            if len(parts) == 2:
                return None  # Table - we don't drop tables in fix
            if len(parts) != 3:
                return None
            part1, part2, part3 = parts
            type_map = {
                "VIEW": ("VIEW", "VIEW"),
                "SQL_STORED_PROCEDURE": ("PROCEDURE", "PROCEDURE"),
                "SQL_STOR": ("PROCEDURE", "PROCEDURE"),
                "SQL_SCALAR_FUNCTION": ("FUNCTION", "FUNCTION"),
                "SQL_TABLE_VALUED_FUNCTION": ("FUNCTION", "FUNCTION"),
                "SQL_INLINE_TABLE_VALUED_FUNCTION": ("FUNCTION", "FUNCTION"),
                "SQL_FUNCTION": ("FUNCTION", "FUNCTION"),
                "CLR_STORED_PROCEDURE": ("PROCEDURE", "PROCEDURE"),
                "CLR_SCALAR_FUNCTION": ("FUNCTION", "FUNCTION"),
                "CLR_TABLE_VALUED_FUNCTION": ("FUNCTION", "FUNCTION"),
            }
            if part1 in type_map:
                kind, _ = type_map[part1]
                schema, name = part2, part3
                full = f"{qident(schema)}.{qident(name)}"
                if kind == "VIEW":
                    return f"DROP VIEW IF EXISTS {full};\n"
                if kind == "PROCEDURE":
                    return f"DROP PROCEDURE IF EXISTS {full};\n"
                if kind == "FUNCTION":
                    return f"DROP FUNCTION IF EXISTS {full};\n"
                return None
            schema, table, obj_name_only = parts
            # Index (non-PK, non-unique-constraint)
            cur.execute("""
                SELECT 1 FROM sys.indexes i
                JOIN sys.tables t ON t.object_id = i.object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ? AND i.name = ?
                  AND i.is_primary_key = 0 AND i.is_unique_constraint = 0
            """, (schema, table, obj_name_only))
            if cur.fetchone():
                return f"DROP INDEX IF EXISTS {qident(obj_name_only)} ON {qident(schema)}.{qident(table)};\n"
            # Primary key: drop dependent FKs first, then PK (so DROP CONSTRAINT PK succeeds)
            cur.execute("""
                SELECT 1 FROM sys.key_constraints kc
                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ? AND kc.name = ? AND kc.type = 'PK'
            """, (schema, table, obj_name_only))
            if cur.fetchone():
                drops = []
                cur.execute("""
                    SELECT fk.name AS fk_name,
                           OBJECT_SCHEMA_NAME(fk.parent_object_id) AS child_schema,
                           OBJECT_NAME(fk.parent_object_id) AS child_table
                    FROM sys.foreign_keys fk
                    JOIN sys.tables t ON t.object_id = fk.referenced_object_id
                    JOIN sys.schemas s ON s.schema_id = t.schema_id
                    WHERE s.name = ? AND t.name = ?
                """, (schema, table))
                for row in cur.fetchall():
                    fk_name, child_schema, child_table = row[0], row[1], row[2]
                    drops.append(
                        f"ALTER TABLE {qident(child_schema)}.{qident(child_table)} "
                        f"DROP CONSTRAINT IF EXISTS {qident(fk_name)};\n"
                    )
                drops.append(
                    f"ALTER TABLE {qident(schema)}.{qident(table)} DROP CONSTRAINT IF EXISTS {qident(obj_name_only)};\n"
                )
                return "".join(drops)
            # Constraint (any other type)
            cur.execute("""
                SELECT 1 FROM sys.objects o
                JOIN sys.tables t ON t.object_id = o.parent_object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ? AND o.name = ?
            """, (schema, table, obj_name_only))
            if cur.fetchone():
                return f"ALTER TABLE {qident(schema)}.{qident(table)} DROP CONSTRAINT IF EXISTS {qident(obj_name_only)};\n"
            # Column
            cur.execute("""
                SELECT 1 FROM sys.columns c
                JOIN sys.tables t ON t.object_id = c.object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ? AND c.name = ?
            """, (schema, table, obj_name_only))
            if cur.fetchone():
                return f"ALTER TABLE {qident(schema)}.{qident(table)} DROP COLUMN {qident(obj_name_only)};\n"
            return None
        except Exception:
            return None

    def _build_column_mismatch_script_from_cursors(self, src_cur, dest_cur, obj_name: str):
        """
        For Schema.Table.Column mismatches, build a Red Gate–style script (drop deps on dest,
        ALTER COLUMN to match source including COLLATE, recreate dropped indexes/PK/FKs from source).
        Returns None if not applicable (programmable object path, missing table, or column missing on dest).
        """
        parts = obj_name.split(".")
        if len(parts) != 3:
            return None
        schema, table, column = parts
        if schema in _SCHEMA_VALIDATION_OBJECT_TYPE_PREFIXES:
            return None
        try:
            src_cur.execute(
                """
                SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                """,
                (schema, table),
            )
            if src_cur.fetchone()[0] == 0:
                return None
        except Exception:
            return None
        try:
            from gui.utils.schema_column_align import build_column_mismatch_fix_script
        except ImportError:
            try:
                from azure_migration_tool.gui.utils.schema_column_align import (
                    build_column_mismatch_fix_script,
                )
            except ImportError:
                return None
        try:
            from src.backup.exporters import qident, type_sql
        except ImportError:
            try:
                from sechma_backup.exporters import qident, type_sql
            except ImportError:
                from backup.exporters import qident, type_sql

        def _get_code(cur, oname):
            c, _ = self._get_object_code(cur, oname)
            return c if c else None

        try:
            script, err = build_column_mismatch_fix_script(
                src_cur,
                dest_cur,
                schema,
                table,
                column,
                _get_code,
                qident,
                type_sql,
            )
        except Exception as ex:
            return f"-- Error building column align script: {ex}"

        if err:
            return f"-- {err}"
        if not script:
            return None
        return script

    def _get_table_index_replication_script(self, connection_map, db_name, schema, table):
        """
        Build a single script to replicate source table's index/PK layout on destination:
        1. DROP FKs that reference this table (on destination)
        2. DROP all non-PK indexes on this table (destination)
        3. DROP PK on this table (destination)
        4. ADD PK from source (correct type: CLUSTERED/NONCLUSTERED)
        5. ADD all indexes from source (order: CLUSTERED then NONCLUSTERED)
        6. RESTORE FKs from source (ADD CONSTRAINT ... FOREIGN KEY ...)
        Returns (script_string, None) or (None, error_message). SQL Server only.
        """
        try:
            try:
                from src.backup.exporters import qident
            except ImportError:
                try:
                    from backup.exporters import qident
                except ImportError:
                    from utils.paths import qident
        except ImportError:
            def qident(n):
                return n if n is None else f"[{str(n).replace(']', ']]')}]"

        if db_name not in connection_map:
            return None, "Database not in connection map"
        _, _, src_info, dest_info = connection_map[db_name]
        try:
            src_conn = connect_to_any_database(
                server=src_info[0], database=src_info[1], auth=src_info[2],
                user=src_info[3], password=src_info[4] or None,
                db_type=self.src_db_type_var.get(),
                port=int(self.src_port_var.get() or 50000), timeout=30
            )
            dest_conn = connect_to_any_database(
                server=dest_info[0], database=dest_info[1], auth=dest_info[2],
                user=dest_info[3], password=dest_info[4] or None,
                db_type=self.dest_db_type_var.get(),
                port=int(self.dest_port_var.get() or 50000), timeout=30
            )
        except Exception as e:
            return None, str(e)

        src_cur = src_conn.cursor()
        dest_cur = dest_conn.cursor()
        parts = []

        try:
            # ---- 1. DROP FKs that reference this table (on destination) ----
            dest_cur.execute("""
                SELECT fk.name AS fk_name,
                       OBJECT_SCHEMA_NAME(fk.parent_object_id) AS child_schema,
                       OBJECT_NAME(fk.parent_object_id) AS child_table
                FROM sys.foreign_keys fk
                JOIN sys.tables t ON t.object_id = fk.referenced_object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ?
            """, (schema, table))
            fks_to_restore = list(dest_cur.fetchall())
            parts.append("-- 1. DROP foreign keys that reference this table")
            for row in fks_to_restore:
                fk_name, child_schema, child_table = row[0], row[1], row[2]
                parts.append(
                    f"ALTER TABLE {qident(child_schema)}.{qident(child_table)} "
                    f"DROP CONSTRAINT IF EXISTS {qident(fk_name)};"
                )
            if fks_to_restore:
                parts.append("")

            # ---- 2. DROP non-PK indexes on this table (destination) ----
            dest_cur.execute("""
                SELECT i.name
                FROM sys.indexes i
                JOIN sys.tables t ON t.object_id = i.object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ?
                  AND i.is_primary_key = 0 AND i.is_unique_constraint = 0
                  AND i.type > 0
            """, (schema, table))
            dest_indexes = [row[0] for row in dest_cur.fetchall()]
            parts.append("-- 2. DROP non-PK indexes on this table")
            for idx_name in dest_indexes:
                parts.append(
                    f"DROP INDEX IF EXISTS {qident(idx_name)} ON {qident(schema)}.{qident(table)};"
                )
            if dest_indexes:
                parts.append("")

            # ---- 3. DROP PK on this table (destination) ----
            dest_cur.execute("""
                SELECT kc.name
                FROM sys.key_constraints kc
                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ? AND kc.type = 'PK'
            """, (schema, table))
            pk_row = dest_cur.fetchone()
            parts.append("-- 3. DROP primary key on this table")
            if pk_row:
                parts.append(
                    f"ALTER TABLE {qident(schema)}.{qident(table)} "
                    f"DROP CONSTRAINT IF EXISTS {qident(pk_row[0])};"
                )
                parts.append("")

            # ---- 4 & 5. From SOURCE: get PK and all indexes, ordered (PK first, then CLUSTERED, then NONCLUSTERED) ----
            src_cur.execute("""
                SELECT kc.name, i.type_desc
                FROM sys.key_constraints kc
                JOIN sys.indexes i ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
                JOIN sys.tables t ON t.object_id = kc.parent_object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ? AND kc.type = 'PK'
            """, (schema, table))
            src_pk_row = src_cur.fetchone()
            src_cur.execute("""
                SELECT i.name, i.type_desc
                FROM sys.indexes i
                JOIN sys.tables t ON t.object_id = i.object_id
                JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE s.name = ? AND t.name = ?
                  AND i.is_primary_key = 0 AND i.is_unique_constraint = 0
                  AND i.type > 0
            """, (schema, table))
            src_index_rows = src_cur.fetchall()
            # Order: PK first, then CLUSTERED indexes, then NONCLUSTERED
            create_order = []
            if src_pk_row:
                create_order.append((src_pk_row[0], "PK"))
            for r in src_index_rows:
                create_order.append((r[0], (r[1] or "").strip().upper()))
            create_order.sort(key=lambda x: (0 if x[1] == "PK" else (1 if x[1] == "CLUSTERED" else 2), x[0]))

            parts.append("-- 4. ADD primary key and indexes from source (exact replication)")
            for obj_name_only, _ in create_order:
                full_name = f"{schema}.{table}.{obj_name_only}"
                code, _ = self._get_object_code(src_cur, full_name)
                if code and code.strip():
                    # Strip leading comment line if present for cleaner script
                    lines = code.strip().split("\n")
                    if lines and lines[0].strip().startswith("--"):
                        lines = lines[1:]
                    parts.append("\n".join(lines).strip())
                    parts.append("")

            # ---- 6. RESTORE FKs from source ----
            if fks_to_restore:
                parts.append("-- 5. RESTORE foreign keys that reference this table")
                for row in fks_to_restore:
                    fk_name, child_schema, child_table = row[0], row[1], row[2]
                    # FK lives on child table: child_schema.child_table.fk_name
                    full_name = f"{child_schema}.{child_table}.{fk_name}"
                    code, _ = self._get_object_code(src_cur, full_name)
                    if code and code.strip():
                        lines = code.strip().split("\n")
                        if lines and lines[0].strip().startswith("--"):
                            lines = lines[1:]
                        parts.append("\n".join(lines).strip())
                        parts.append("")

            src_conn.close()
            dest_conn.close()
            return "\n".join(parts).strip(), None
        except Exception as e:
            try:
                src_conn.close()
            except Exception:
                pass
            try:
                dest_conn.close()
            except Exception:
                pass
            return None, str(e)

    def _fix_missing_objects(self):
        """Open comparison dialog for missing objects."""
        # Check if we have Excel configs (bulk validation)
        using_excel = bool(self.excel_configs)
        
        # Validate that we have connection info (only if not using Excel)
        if not using_excel:
            src_server = self.src_server_var.get().strip() if self.src_server_var.get() else ""
            src_db = self.src_db_var.get().strip() if self.src_db_var.get() else ""
            dest_server = self.dest_server_var.get().strip() if self.dest_server_var.get() else ""
            dest_db = self.dest_db_var.get().strip() if self.dest_db_var.get() else ""
            
            if not src_server or not src_db:
                messagebox.showerror("Error", 
                    "Source server and database are required!\n\n" +
                    "Please enter the source server and database in the form above, then run validation first.")
                return
            
            if not dest_server or not dest_db:
                messagebox.showerror("Error", 
                    "Destination server and database are required!\n\n" +
                    "Please enter the destination server and database in the form above, then run validation first.")
                return
        
        # Collect missing and mismatch objects from ALL items (ignore current filter so we always see fixable objects)
        missing_objects = []
        db_configs = {}  # Map database name to config

        if using_excel:
            for cfg in self.excel_configs:
                db_key = f"{cfg.get('src_db')} vs {cfg.get('dest_db')}"
                db_configs[db_key] = cfg

        # Use all_tree_items so filter doesn't hide Missing/Mismatch; if empty, user hasn't run validation
        items_to_scan = getattr(self, "all_tree_items", None) or []
        for item in items_to_scan:
            try:
                status = self.results_tree.set(item, "Status")
            except Exception:
                continue
            is_missing = status and "Missing" in status
            is_mismatch = status and "Mismatch" in status
            if not (is_missing or is_mismatch):
                continue
            obj_name = self.results_tree.item(item, "text")
            db_name = self.results_tree.set(item, "DB")
            if using_excel and (not db_name or not db_name.strip()):
                continue
            if not db_name or not db_name.strip():
                db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
            row_status = "Mismatch" if is_mismatch else "Missing"
            missing_objects.append((obj_name, db_name, row_status))

        if not missing_objects:
            messagebox.showinfo("Info", "No missing or mismatch objects found to fix!\n\n"
                                "Run validation first, then look for 'Missing' or 'Mismatch' in the results.\n"
                                "Clear any status filter (set to 'All') if you expect to see items.")
            return

        # Sort so fix order is deterministic: PK first, then indexes, then FKs, then columns, then tables/other
        def _fix_order_key(item):
            obj_name, db_name, status = item
            parts = obj_name.split(".")
            last = parts[-1] if len(parts) >= 3 else (parts[-1] if parts else "")
            if last.startswith("PK_"):
                return (0, obj_name)
            if last.startswith("IX_") or last.startswith("IDX_") or last.startswith("AK_"):
                return (1, obj_name)
            if last.startswith("FK_"):
                return (2, obj_name)
            if len(parts) == 3:
                return (3, obj_name)  # e.g. schema.table.column or other
            if len(parts) == 2:
                return (4, obj_name)  # table
            return (5, obj_name)
        missing_objects.sort(key=_fix_order_key)

        selected_items = self.results_tree.selection()
        objects_to_fix = []
        if selected_items:
            for item in selected_items:
                try:
                    status = self.results_tree.set(item, "Status")
                except Exception:
                    continue
                if not status or ("Missing" not in status and "Mismatch" not in status):
                    continue
                obj_name = self.results_tree.item(item, "text")
                db_name = self.results_tree.set(item, "DB")
                if not db_name or not db_name.strip():
                    if using_excel:
                        continue
                    db_name = f"{self.src_db_var.get()} vs {self.dest_db_var.get()}"
                row_status = "Mismatch" if "Mismatch" in status else "Missing"
                objects_to_fix.append((obj_name, db_name, row_status))
            if objects_to_fix:
                objects_to_fix.sort(key=_fix_order_key)
        else:
            objects_to_fix = missing_objects

        if not objects_to_fix:
            messagebox.showinfo("Info", "No objects selected to fix!")
            return

        mismatch_count = sum(1 for _o, _d, s in objects_to_fix if s == "Mismatch")
        msg = f"Opening comparison for {len(objects_to_fix)} object(s)"
        if mismatch_count:
            msg += f" ({mismatch_count} with definition difference: will generate DROP then recreate from source)"
        msg += f"\nacross {len(set(d for _o, d, _s in objects_to_fix))} database pair(s)."
        messagebox.showinfo("Opening Bulk Comparison", msg)

        self._open_bulk_comparison_dialog(objects_to_fix, db_configs, using_excel)
    
    def _open_bulk_comparison_dialog(self, objects_to_fix, db_configs, using_excel):
        """Open bulk comparison dialog for all missing objects."""
        if not objects_to_fix:
            return
        
        # Group objects by database pair; each item is (obj_name, db_name, status) with status Missing or Mismatch
        objects_by_db = {}
        for obj_name, db_name, status in objects_to_fix:
            if db_name not in objects_by_db:
                objects_by_db[db_name] = []
            objects_by_db[db_name].append((obj_name, db_name, status))
        
        # Create connection info map for each database pair
        connection_map = {}  # db_name -> (src_conn_str, dest_conn_str, src_info, dest_info)
        
        for db_name in objects_by_db.keys():
            # Get connection info for this database pair
            if using_excel:
                if db_name in db_configs:
                    cfg = db_configs[db_name]
                    src_server = cfg.get("src_server", "")
                    src_db = cfg.get("src_db", "")
                    dest_server = cfg.get("dest_server", "")
                    dest_db = cfg.get("dest_db", "")
                    src_auth = cfg.get("src_auth", self.src_auth_var.get())
                    src_user = cfg.get("src_user", cfg.get("user", self.src_user_var.get()))
                    src_password = cfg.get("src_password", self.src_password_var.get())
                    dest_auth = cfg.get("dest_auth", self.dest_auth_var.get())
                    dest_user = cfg.get("dest_user", cfg.get("user", self.dest_user_var.get()))
                    dest_password = cfg.get("dest_password", self.dest_password_var.get())
                else:
                    messagebox.showerror("Error", f"Database configuration not found for: {db_name}")
                    continue
            else:
                src_server = self.src_server_var.get().strip()
                src_db = self.src_db_var.get().strip()
                dest_server = self.dest_server_var.get().strip()
                dest_db = self.dest_db_var.get().strip()
                src_auth = self.src_auth_var.get()
                src_user = self.src_user_var.get()
                src_password = self.src_password_var.get()
                dest_auth = self.dest_auth_var.get()
                dest_user = self.dest_user_var.get()
                dest_password = self.dest_password_var.get()
            
            # Build connection strings
            src_conn_str = self._get_connection_string(src_server, src_db, src_auth, src_user, src_password)
            dest_conn_str = self._get_connection_string(dest_server, dest_db, dest_auth, dest_user, dest_password)
            
            if "Timeout" not in src_conn_str:
                src_conn_str += ";Connection Timeout=30"
            if "Timeout" not in dest_conn_str:
                dest_conn_str += ";Connection Timeout=30"
            
            connection_map[db_name] = (
                src_conn_str,
                dest_conn_str,
                (src_server, src_db, src_auth, src_user, src_password),
                (dest_server, dest_db, dest_auth, dest_user, dest_password)
            )
        
        # Create fetch code callback that uses the correct connection for each object
        def fetch_code(obj_name, is_source):
            # Find which database pair this object belongs to; obj_info = (obj_name, db_name, status)
            for obj_info in objects_to_fix:
                if obj_info[0] == obj_name:
                    db_name = obj_info[1]
                    fix_status = obj_info[2] if len(obj_info) > 2 else ""
                    if db_name in connection_map:
                        _, _, src_info, dest_info = connection_map[db_name]
                        try:
                            if (
                                is_source
                                and fix_status == "Mismatch"
                                and len(obj_name.split(".")) == 3
                                and obj_name.split(".")[0]
                                not in _SCHEMA_VALIDATION_OBJECT_TYPE_PREFIXES
                            ):
                                src_conn = connect_to_any_database(
                                    server=src_info[0],
                                    database=src_info[1],
                                    auth=src_info[2],
                                    user=src_info[3],
                                    password=src_info[4] or None,
                                    db_type=self.src_db_type_var.get(),
                                    port=int(self.src_port_var.get() or 50000),
                                    timeout=30,
                                )
                                dest_conn = connect_to_any_database(
                                    server=dest_info[0],
                                    database=dest_info[1],
                                    auth=dest_info[2],
                                    user=dest_info[3],
                                    password=dest_info[4] or None,
                                    db_type=self.dest_db_type_var.get(),
                                    port=int(self.dest_port_var.get() or 50000),
                                    timeout=30,
                                )
                                try:
                                    sc = src_conn.cursor()
                                    dc = dest_conn.cursor()
                                    alt = self._build_column_mismatch_script_from_cursors(
                                        sc, dc, obj_name
                                    )
                                    if alt:
                                        return alt
                                finally:
                                    try:
                                        src_conn.close()
                                    except Exception:
                                        pass
                                    try:
                                        dest_conn.close()
                                    except Exception:
                                        pass
                            # Use connection info to connect
                            conn_info = src_info if is_source else dest_info
                            db_type_var = self.src_db_type_var if is_source else self.dest_db_type_var
                            port_var = self.src_port_var if is_source else self.dest_port_var
                            conn = connect_to_any_database(
                                server=conn_info[0],
                                database=conn_info[1],
                                auth=conn_info[2],
                                user=conn_info[3],
                                password=conn_info[4] or None,
                                db_type=db_type_var.get(),
                                port=int(port_var.get() or 50000),
                                timeout=30
                            )
                            cur = conn.cursor()
                            code, _ = self._get_object_code(cur, obj_name)
                            conn.close()
                            return code if code else ""
                        except Exception as e:
                            # For destination, if object doesn't exist, return empty string (expected)
                            # For source, this is an error
                            if is_source:
                                return f"-- Error fetching source code: {str(e)}"
                            return ""  # Destination object doesn't exist - this is expected
            return ""
        
        # Create deploy callback that uses the correct connection for each object
        def deploy_callback(obj_name, sql_script):
            # Find which database pair this object belongs to; obj_info = (obj_name, db_name, status)
            for obj_info in objects_to_fix:
                if obj_info[0] == obj_name:
                    db_name = obj_info[1]
                    if db_name in connection_map:
                        _, _, src_info, dest_info = connection_map[db_name]
                        return self._deploy_object(obj_name, sql_script,
                            src_info[0], src_info[1], src_info[2], src_info[3], src_info[4],
                            dest_info[0], dest_info[1], dest_info[2], dest_info[3], dest_info[4])
            return False, "Database configuration not found for object"

        # Callback to get DROP statement for destination object (for Mismatch: DROP then recreate from source)
        def get_drop_for_object(obj_name, db_name):
            parts = obj_name.split(".")
            if (
                len(parts) == 3
                and parts[0] not in _SCHEMA_VALIDATION_OBJECT_TYPE_PREFIXES
            ):
                # Column mismatch: full script (drops + ALTER + recreate) is returned from fetch_code; do not prepend DROP COLUMN.
                return ""
            if db_name not in connection_map:
                return ""
            _, _, src_info, dest_info = connection_map[db_name]
            try:
                conn = connect_to_any_database(
                    server=dest_info[0],
                    database=dest_info[1],
                    auth=dest_info[2],
                    user=dest_info[3],
                    password=dest_info[4] or None,
                    db_type=self.dest_db_type_var.get(),
                    port=int(self.dest_port_var.get() or 50000),
                    timeout=15
                )
                cur = conn.cursor()
                drop_sql = self._get_drop_statement(cur, obj_name)
                conn.close()
                return (drop_sql or "").strip()
            except Exception:
                return ""

        # Callback to get full table index/PK replication script (drop FKs, drop indexes/PK, recreate from source, restore FKs)
        def get_table_script_callback(db_name, schema, table):
            return self._get_table_index_replication_script(connection_map, db_name, schema, table)

        # Open bulk dialog
        dialog = BulkSchemaComparisonDialog(
            self.frame.winfo_toplevel(),
            objects_to_fix,
            fetch_code,
            deploy_callback,
            get_drop_callback=get_drop_for_object,
            get_table_script_callback=get_table_script_callback
        )
    
    def _open_comparison_dialog(self, obj_info, all_objects, db_configs, using_excel):
        """Open comparison dialog for an object."""
        obj_name, source_status, db_name = obj_info
        
        # Get connection info
        if using_excel:
            if db_name in db_configs:
                cfg = db_configs[db_name]
                src_server = cfg.get("src_server", "")
                src_db = cfg.get("src_db", "")
                dest_server = cfg.get("dest_server", "")
                dest_db = cfg.get("dest_db", "")
                src_auth = cfg.get("src_auth", self.src_auth_var.get())
                src_user = cfg.get("src_user", cfg.get("user", self.src_user_var.get()))
                src_password = cfg.get("src_password", self.src_password_var.get())
                dest_auth = cfg.get("dest_auth", self.dest_auth_var.get())
                dest_user = cfg.get("dest_user", cfg.get("user", self.dest_user_var.get()))
                dest_password = cfg.get("dest_password", self.dest_password_var.get())
            else:
                messagebox.showerror("Error", f"Database configuration not found for: {db_name}")
                return
        else:
            src_server = self.src_server_var.get().strip()
            src_db = self.src_db_var.get().strip()
            dest_server = self.dest_server_var.get().strip()
            dest_db = self.dest_db_var.get().strip()
            src_auth = self.src_auth_var.get()
            src_user = self.src_user_var.get()
            src_password = self.src_password_var.get()
            dest_auth = self.dest_auth_var.get()
            dest_user = self.dest_user_var.get()
            dest_password = self.dest_password_var.get()
        
        # Get source and destination code in a thread
        loading_dialog = tk.Toplevel(self.frame.winfo_toplevel())
        loading_dialog.title("Loading...")
        loading_dialog.geometry("300x100")
        tk.Label(loading_dialog, text=f"Loading object definition for:\n{obj_name}").pack(expand=True)
        loading_dialog.update()
        
        def fetch_codes():
            try:
                _oname, source_status, _dbn = obj_info
                src_conn = connect_to_any_database(
                    server=src_server,
                    database=src_db,
                    auth=src_auth,
                    user=src_user,
                    password=src_password or None,
                    db_type=self.src_db_type_var.get(),
                    port=int(self.src_port_var.get() or 50000),
                    timeout=30
                )
                dest_conn = connect_to_any_database(
                    server=dest_server,
                    database=dest_db,
                    auth=dest_auth,
                    user=dest_user,
                    password=dest_password or None,
                    db_type=self.dest_db_type_var.get(),
                    port=int(self.dest_port_var.get() or 50000),
                    timeout=30
                )
                
                src_cur = src_conn.cursor()
                dest_cur = dest_conn.cursor()
                
                source_code, _ = self._get_object_code(src_cur, obj_name)
                dest_code, _ = self._get_object_code(dest_cur, obj_name)
                if source_status == "Mismatch":
                    alt_script = self._build_column_mismatch_script_from_cursors(
                        src_cur, dest_cur, obj_name
                    )
                    if alt_script:
                        source_code = alt_script
                
                src_conn.close()
                dest_conn.close()
                
                # Open dialog on main thread
                loading_dialog.after(0, lambda: self._show_dialog(loading_dialog, obj_name, source_code, dest_code, 
                    obj_info, all_objects, db_configs, using_excel,
                    src_server, src_db, src_auth, src_user, src_password,
                    dest_server, dest_db, dest_auth, dest_user, dest_password))
            except Exception as e:
                loading_dialog.after(0, lambda: self._show_error(loading_dialog, f"Failed to fetch object code:\n{str(e)}"))
        
        threading.Thread(target=fetch_codes, daemon=True).start()
    
    def _show_error(self, loading_dialog, error_msg):
        """Show error message and close loading dialog."""
        loading_dialog.destroy()
        messagebox.showerror("Error", error_msg)
    
    def _show_dialog(self, loading_dialog, obj_name, source_code, dest_code, obj_info, all_objects, db_configs, using_excel,
                    src_server, src_db, src_auth, src_user, src_password,
                    dest_server, dest_db, dest_auth, dest_user, dest_password):
        """Show the comparison dialog."""
        loading_dialog.destroy()  # Close loading window
        
        if not source_code:
            messagebox.showerror("Error", f"Could not retrieve source code for: {obj_name}")
            return
        
        # Create deploy callback
        def deploy_callback(obj_name, sql_script):
            return self._deploy_object(obj_name, sql_script,
                src_server, src_db, src_auth, src_user, src_password,
                dest_server, dest_db, dest_auth, dest_user, dest_password)
        
        # Open comparison dialog
        dialog = SchemaComparisonDialog(
            self.frame.winfo_toplevel(),
            obj_name,
            source_code,
            dest_code,
            deploy_callback
        )
        
        # Store references for processing next object
        dialog.remaining_objects = all_objects[1:] if len(all_objects) > 1 else []
        dialog.db_configs = db_configs
        dialog.using_excel = using_excel
        dialog.src_server = src_server
        dialog.src_db = src_db
        dialog.dest_server = dest_server
        dialog.dest_db = dest_db
        dialog.src_auth = src_auth
        dialog.src_user = src_user
        dialog.src_password = src_password
        dialog.dest_auth = dest_auth
        dialog.dest_user = dest_user
        dialog.dest_password = dest_password
        dialog.parent_tab = self
    
    def _deploy_object(self, obj_name, sql_script,
                      src_server, src_db, src_auth, src_user, src_password,
                      dest_server, dest_db, dest_auth, dest_user, dest_password):
        """Deploy SQL script to destination database."""
        try:
            dest_conn = connect_to_any_database(
                server=dest_server,
                database=dest_db,
                auth=dest_auth,
                user=dest_user,
                password=dest_password or None,
                db_type=self.dest_db_type_var.get(),
                port=int(self.dest_port_var.get() or 50000),
                timeout=30
            )
            dest_cur = dest_conn.cursor()
            
            # Execute script
            batch_count = 0
            for batch in sql_script.split('GO'):
                batch = batch.strip()
                if batch:
                    batch_count += 1
                    dest_cur.execute(batch)
            
            dest_conn.commit()
            dest_conn.close()
            
            return True, f"Successfully deployed {obj_name}"
        except Exception as e:
            return False, f"Deployment failed: {str(e)}"
    
    def _fix_objects_for_db(self, missing_objects, force_fix_mode,
                           src_server, src_db, src_auth, src_user, src_password,
                           dest_server, dest_db, dest_auth, dest_user, dest_password):
        """Fix missing objects for a specific database pair."""
        try:
            # Get connection strings
            self.validation_log.insert(tk.END, "\n" + "="*60 + "\n")
            self.validation_log.insert(tk.END, "FIX OPERATION: Starting to fix missing objects...\n")
            self.validation_log.insert(tk.END, f"  Total objects to fix: {len(missing_objects)}\n")
            self.validation_log.insert(tk.END, f"  Force fix mode: {force_fix_mode}\n")
            self.validation_log.insert(tk.END, "\nStep 1: Preparing connections...\n")
            self.validation_log.see(tk.END)
            
            self.validation_log.insert(tk.END, f"  [OK] Preparing connections\n")
            self.validation_log.insert(tk.END, f"    Source: {src_server} | {src_db} | Auth: {src_auth}\n")
            self.validation_log.insert(tk.END, f"    Destination: {dest_server} | {dest_db} | Auth: {dest_auth}\n")
            self.validation_log.see(tk.END)
            
            src_conn = None
            dest_conn = None
            
            self.validation_log.insert(tk.END, "\nStep 2: Connecting to source database...\n")
            self.validation_log.see(tk.END)
            try:
                src_conn = connect_to_any_database(
                    server=src_server,
                    database=src_db,
                    auth=src_auth,
                    user=src_user,
                    password=src_password or None,
                    db_type=self.src_db_type_var.get(),
                    port=int(self.src_port_var.get() or 50000),
                    timeout=30
                )
                self.validation_log.insert(tk.END, "  [OK] Connected to source database successfully\n")
                self.validation_log.insert(tk.END, f"    Server: {src_server}\n")
                self.validation_log.insert(tk.END, f"    Database: {src_db}\n")
                self.validation_log.see(tk.END)
            except Exception as e:
                self.validation_log.insert(tk.END, f"  [X] Error connecting to source: {str(e)}\n")
                self.validation_log.insert(tk.END, f"    Source: {src_server} | {src_db}\n")
                self.validation_log.insert(tk.END, f"    Auth: {src_auth}\n")
                self.validation_log.insert(tk.END, f"    User: {src_user}\n")
                import traceback
                self.validation_log.insert(tk.END, f"    Traceback: {traceback.format_exc()}\n")
                self.validation_log.see(tk.END)
                if 'src_conn' in locals() and src_conn:
                    try:
                        src_conn.close()
                    except:
                        pass
                messagebox.showerror("Connection Error", f"Failed to connect to source database:\n{str(e)}\n\nSource: {src_server} | {src_db}")
                return
            
            self.validation_log.insert(tk.END, "\nStep 3: Connecting to destination database...\n")
            self.validation_log.see(tk.END)
            try:
                dest_conn = connect_to_any_database(
                    server=dest_server,
                    database=dest_db,
                    auth=dest_auth,
                    user=dest_user,
                    password=dest_password or None,
                    db_type=self.dest_db_type_var.get(),
                    port=int(self.dest_port_var.get() or 50000),
                    timeout=30
                )
                self.validation_log.insert(tk.END, "  [OK] Connected to destination database successfully\n")
                self.validation_log.insert(tk.END, f"    Server: {dest_server}\n")
                self.validation_log.insert(tk.END, f"    Database: {dest_db}\n")
                self.validation_log.see(tk.END)
            except Exception as e:
                if src_conn:
                    try:
                        src_conn.close()
                    except:
                        pass
                self.validation_log.insert(tk.END, f"  [FAILED] Error connecting to destination: {str(e)}\n")
                self.validation_log.insert(tk.END, f"    Destination: {dest_server} | {dest_db}\n")
                self.validation_log.insert(tk.END, f"    Auth: {dest_auth}\n")
                self.validation_log.insert(tk.END, f"    User: {dest_user}\n")
                import traceback
                self.validation_log.insert(tk.END, f"    Traceback: {traceback.format_exc()}\n")
                self.validation_log.see(tk.END)
                messagebox.showerror("Connection Error", f"Failed to connect to destination database:\n{str(e)}\n\nDestination: {dest_server} | {dest_db}")
                return
            
            # Process objects with cursors
            try:
                self.validation_log.insert(tk.END, "\nStep 4: Creating cursors...\n")
                self.validation_log.see(tk.END)
                src_cur = src_conn.cursor()
                dest_cur = dest_conn.cursor()
                self.validation_log.insert(tk.END, "  [OK] Cursors created\n")
                self.validation_log.see(tk.END)
                
                success_count = 0
                error_count = 0
                
                # Import functions from sechma_backup
                self.validation_log.insert(tk.END, "\nStep 5: Importing schema backup functions...\n")
                self.validation_log.see(tk.END)
                import sys
                from pathlib import Path
                parent_dir = Path(__file__).parent.parent.parent.parent
                sys.path.insert(0, str(parent_dir))
                
                try:
                    try:
                        from src.backup.exporters import (
                            fetch_tables, fetch_columns, fetch_primary_key,
                            build_create_table_sql, qident, type_sql, parse_int_or_default,
                            fetch_objects, object_definition, wrap_create_or_alter,
                            export_indexes, export_foreign_keys, export_check_constraints
                        )
                    except ImportError:
                        try:
                            from schema_backup import (
                                fetch_tables, fetch_columns, fetch_primary_key,
                                build_create_table_sql, qident, type_sql, parse_int_or_default,
                                fetch_objects, object_definition, wrap_create_or_alter,
                                export_indexes, export_foreign_keys, export_check_constraints
                            )
                        except ImportError:
                            from sechma_backup import (
                                fetch_tables, fetch_columns, fetch_primary_key,
                                build_create_table_sql, qident, type_sql, parse_int_or_default,
                                fetch_objects, object_definition, wrap_create_or_alter,
                                export_indexes, export_foreign_keys, export_check_constraints
                            )
                    self.validation_log.insert(tk.END, "  [OK] Schema backup functions imported successfully\n")
                    self.validation_log.see(tk.END)
                except ImportError as import_err:
                    self.validation_log.insert(tk.END, f"  [X] Error: Could not import schema backup functions\n")
                    self.validation_log.insert(tk.END, f"    Error: {str(import_err)}\n")
                    import traceback
                    self.validation_log.insert(tk.END, f"    Traceback: {traceback.format_exc()}\n")
                    self.validation_log.see(tk.END)
                    messagebox.showerror("Error", f"Could not import required functions from schema_backup.py or sechma_backup.py:\n{str(import_err)}")
                    return
                
                self.validation_log.insert(tk.END, f"\nStep 6: Processing {len(missing_objects)} object(s)...\n")
                self.validation_log.insert(tk.END, "="*60 + "\n")
                self.validation_log.see(tk.END)
                
                obj_count = 0
                for obj_name, source_status in missing_objects:
                    obj_count += 1
                    self.validation_log.insert(tk.END, f"\n[{obj_count}/{len(missing_objects)}] Processing: {obj_name}\n")
                    self.validation_log.insert(tk.END, f"  Status: {source_status}\n")
                    self.validation_log.see(tk.END)
                    try:
                        # Parse object name (format: "ObjectType.Schema.Object" or "Schema.Table" or "Schema.Table.Column")
                        parts = obj_name.split('.')
                        
                        if len(parts) == 2:
                            # Table: Schema.Table
                            schema, table = parts
                            self.validation_log.insert(tk.END, f"  Object type: Table\n")
                            self.validation_log.insert(tk.END, f"  Fetching table definition from source...\n")
                            self.validation_log.see(tk.END)
                            
                            # Get table definition from source
                            try:
                                cols = fetch_columns(src_cur, schema, table)
                                self.validation_log.insert(tk.END, f"    [OK] Found {len(cols)} column(s)\n")
                                pk_rows = fetch_primary_key(src_cur, schema, table)
                                self.validation_log.insert(tk.END, f"    [OK] Found {len(pk_rows)} primary key column(s)\n")
                                create_sql = build_create_table_sql(schema, table, cols, pk_rows)
                                self.validation_log.insert(tk.END, f"    [OK] Generated CREATE TABLE SQL\n")
                                self.validation_log.see(tk.END)
                            except Exception as fetch_err:
                                self.validation_log.insert(tk.END, f"    [X] Error fetching table definition: {str(fetch_err)}\n")
                                error_count += 1
                                continue
                            
                            # Execute on destination
                            self.validation_log.insert(tk.END, f"  Executing CREATE TABLE on destination...\n")
                            self.validation_log.see(tk.END)
                            try:
                                batch_count = 0
                                for batch in create_sql.split('GO'):
                                    batch = batch.strip()
                                    if batch:
                                        batch_count += 1
                                        self.validation_log.insert(tk.END, f"    Executing batch {batch_count}...\n")
                                        dest_cur.execute(batch)
                                dest_conn.commit()
                                self.validation_log.insert(tk.END, f"  [OK] Created table: {schema}.{table}\n")
                                success_count += 1
                            except Exception as exec_err:
                                self.validation_log.insert(tk.END, f"  [X] Error executing CREATE TABLE: {str(exec_err)}\n")
                                self.validation_log.insert(tk.END, f"    SQL: {create_sql[:500]}...\n")
                                error_count += 1
                                dest_conn.rollback()
                            self.validation_log.see(tk.END)
                            
                        elif len(parts) == 3:
                            # Could be: ObjectType.Schema.Object, Schema.Table.Column, Schema.Table.Index, or Schema.Table.Constraint
                            part1, part2, part3 = parts
                            
                            # Check if it's an index (various prefixes) or foreign key (starts with FK_)
                            is_index = (
                                part3.startswith('IX_') or 
                                part3.startswith('PK_') or 
                                part3.startswith('UQ_') or
                                part3.startswith('IDX_') or
                                part3.startswith('NCIX_') or
                                part3.startswith('pgndx_') or
                                part3.startswith('idx_')
                            )
                            
                            if is_index:
                                # It's an index: Schema.Table.IndexName
                                schema, table, index_name = parts
                                self.validation_log.insert(tk.END, f"  Object type: Index\n")
                                self.validation_log.insert(tk.END, f"  Fetching index definition from source...\n")
                                self.validation_log.see(tk.END)
                                
                                # Get index definition from source with all properties
                                try:
                                    src_cur.execute("""
                                    SELECT 
                                        i.name AS index_name,
                                        i.is_unique,
                                        i.type_desc,
                                        i.has_filter,
                                        CAST(i.filter_definition AS NVARCHAR(MAX)) AS filter_definition,
                                        i.ignore_dup_key,
                                        i.allow_row_locks,
                                        i.allow_page_locks,
                                        i.fill_factor,
                                        i.is_padded,
                                        ds.name AS filegroup_name
                                    FROM sys.indexes i
                                    JOIN sys.tables t ON t.object_id = i.object_id
                                    JOIN sys.schemas s ON s.schema_id = t.schema_id
                                    LEFT JOIN sys.data_spaces ds ON ds.data_space_id = i.data_space_id
                                    WHERE s.name = ? AND t.name = ? AND i.name = ?
                                      AND i.is_primary_key = 0
                                      AND i.is_unique_constraint = 0
                                """, schema, table, index_name)
                                    
                                    idx_row = src_cur.fetchone()
                                    if not idx_row:
                                        self.validation_log.insert(tk.END, f"    [X] Index not found in source\n")
                                        error_count += 1
                                        continue
                                    
                                    self.validation_log.insert(tk.END, f"    [OK] Found index definition\n")
                                    self.validation_log.insert(tk.END, f"      Type: {idx_row.type_desc}\n")
                                    self.validation_log.insert(tk.END, f"      Unique: {idx_row.is_unique}\n")
                                    self.validation_log.see(tk.END)
                                except Exception as fetch_err:
                                    self.validation_log.insert(tk.END, f"    [X] Error fetching index definition: {str(fetch_err)}\n")
                                    error_count += 1
                                    continue
                                
                                if idx_row:
                                    # Get index columns
                                    self.validation_log.insert(tk.END, f"    Fetching index columns...\n")
                                    self.validation_log.see(tk.END)
                                    try:
                                        src_cur.execute("""
                                            SELECT
                                                c.name AS col_name,
                                                ic.is_descending_key,
                                                ic.is_included_column,
                                                ic.key_ordinal
                                            FROM sys.index_columns ic
                                            JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                            WHERE ic.object_id = OBJECT_ID(?, 'U')
                                              AND ic.index_id = (SELECT index_id FROM sys.indexes WHERE object_id = OBJECT_ID(?, 'U') AND name = ?)
                                            ORDER BY ic.is_included_column, ic.key_ordinal;
                                        """, f"{schema}.{table}", f"{schema}.{table}", index_name)
                                        
                                        idx_cols = src_cur.fetchall()
                                        self.validation_log.insert(tk.END, f"      [OK] Found {len(idx_cols)} column(s)\n")
                                        key_cols = []
                                        inc_cols = []
                                        for c in idx_cols:
                                            if c.is_included_column:
                                                inc_cols.append(qident(c.col_name))
                                            else:
                                                direction = " DESC" if c.is_descending_key else " ASC"
                                                key_cols.append(qident(c.col_name) + direction)
                                        
                                        self.validation_log.insert(tk.END, f"      Key columns: {len(key_cols)}, Include columns: {len(inc_cols)}\n")
                                        self.validation_log.see(tk.END)
                                        
                                        if not key_cols:
                                            self.validation_log.insert(tk.END, f"    [X] Index has no key columns: {obj_name}\n")
                                            error_count += 1
                                            continue
                                    except Exception as col_err:
                                        self.validation_log.insert(tk.END, f"    [X] Error fetching index columns: {str(col_err)}\n")
                                        error_count += 1
                                        continue
                                    
                                    table_name = f"{qident(schema)}.{qident(table)}"
                                    unique = "UNIQUE " if idx_row.is_unique else ""
                                    
                                    # Get the actual type_desc value and log it for debugging
                                    type_desc_value = (idx_row.type_desc or "").strip().upper()
                                    source_typ = "CLUSTERED" if type_desc_value == "CLUSTERED" else "NONCLUSTERED"
                                    typ = source_typ  # May be changed below if conflicts detected
                                    
                                    # Check if index already exists in destination FIRST (before clustered check)
                                    self.validation_log.insert(tk.END, f"    Checking if index exists in destination...\n")
                                    self.validation_log.see(tk.END)
                                    try:
                                        dest_cur.execute("""
                                            SELECT i.type_desc, COUNT(*) as cnt
                                            FROM sys.indexes i
                                            JOIN sys.tables t ON t.object_id = i.object_id
                                            JOIN sys.schemas s ON s.schema_id = t.schema_id
                                            WHERE s.name = ? AND t.name = ? AND i.name = ?
                                              AND i.is_primary_key = 0
                                              AND i.is_unique_constraint = 0
                                            GROUP BY i.type_desc
                                        """, schema, table, index_name)
                                        
                                        dest_index_row = dest_cur.fetchone()
                                        if dest_index_row and dest_index_row.cnt > 0:
                                            self.validation_log.insert(tk.END, f"    [OK] Index already exists in destination (type: {dest_index_row.type_desc})\n")
                                            self.validation_log.insert(tk.END, f"  [OK] Skipped (already exists): {obj_name}\n")
                                            success_count += 1
                                            continue
                                        self.validation_log.insert(tk.END, f"    [OK] Index does not exist in destination, proceeding...\n")
                                        self.validation_log.see(tk.END)
                                    except Exception as check_err:
                                        self.validation_log.insert(tk.END, f"    [WARN] Error checking destination: {str(check_err)}, proceeding anyway...\n")
                                        self.validation_log.see(tk.END)
                                    
                                    # For exact replication: keep CLUSTERED; do not convert to NONCLUSTERED.
                                    # If destination has a clustered PK, user must fix PK first (Fix Missing Objects, deploy PK then this index).
                                    if typ == "CLUSTERED":
                                        self.validation_log.insert(tk.END, f"    Checking for clustered index conflicts...\n")
                                        self.validation_log.see(tk.END)
                                        try:
                                            dest_cur.execute("""
                                                SELECT i.name, i.is_primary_key, i.is_unique_constraint 
                                                FROM sys.indexes i
                                                JOIN sys.tables t ON t.object_id = i.object_id
                                                JOIN sys.schemas s ON s.schema_id = t.schema_id
                                                WHERE s.name = ? AND t.name = ? AND i.type_desc = 'CLUSTERED' AND i.name != ?
                                            """, schema, table, index_name)
                                            existing_clustered = dest_cur.fetchone()
                                            if existing_clustered:
                                                existing_name = existing_clustered.name
                                                self.validation_log.insert(tk.END, f"    [WARN] Table already has clustered index '{existing_name}' (e.g. primary key). For exact replication: use Fix Missing Objects and deploy primary key first (recreate as NONCLUSTERED), then deploy this index as CLUSTERED.\n")
                                                self.validation_log.see(tk.END)
                                                # Keep typ = CLUSTERED so we generate correct script; deploy may fail until PK is fixed
                                            else:
                                                self.validation_log.insert(tk.END, f"    [OK] No clustered index conflict\n")
                                                self.validation_log.see(tk.END)
                                        except Exception as conflict_err:
                                            self.validation_log.insert(tk.END, f"    [WARN] Error checking conflicts: {str(conflict_err)}, proceeding...\n")
                                            self.validation_log.see(tk.END)
                                    
                                    # SQL Server doesn't allow INCLUDE columns on clustered indexes
                                    # If this is a clustered index with INCLUDE columns, convert to NONCLUSTERED
                                    if typ == "CLUSTERED" and inc_cols:
                                        typ = "NONCLUSTERED"
                                        self.validation_log.insert(tk.END, f"    [WARN] Clustered indexes cannot have INCLUDE columns, converting to NONCLUSTERED\n")
                                        self.validation_log.see(tk.END)
                                    
                                    # Debug: Log the index type from source and final type
                                    if source_typ != typ:
                                        self.validation_log.insert(tk.END, f"    Index type changed: {source_typ} -> {typ}\n")
                                    else:
                                        self.validation_log.insert(tk.END, f"    Index type: {typ}\n")
                                    self.validation_log.see(tk.END)
                                    
                                    self.validation_log.insert(tk.END, f"    Building CREATE INDEX SQL...\n")
                                    self.validation_log.see(tk.END)
                                    
                                    # Format column list with proper indentation (matching user's example format)
                                    key_cols_formatted = ',\n\t'.join(key_cols)
                                    create_sql = f"CREATE {unique}{typ} INDEX {qident(index_name)} ON {table_name}\n(\n\t{key_cols_formatted}\n)"
                                    if inc_cols:
                                        inc_cols_formatted = ',\n\t'.join(inc_cols)
                                        create_sql += f"\nINCLUDE (\n\t{inc_cols_formatted}\n)"
                                    
                                    # Add WHERE clause for filtered indexes
                                    if idx_row.has_filter and idx_row.filter_definition:
                                        filter_def = str(idx_row.filter_definition).strip()
                                        if filter_def:
                                            if not filter_def.upper().startswith('WHERE'):
                                                create_sql += f"\nWHERE {filter_def}"
                                            else:
                                                create_sql += f"\n{filter_def}"
                                    
                                    # Build WITH clause options (matching user's example format)
                                    with_options = []
                                    
                                    # STATISTICS_NORECOMPUTE (default to OFF)
                                    with_options.append("STATISTICS_NORECOMPUTE = OFF")
                                    
                                    # IGNORE_DUP_KEY (only for unique indexes)
                                    if idx_row.is_unique and idx_row.ignore_dup_key is not None:
                                        with_options.append(f"IGNORE_DUP_KEY = {'ON' if idx_row.ignore_dup_key else 'OFF'}")
                                    
                                    # DROP_EXISTING (always OFF for new indexes)
                                    with_options.append("DROP_EXISTING = OFF")
                                    
                                    # ONLINE (always OFF for compatibility)
                                    with_options.append("ONLINE = OFF")
                                    
                                    # ALLOW_ROW_LOCKS
                                    if idx_row.allow_row_locks is not None:
                                        with_options.append(f"ALLOW_ROW_LOCKS = {'ON' if idx_row.allow_row_locks else 'OFF'}")
                                    
                                    # ALLOW_PAGE_LOCKS
                                    if idx_row.allow_page_locks is not None:
                                        with_options.append(f"ALLOW_PAGE_LOCKS = {'ON' if idx_row.allow_page_locks else 'OFF'}")
                                    
                                    # FILLFACTOR (only if specified)
                                    if idx_row.fill_factor is not None and idx_row.fill_factor > 0 and idx_row.fill_factor < 100:
                                        with_options.append(f"FILLFACTOR = {idx_row.fill_factor}")
                                    
                                    # PAD_INDEX (only if FILLFACTOR is specified)
                                    if idx_row.fill_factor is not None and idx_row.fill_factor > 0 and idx_row.fill_factor < 100:
                                        if idx_row.is_padded is not None:
                                            with_options.append(f"PAD_INDEX = {'ON' if idx_row.is_padded else 'OFF'}")
                                    
                                    # Add WITH clause if we have options (matching user's example format)
                                    if with_options:
                                        options_str = ', \n    '.join(with_options)
                                        create_sql += f"\nWITH (\n    {options_str}\n)"
                                    
                                    # Add filegroup specification
                                    filegroup_name = getattr(idx_row, 'filegroup_name', None) if idx_row else None
                                    if filegroup_name and str(filegroup_name).upper() != 'PRIMARY':
                                        create_sql += f" ON [{filegroup_name}]"
                                    else:
                                        create_sql += " ON [PRIMARY]"
                                    
                                    create_sql += ";"
                                    
                                    # Log the generated SQL for debugging
                                    self.validation_log.insert(tk.END, f"    [OK] SQL generated (length: {len(create_sql)} chars)\n")
                                    self.validation_log.insert(tk.END, f"    Generated SQL:\n{create_sql}\n")
                                    self.validation_log.see(tk.END)
                                    
                                    self.validation_log.insert(tk.END, f"    Executing CREATE INDEX on destination...\n")
                                    self.validation_log.see(tk.END)
                                    try:
                                        dest_cur.execute(create_sql)
                                        dest_conn.commit()
                                        
                                        self.validation_log.insert(tk.END, f"    [OK] CREATE INDEX executed successfully\n")
                                        self.validation_log.insert(tk.END, f"  [OK] Created index: {schema}.{table}.{index_name}\n")
                                        success_count += 1
                                    except Exception as idx_err:
                                        # Check if index already exists
                                        if "already exists" in str(idx_err).lower() or "duplicate" in str(idx_err).lower():
                                            self.validation_log.insert(tk.END, f"    [WARN] Index already exists (detected during execution)\n")
                                            self.validation_log.insert(tk.END, f"  [OK] Index already exists in destination: {obj_name}\n")
                                            success_count += 1  # Count as success
                                        else:
                                            self.validation_log.insert(tk.END, f"    [X] Error executing CREATE INDEX: {str(idx_err)}\n")
                                            self.validation_log.insert(tk.END, f"    SQL that failed:\n{create_sql}\n")
                                            import traceback
                                            self.validation_log.insert(tk.END, f"    Traceback: {traceback.format_exc()}\n")
                                            self.validation_log.insert(tk.END, f"  [X] Error creating index {obj_name}: {str(idx_err)}\n")
                                            error_count += 1
                                            dest_conn.rollback()
                                    self.validation_log.see(tk.END)
                                elif force_fix_mode:
                                    # Try to get from export_indexes
                                    self.validation_log.insert(tk.END, f"[WARN] Index not found, skipping: {obj_name}\n")
                                    error_count += 1
                                else:
                                    self.validation_log.insert(tk.END, f"[X] Index not found in source: {obj_name}\n")
                                    error_count += 1
                                    
                            elif part3.startswith('FK_'):
                                # It's a foreign key: Schema.Table.FKName
                                schema, table, fk_name = parts
                                self.validation_log.insert(tk.END, f"  Object type: Foreign Key\n")
                                self.validation_log.insert(tk.END, f"  Fetching foreign key definition from source...\n")
                                self.validation_log.see(tk.END)
                                
                                # Get FK definition from source
                                try:
                                    src_cur.execute("""
                                    SELECT
                                        fk.name AS fk_name,
                                        s2.name AS ref_schema_name,
                                        t2.name AS ref_table_name,
                                        fk.delete_referential_action_desc,
                                        fk.update_referential_action_desc
                                    FROM sys.foreign_keys fk
                                    JOIN sys.tables t1 ON t1.object_id = fk.parent_object_id
                                    JOIN sys.schemas s1 ON s1.schema_id = t1.schema_id
                                    JOIN sys.tables t2 ON t2.object_id = fk.referenced_object_id
                                    JOIN sys.schemas s2 ON s2.schema_id = t2.schema_id
                                    WHERE s1.name = ? AND t1.name = ? AND fk.name = ?
                                """, schema, table, fk_name)
                                    
                                    fk_row = src_cur.fetchone()
                                    if not fk_row:
                                        self.validation_log.insert(tk.END, f"    [X] Foreign key not found in source\n")
                                        error_count += 1
                                        continue
                                    
                                    self.validation_log.insert(tk.END, f"    [OK] Found foreign key definition\n")
                                    self.validation_log.insert(tk.END, f"      References: {fk_row.ref_schema_name}.{fk_row.ref_table_name}\n")
                                    self.validation_log.see(tk.END)
                                except Exception as fetch_err:
                                    self.validation_log.insert(tk.END, f"    [X] Error fetching foreign key definition: {str(fetch_err)}\n")
                                    error_count += 1
                                    continue
                                
                                if fk_row:
                                    # Get FK columns
                                    self.validation_log.insert(tk.END, f"    Fetching foreign key columns...\n")
                                    self.validation_log.see(tk.END)
                                    try:
                                        src_cur.execute("""
                                            SELECT
                                                pc.name AS parent_column,
                                                rc.name AS ref_column,
                                                fkc.constraint_column_id
                                            FROM sys.foreign_key_columns fkc
                                            JOIN sys.columns pc
                                                ON pc.object_id = fkc.parent_object_id AND pc.column_id = fkc.parent_column_id
                                            JOIN sys.columns rc
                                                ON rc.object_id = fkc.referenced_object_id AND rc.column_id = fkc.referenced_column_id
                                            WHERE fkc.constraint_object_id = OBJECT_ID(?, 'F')
                                            ORDER BY fkc.constraint_column_id;
                                        """, fk_name)
                                        
                                        fk_cols = src_cur.fetchall()
                                        self.validation_log.insert(tk.END, f"      [OK] Found {len(fk_cols)} column pair(s)\n")
                                        self.validation_log.see(tk.END)
                                        
                                        if not fk_cols:
                                            self.validation_log.insert(tk.END, f"    [X] Foreign key has no columns: {obj_name}\n")
                                            error_count += 1
                                            continue
                                    except Exception as col_err:
                                        self.validation_log.insert(tk.END, f"    [X] Error fetching foreign key columns: {str(col_err)}\n")
                                        error_count += 1
                                        continue
                                    
                                    parent_cols = ", ".join(qident(c.parent_column) for c in fk_cols)
                                    ref_cols = ", ".join(qident(c.ref_column) for c in fk_cols)
                                    
                                    parent = f"{qident(schema)}.{qident(table)}"
                                    ref = f"{qident(fk_row.ref_schema_name)}.{qident(fk_row.ref_table_name)}"
                                    
                                    actions = []
                                    if fk_row.delete_referential_action_desc and fk_row.delete_referential_action_desc.upper() != "NO_ACTION":
                                        actions.append(f"ON DELETE {fk_row.delete_referential_action_desc.replace('_', ' ')}")
                                    if fk_row.update_referential_action_desc and fk_row.update_referential_action_desc.upper() != "NO_ACTION":
                                        actions.append(f"ON UPDATE {fk_row.update_referential_action_desc.replace('_', ' ')}")
                                    
                                    create_sql = f"ALTER TABLE {parent} WITH CHECK ADD CONSTRAINT {qident(fk_name)} FOREIGN KEY ({parent_cols}) REFERENCES {ref} ({ref_cols})"
                                    if actions:
                                        create_sql += " " + " ".join(actions)
                                    create_sql += ";"
                                    
                                    try:
                                        dest_cur.execute(create_sql)
                                        dest_conn.commit()
                                        
                                        self.validation_log.insert(tk.END, f"[OK] Created foreign key: {schema}.{table}.{fk_name}\n")
                                        success_count += 1
                                    except Exception as fk_err:
                                        # Check if FK already exists
                                        if "already exists" in str(fk_err).lower() or "duplicate" in str(fk_err).lower():
                                            self.validation_log.insert(tk.END, f"[WARN] Foreign key already exists: {obj_name}\n")
                                            error_count += 1
                                        else:
                                            raise
                                elif force_fix_mode:
                                    self.validation_log.insert(tk.END, f"[WARN] Foreign key not found, skipping: {obj_name}\n")
                                    error_count += 1
                                else:
                                    self.validation_log.insert(tk.END, f"[X] Foreign key not found in source: {obj_name}\n")
                                    error_count += 1
                                    
                            else:
                                # Check if part3 looks like an index, constraint, or default (not a column)
                                # Common prefixes: IX_, PK_, UQ_, FK_, IDX_, NCIX_, pgndx_, UX_, UC_, DEF_, DF__
                                is_likely_index_or_constraint = (
                                    part3.startswith('IX_') or 
                                    part3.startswith('PK_') or 
                                    part3.startswith('UQ_') or 
                                    part3.startswith('FK_') or
                                    part3.startswith('IDX_') or
                                    part3.startswith('NCIX_') or
                                    part3.startswith('pgndx_') or
                                    part3.startswith('UX_') or
                                    part3.startswith('UC_') or
                                    part3.startswith('DEF_') or
                                    part3.startswith('DF__')
                                )
                                
                                if is_likely_index_or_constraint:
                                    # This is likely an index, constraint, or default, not a column
                                    # Try to find it in sys.indexes, sys.foreign_keys, sys.check_constraints, or sys.default_constraints
                                    schema, table, obj_name_only = parts
                                    
                                    # Try as index first - recursively call index creation logic
                                    src_cur.execute("""
                                        SELECT COUNT(*) FROM sys.indexes i
                                        JOIN sys.tables t ON t.object_id = i.object_id
                                        JOIN sys.schemas s ON s.schema_id = t.schema_id
                                        WHERE s.name = ? AND t.name = ? AND i.name = ?
                                          AND i.is_primary_key = 0
                                          AND i.is_unique_constraint = 0
                                    """, schema, table, obj_name_only)
                                    if src_cur.fetchone()[0] > 0:
                                        # It's an index - handle it as an index
                                        # Re-enter the index creation logic by treating it as if it started with IX_
                                        # We'll use the same logic but with the correct index name
                                        self.validation_log.insert(tk.END, f"Creating index: {schema}.{table}.{obj_name_only}...\n")
                                        self.validation_log.see(tk.END)
                                        
                                        # Get index definition (same query as above)
                                        src_cur.execute("""
                                            SELECT 
                                                i.name AS index_name,
                                                i.is_unique,
                                                i.type_desc,
                                                i.has_filter,
                                                CAST(i.filter_definition AS NVARCHAR(MAX)) AS filter_definition,
                                                i.ignore_dup_key,
                                                i.allow_row_locks,
                                                i.allow_page_locks,
                                                i.fill_factor,
                                                i.is_padded,
                                                ds.name AS filegroup_name
                                            FROM sys.indexes i
                                            JOIN sys.tables t ON t.object_id = i.object_id
                                            JOIN sys.schemas s ON s.schema_id = t.schema_id
                                            LEFT JOIN sys.data_spaces ds ON ds.data_space_id = i.data_space_id
                                            WHERE s.name = ? AND t.name = ? AND i.name = ?
                                              AND i.is_primary_key = 0
                                              AND i.is_unique_constraint = 0
                                        """, schema, table, obj_name_only)
                                        
                                        idx_row = src_cur.fetchone()
                                        if idx_row:
                                            # Get index columns
                                            src_cur.execute("""
                                                SELECT
                                                    c.name AS col_name,
                                                    ic.is_descending_key,
                                                    ic.is_included_column,
                                                    ic.key_ordinal
                                                FROM sys.index_columns ic
                                                JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                                WHERE ic.object_id = OBJECT_ID(?, 'U')
                                                  AND ic.index_id = (SELECT index_id FROM sys.indexes WHERE object_id = OBJECT_ID(?, 'U') AND name = ?)
                                                ORDER BY ic.is_included_column, ic.key_ordinal;
                                            """, f"{schema}.{table}", f"{schema}.{table}", obj_name_only)
                                            
                                            idx_cols = src_cur.fetchall()
                                            key_cols = []
                                            inc_cols = []
                                            for c in idx_cols:
                                                if c.is_included_column:
                                                    inc_cols.append(qident(c.col_name))
                                                else:
                                                    direction = " DESC" if c.is_descending_key else " ASC"
                                                    key_cols.append(qident(c.col_name) + direction)
                                            
                                            if not key_cols:
                                                self.validation_log.insert(tk.END, f"[X] Index has no key columns: {obj_name}\n")
                                                error_count += 1
                                                continue
                                            
                                            table_name = f"{qident(schema)}.{qident(table)}"
                                            unique = "UNIQUE " if idx_row.is_unique else ""
                                            
                                            # Get the actual type_desc value and log it for debugging
                                            type_desc_value = (idx_row.type_desc or "").strip().upper()
                                            typ = "CLUSTERED" if type_desc_value == "CLUSTERED" else "NONCLUSTERED"
                                            
                                            # Debug: Log the index type from source
                                            self.validation_log.insert(tk.END, f"  Source index type: {idx_row.type_desc} -> {typ}\n")
                                            self.validation_log.see(tk.END)
                                            
                                            # Check if index already exists in destination FIRST
                                            dest_cur.execute("""
                                                SELECT i.type_desc, COUNT(*) as cnt
                                                FROM sys.indexes i
                                                JOIN sys.tables t ON t.object_id = i.object_id
                                                JOIN sys.schemas s ON s.schema_id = t.schema_id
                                                WHERE s.name = ? AND t.name = ? AND i.name = ?
                                                  AND i.is_primary_key = 0
                                                  AND i.is_unique_constraint = 0
                                                GROUP BY i.type_desc
                                            """, schema, table, obj_name_only)
                                            
                                            dest_index_row = dest_cur.fetchone()
                                            if dest_index_row and dest_index_row.cnt > 0:
                                                self.validation_log.insert(tk.END, f"[OK] Index already exists in destination (type: {dest_index_row.type_desc}): {obj_name}\n")
                                                success_count += 1
                                                continue
                                            
                                            # If this index is CLUSTERED and destination has another clustered (e.g. PK), skip with message for exact replication
                                            if typ == "CLUSTERED":
                                                dest_cur.execute("""
                                                    SELECT i.name FROM sys.indexes i
                                                    JOIN sys.tables t ON t.object_id = i.object_id
                                                    JOIN sys.schemas s ON s.schema_id = t.schema_id
                                                    WHERE s.name = ? AND t.name = ? AND i.type_desc = 'CLUSTERED' AND i.name != ?
                                                """, schema, table, obj_name_only)
                                                existing = dest_cur.fetchone()
                                                if existing:
                                                    self.validation_log.insert(tk.END, f"[WARN] Skipping: table already has clustered index '{existing.name}'. For exact replication use Fix Missing Objects: deploy primary key first (as NONCLUSTERED), then this index: {obj_name}\n")
                                                    error_count += 1
                                                    continue
                                            # For nonclustered indexes (or no conflict), proceed with creation
                                            
                                            # Format column list with proper indentation (matching user's example format)
                                            key_cols_formatted = ',\n\t'.join(key_cols)
                                            create_sql = f"CREATE {unique}{typ} INDEX {qident(obj_name_only)} ON {table_name}\n(\n\t{key_cols_formatted}\n)"
                                            if inc_cols:
                                                inc_cols_formatted = ',\n\t'.join(inc_cols)
                                                create_sql += f"\nINCLUDE (\n\t{inc_cols_formatted}\n)"
                                            
                                            # Add WHERE clause for filtered indexes
                                            if idx_row.has_filter and idx_row.filter_definition:
                                                filter_def = str(idx_row.filter_definition).strip()
                                                if filter_def:
                                                    if not filter_def.upper().startswith('WHERE'):
                                                        create_sql += f"\nWHERE {filter_def}"
                                                    else:
                                                        create_sql += f"\n{filter_def}"
                                            
                                            # Build WITH clause options (matching user's example format)
                                            with_options = []
                                            with_options.append("STATISTICS_NORECOMPUTE = OFF")
                                            
                                            if idx_row.is_unique and idx_row.ignore_dup_key is not None:
                                                with_options.append(f"IGNORE_DUP_KEY = {'ON' if idx_row.ignore_dup_key else 'OFF'}")
                                            
                                            with_options.append("DROP_EXISTING = OFF")
                                            with_options.append("ONLINE = OFF")
                                            
                                            if idx_row.allow_row_locks is not None:
                                                with_options.append(f"ALLOW_ROW_LOCKS = {'ON' if idx_row.allow_row_locks else 'OFF'}")
                                            
                                            if idx_row.allow_page_locks is not None:
                                                with_options.append(f"ALLOW_PAGE_LOCKS = {'ON' if idx_row.allow_page_locks else 'OFF'}")
                                            
                                            if idx_row.fill_factor is not None and idx_row.fill_factor > 0 and idx_row.fill_factor < 100:
                                                with_options.append(f"FILLFACTOR = {idx_row.fill_factor}")
                                            
                                            if idx_row.fill_factor is not None and idx_row.fill_factor > 0 and idx_row.fill_factor < 100:
                                                if idx_row.is_padded is not None:
                                                    with_options.append(f"PAD_INDEX = {'ON' if idx_row.is_padded else 'OFF'}")
                                            
                                            if with_options:
                                                options_str = ', \n    '.join(with_options)
                                                create_sql += f"\nWITH (\n    {options_str}\n)"
                                            
                                            filegroup_name = getattr(idx_row, 'filegroup_name', None) if idx_row else None
                                            if filegroup_name and str(filegroup_name).upper() != 'PRIMARY':
                                                create_sql += f" ON [{filegroup_name}]"
                                            else:
                                                create_sql += " ON [PRIMARY]"
                                            
                                            create_sql += ";"
                                            
                                            # Log the generated SQL for debugging
                                            self.validation_log.insert(tk.END, f"  Generated SQL:\n{create_sql}\n")
                                            self.validation_log.see(tk.END)
                                            
                                            try:
                                                dest_cur.execute(create_sql)
                                                dest_conn.commit()
                                                self.validation_log.insert(tk.END, f"[OK] Created index: {schema}.{table}.{obj_name_only}\n")
                                                success_count += 1
                                            except Exception as idx_err:
                                                if "already exists" in str(idx_err).lower() or "duplicate" in str(idx_err).lower():
                                                    self.validation_log.insert(tk.END, f"[OK] Index already exists in destination: {obj_name}\n")
                                                    success_count += 1
                                                else:
                                                    self.validation_log.insert(tk.END, f"[X] Error creating index {obj_name}: {str(idx_err)}\n")
                                                    self.validation_log.insert(tk.END, f"  SQL that failed:\n{create_sql}\n")
                                                    error_count += 1
                                                    dest_conn.rollback()
                                            continue  # Skip constraint checks since this was an index
                                    
                                    # Try as constraint (unique, check, default)
                                    # Check for default constraint
                                    src_cur.execute("""
                                        SELECT 
                                            dc.name AS constraint_name,
                                            c.name AS column_name,
                                            dc.definition
                                        FROM sys.default_constraints dc
                                        JOIN sys.columns c ON c.object_id = dc.parent_object_id AND c.column_id = dc.parent_column_id
                                        WHERE dc.parent_object_id = OBJECT_ID(?, 'U') AND dc.name = ?
                                    """, f"{schema}.{table}", obj_name_only)
                                    default_row = src_cur.fetchone()
                                    if default_row:
                                        # It's a default constraint - try to create it
                                        self.validation_log.insert(tk.END, f"Creating default constraint: {schema}.{table}.{obj_name_only}...\n")
                                        self.validation_log.see(tk.END)
                                        
                                        try:
                                            create_sql = f"ALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(default_row.constraint_name)} DEFAULT {default_row.definition} FOR {qident(default_row.column_name)}"
                                            dest_cur.execute(create_sql)
                                            dest_conn.commit()
                                            self.validation_log.insert(tk.END, f"[OK] Created default constraint: {schema}.{table}.{obj_name_only}\n")
                                            success_count += 1
                                        except Exception as dc_err:
                                            if "already exists" in str(dc_err).lower():
                                                self.validation_log.insert(tk.END, f"[OK] Default constraint already exists: {obj_name}\n")
                                                success_count += 1
                                            else:
                                                self.validation_log.insert(tk.END, f"[X] Error creating default constraint {obj_name}: {str(dc_err)}\n")
                                                error_count += 1
                                                dest_conn.rollback()
                                        continue
                                    
                                    # Check for check constraint
                                    src_cur.execute("""
                                        SELECT 
                                            cc.name AS constraint_name,
                                            cc.definition
                                        FROM sys.check_constraints cc
                                        WHERE cc.parent_object_id = OBJECT_ID(?, 'U') AND cc.name = ?
                                    """, f"{schema}.{table}", obj_name_only)
                                    check_row = src_cur.fetchone()
                                    if check_row:
                                        # It's a check constraint - try to create it
                                        self.validation_log.insert(tk.END, f"Creating check constraint: {schema}.{table}.{obj_name_only}...\n")
                                        self.validation_log.see(tk.END)
                                        
                                        try:
                                            create_sql = f"ALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(check_row.constraint_name)} CHECK {check_row.definition}"
                                            dest_cur.execute(create_sql)
                                            dest_conn.commit()
                                            self.validation_log.insert(tk.END, f"[OK] Created check constraint: {schema}.{table}.{obj_name_only}\n")
                                            success_count += 1
                                        except Exception as cc_err:
                                            if "already exists" in str(cc_err).lower():
                                                self.validation_log.insert(tk.END, f"[OK] Check constraint already exists: {obj_name}\n")
                                                success_count += 1
                                            else:
                                                self.validation_log.insert(tk.END, f"[X] Error creating check constraint {obj_name}: {str(cc_err)}\n")
                                                error_count += 1
                                                dest_conn.rollback()
                                        continue
                                    
                                    # Check for unique constraint
                                    src_cur.execute("""
                                        SELECT 
                                            kc.name AS constraint_name,
                                            kc.type_desc,
                                            kc.is_system_named
                                        FROM sys.key_constraints kc
                                        WHERE kc.parent_object_id = OBJECT_ID(?, 'U') AND kc.name = ? AND kc.type = 'UQ'
                                    """, f"{schema}.{table}", obj_name_only)
                                    unique_row = src_cur.fetchone()
                                    if unique_row:
                                        # It's a unique constraint - get columns and create it
                                        self.validation_log.insert(tk.END, f"Creating unique constraint: {schema}.{table}.{obj_name_only}...\n")
                                        self.validation_log.see(tk.END)
                                        
                                        # Get constraint columns
                                        src_cur.execute("""
                                            SELECT c.name AS column_name, ic.is_descending_key
                                            FROM sys.index_columns ic
                                            JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
                                            JOIN sys.key_constraints kc ON kc.unique_index_id = ic.index_id AND kc.parent_object_id = ic.object_id
                                            WHERE kc.parent_object_id = OBJECT_ID(?, 'U') AND kc.name = ?
                                            ORDER BY ic.key_ordinal
                                        """, f"{schema}.{table}", obj_name_only)
                                        uq_cols = src_cur.fetchall()
                                        
                                        if uq_cols:
                                            cols_list = ', '.join([f"{qident(c.column_name)} {'DESC' if c.is_descending_key else 'ASC'}" for c in uq_cols])
                                            try:
                                                create_sql = f"ALTER TABLE {qident(schema)}.{qident(table)} ADD CONSTRAINT {qident(unique_row.constraint_name)} UNIQUE ({cols_list})"
                                                dest_cur.execute(create_sql)
                                                dest_conn.commit()
                                                self.validation_log.insert(tk.END, f"[OK] Created unique constraint: {schema}.{table}.{obj_name_only}\n")
                                                success_count += 1
                                            except Exception as uq_err:
                                                if "already exists" in str(uq_err).lower():
                                                    self.validation_log.insert(tk.END, f"[OK] Unique constraint already exists: {obj_name}\n")
                                                    success_count += 1
                                                else:
                                                    self.validation_log.insert(tk.END, f"[X] Error creating unique constraint {obj_name}: {str(uq_err)}\n")
                                                    error_count += 1
                                                    dest_conn.rollback()
                                        else:
                                            self.validation_log.insert(tk.END, f"[X] Could not find columns for unique constraint: {obj_name}\n")
                                            error_count += 1
                                        continue
                                    
                                    # Check for primary key constraint (skip - can't add another PK)
                                    src_cur.execute("""
                                        SELECT name FROM sys.key_constraints 
                                        WHERE parent_object_id = OBJECT_ID(?, 'U') AND name = ? AND type = 'PK'
                                    """, f"{schema}.{table}", obj_name_only)
                                    if src_cur.fetchone():
                                        self.validation_log.insert(tk.END, f"[WARN] Primary key constraint {obj_name} - table should already have a primary key. Skipping.\n")
                                        error_count += 1
                                        continue
                                
                                # Check if it's a column (table exists in source)
                                src_cur.execute("""
                                    SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES 
                                    WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                                """, part1, part2)
                                if src_cur.fetchone()[0] > 0:
                                    # It's a column: Schema.Table.Column
                                    schema, table, column = parts
                                    self.validation_log.insert(tk.END, f"Adding column: {schema}.{table}.{column}...\n")
                                    self.validation_log.see(tk.END)
                                    
                                    # Get column definition from source
                                    cols = fetch_columns(src_cur, schema, table)
                                    col_def = None
                                    for col in cols:
                                        if col.column_name == column:
                                            col_def = col
                                            break
                                    
                                    if col_def:
                                        dest_cur.execute(
                                            """
                                            SELECT 1 FROM sys.columns c
                                            JOIN sys.tables t ON t.object_id = c.object_id
                                            JOIN sys.schemas s ON s.schema_id = t.schema_id
                                            WHERE s.name = ? AND t.name = ? AND c.name = ?
                                            """,
                                            (schema, table, column),
                                        )
                                        column_already_on_dest = dest_cur.fetchone() is not None

                                        if column_already_on_dest and source_status == "Mismatch":
                                            try:
                                                from gui.utils.schema_column_align import (
                                                    build_column_mismatch_fix_script,
                                                )
                                            except ImportError:
                                                try:
                                                    from azure_migration_tool.gui.utils.schema_column_align import (
                                                        build_column_mismatch_fix_script,
                                                    )
                                                except ImportError:
                                                    build_column_mismatch_fix_script = None
                                            if build_column_mismatch_fix_script is None:
                                                self.validation_log.insert(
                                                    tk.END,
                                                    "[X] Column align module not available.\n",
                                                )
                                                error_count += 1
                                                continue

                                            def _gc(cur, oname):
                                                c, _ = self._get_object_code(cur, oname)
                                                return c if c else None

                                            script, cerr = build_column_mismatch_fix_script(
                                                src_cur,
                                                dest_cur,
                                                schema,
                                                table,
                                                column,
                                                _gc,
                                                qident,
                                                type_sql,
                                            )
                                            if cerr:
                                                self.validation_log.insert(
                                                    tk.END,
                                                    f"[X] Column align: {cerr}\n",
                                                )
                                                error_count += 1
                                                continue
                                            if not script:
                                                self.validation_log.insert(
                                                    tk.END,
                                                    f"[WARN] No auto-align script for {obj_name}; skipping.\n",
                                                )
                                                error_count += 1
                                                continue
                                            self.validation_log.insert(
                                                tk.END,
                                                f"  Executing column align script (batched)...\n",
                                            )
                                            self.validation_log.see(tk.END)
                                            try:
                                                try:
                                                    from src.utils.sql import split_sql_on_go
                                                except ImportError:
                                                    split_sql_on_go = None
                                                batches = (
                                                    split_sql_on_go(script)
                                                    if split_sql_on_go
                                                    else [b.strip() for b in script.split("GO") if b.strip()]
                                                )
                                                for batch in batches:
                                                    batch = batch.strip()
                                                    if batch:
                                                        dest_cur.execute(batch)
                                                dest_conn.commit()
                                                self.validation_log.insert(
                                                    tk.END,
                                                    f"[OK] Aligned column: {schema}.{table}.{column}\n",
                                                )
                                                success_count += 1
                                            except Exception as al_err:
                                                self.validation_log.insert(
                                                    tk.END,
                                                    f"[X] Column align failed: {al_err}\n",
                                                )
                                                error_count += 1
                                                dest_conn.rollback()
                                            continue

                                        if column_already_on_dest:
                                            self.validation_log.insert(
                                                tk.END,
                                                f"[SKIP] Column already on destination: {schema}.{table}.{column}\n",
                                            )
                                            continue

                                        type_str = type_sql(
                                            col_def.type_name,
                                            col_def.max_length,
                                            col_def.precision,
                                            col_def.scale,
                                        )
                                        nullable = "NULL" if col_def.is_nullable else "NOT NULL"

                                        alter_sql = f"ALTER TABLE {qident(schema)}.{qident(table)} ADD {qident(column)} {type_str} {nullable}"

                                        if col_def.is_identity:
                                            seed = parse_int_or_default(col_def.seed_value_str, 1)
                                            inc = parse_int_or_default(col_def.increment_value_str, 1)
                                            alter_sql = (
                                                f"ALTER TABLE {qident(schema)}.{qident(table)} ADD {qident(column)} "
                                                f"{type_str} IDENTITY({seed},{inc}) {nullable}"
                                            )

                                        dest_cur.execute(alter_sql)
                                        dest_conn.commit()

                                        self.validation_log.insert(
                                            tk.END,
                                            f"[OK] Added column: {schema}.{table}.{column}\n",
                                        )
                                        success_count += 1
                                    else:
                                        self.validation_log.insert(tk.END, f"[X] Column definition not found: {obj_name}\n")
                                        error_count += 1
                                        
                                else:
                                    # Try as programmables: ObjectType.Schema.Object
                                    obj_type, schema, obj_name_only = parts
                                    self.validation_log.insert(tk.END, f"  Object type: {obj_type}\n")
                                    self.validation_log.insert(tk.END, f"  Fetching {obj_type} definition from source...\n")
                                    self.validation_log.see(tk.END)
                                    
                                    # Map object type descriptions to type codes
                                    # Handle both full names and abbreviations
                                    type_map = {
                                        'VIEW': 'V',
                                        'SQL_STORED_PROCEDURE': 'P',
                                        'SQL_STOR': 'P',  # Abbreviation
                                        'SQL_SCALAR_FUNCTION': 'FN',
                                        'SQL_TABLE_VALUED_FUNCTION': 'TF',
                                        'SQL_INLINE_TABLE_VALUED_FUNCTION': 'IF',
                                        'SQL_FUNCTION': 'FN',  # Generic function
                                        'CLR_STORED_PROCEDURE': 'PC',
                                        'CLR_SCALAR_FUNCTION': 'FS',
                                        'CLR_TABLE_VALUED_FUNCTION': 'FT',
                                        'CLR_AGGREGATE_FUNCTION': 'AF',
                                        'SYNONYM': 'SN',
                                        'SEQUENCE': 'SO'
                                    }
                                    
                                    obj_type_code = type_map.get(obj_type, 'P')  # Default to stored procedure
                                    
                                    # Get object definition
                                    try:
                                        src_cur.execute("""
                                            SELECT o.object_id
                                            FROM sys.objects o
                                            JOIN sys.schemas s ON s.schema_id = o.schema_id
                                            WHERE s.name = ? AND o.name = ? AND o.type = ?
                                        """, schema, obj_name_only, obj_type_code)
                                        
                                        row = src_cur.fetchone()
                                        if not row:
                                            self.validation_log.insert(tk.END, f"    [X] {obj_type} not found in source\n")
                                            error_count += 1
                                            continue
                                        
                                        obj_id = row[0]
                                        self.validation_log.insert(tk.END, f"    [OK] Found {obj_type} (object_id: {obj_id})\n")
                                        self.validation_log.insert(tk.END, f"    Getting object definition...\n")
                                        self.validation_log.see(tk.END)
                                        
                                        definition = object_definition(src_cur, obj_id)
                                        
                                        if not definition:
                                            self.validation_log.insert(tk.END, f"    [X] Could not get definition for {obj_type}\n")
                                            error_count += 1
                                            continue
                                        
                                        self.validation_log.insert(tk.END, f"    [OK] Got definition (length: {len(definition)} chars)\n")
                                        self.validation_log.see(tk.END)
                                    except Exception as fetch_err:
                                        self.validation_log.insert(tk.END, f"    [X] Error fetching {obj_type} definition: {str(fetch_err)}\n")
                                        error_count += 1
                                        continue
                                    
                                    if definition:
                                        # Check for Azure SQL incompatible references
                                        azure_incompatible_patterns = [
                                            ('master.dbo.', 'References to master database system tables'),
                                            ('master..', 'References to master database'),
                                            ('msdb.dbo.', 'References to msdb database'),
                                            ('msdb..', 'References to msdb database'),
                                            ('tempdb.dbo.', 'References to tempdb system tables'),
                                            ('syslogins', 'References to deprecated syslogins table'),
                                            ('xp_cmdshell', 'Extended stored procedure not supported'),
                                            ('xp_instance_regread', 'Extended stored procedure not supported'),
                                            ('xp_regread', 'Extended stored procedure not supported'),
                                            ('OPENROWSET', 'OPENROWSET not supported in Azure SQL'),
                                            ('OPENDATASOURCE', 'OPENDATASOURCE not supported in Azure SQL'),
                                            ('sp_addlinkedserver', 'Linked servers not supported in Azure SQL'),
                                            ('sp_configure', 'Server configuration not allowed in Azure SQL')
                                        ]
                                        
                                        incompatible_found = None
                                        definition_upper = definition.upper()
                                        for pattern, reason in azure_incompatible_patterns:
                                            if pattern.upper() in definition_upper:
                                                incompatible_found = (pattern, reason)
                                                break
                                        
                                        if incompatible_found:
                                            pattern, reason = incompatible_found
                                            self.validation_log.insert(tk.END, f"[WARN] Skipping {obj_type} {schema}.{obj_name_only}: {reason} ('{pattern}')\n")
                                            self.validation_log.insert(tk.END, f"  This object is not compatible with Azure SQL Database and requires manual review.\n")
                                            error_count += 1
                                            continue
                                        
                                        # Wrap with CREATE OR ALTER
                                        self.validation_log.insert(tk.END, f"    Wrapping with CREATE OR ALTER...\n")
                                        self.validation_log.see(tk.END)
                                        try:
                                            sql = wrap_create_or_alter(schema, obj_name_only, definition, obj_type)
                                            self.validation_log.insert(tk.END, f"    [OK] SQL wrapped (length: {len(sql)} chars)\n")
                                            self.validation_log.see(tk.END)
                                        except Exception as wrap_err:
                                            self.validation_log.insert(tk.END, f"    [X] Error wrapping SQL: {str(wrap_err)}\n")
                                            error_count += 1
                                            continue
                                        
                                        # Execute on destination
                                        self.validation_log.insert(tk.END, f"    Executing CREATE OR ALTER on destination...\n")
                                        self.validation_log.see(tk.END)
                                        try:
                                            batch_count = 0
                                            for batch in sql.split('GO'):
                                                batch = batch.strip()
                                                if batch:
                                                    batch_count += 1
                                                    self.validation_log.insert(tk.END, f"      Executing batch {batch_count}...\n")
                                                    dest_cur.execute(batch)
                                            dest_conn.commit()
                                            self.validation_log.insert(tk.END, f"    [OK] All batches executed successfully\n")
                                            self.validation_log.insert(tk.END, f"  [OK] Created {obj_type}: {schema}.{obj_name_only}\n")
                                            success_count += 1
                                        except Exception as exec_err:
                                            self.validation_log.insert(tk.END, f"    [X] Error executing CREATE OR ALTER: {str(exec_err)}\n")
                                            self.validation_log.insert(tk.END, f"    SQL (first 500 chars): {sql[:500]}...\n")
                                            import traceback
                                            self.validation_log.insert(tk.END, f"    Traceback: {traceback.format_exc()}\n")
                                            self.validation_log.insert(tk.END, f"  [X] Error creating {obj_type} {obj_name}: {str(exec_err)}\n")
                                            error_count += 1
                                            dest_conn.rollback()
                                        self.validation_log.see(tk.END)
                                        
                        else:
                            self.validation_log.insert(tk.END, f"[X] Unsupported object format: {obj_name}\n")
                            error_count += 1
                                
                    except Exception as e:
                        self.validation_log.insert(tk.END, f"  [X] Error fixing {obj_name}: {str(e)}\n")
                        import traceback
                        self.validation_log.insert(tk.END, f"    Traceback: {traceback.format_exc()}\n")
                        self.validation_log.see(tk.END)
                        error_count += 1
                        try:
                            dest_conn.rollback()
                            self.validation_log.insert(tk.END, f"    [OK] Transaction rolled back\n")
                        except Exception as rollback_err:
                            self.validation_log.insert(tk.END, f"    [WARN] Error during rollback: {str(rollback_err)}\n")
                        self.validation_log.see(tk.END)
                
                # Success message after processing all objects for this database
                self.validation_log.insert(tk.END, f"\n{'='*60}\n")
                self.validation_log.insert(tk.END, f"FIX OPERATION COMPLETE\n")
                self.validation_log.insert(tk.END, f"  Total objects processed: {len(missing_objects)}\n")
                self.validation_log.insert(tk.END, f"  [OK] Succeeded: {success_count}\n")
                self.validation_log.insert(tk.END, f"  [X] Failed: {error_count}\n")
                self.validation_log.insert(tk.END, f"  Success rate: {(success_count/len(missing_objects)*100) if missing_objects else 0:.1f}%\n")
                self.validation_log.see(tk.END)
                return (success_count, error_count)
                    
            except Exception as e:
                self.validation_log.insert(tk.END, f"\n[X] Error processing objects: {str(e)}\n")
                import traceback
                self.validation_log.insert(tk.END, f"Traceback: {traceback.format_exc()}\n")
                return (0, len(missing_objects))
                    
        except Exception as e:
            self.validation_log.insert(tk.END, f"\n[X] Error: {str(e)}\n")
            import traceback
            self.validation_log.insert(tk.END, f"Traceback: {traceback.format_exc()}\n")
            return (0, len(missing_objects))
        finally:
            # Close connections
            try:
                if 'src_conn' in locals():
                    src_conn.close()
            except:
                pass
            try:
                if 'dest_conn' in locals():
                    dest_conn.close()
            except:
                pass
        
        # Return (0, 0) if we reach here without returning earlier
        return (0, 0)
    
    def _filter_results(self):
        """Filter treeview results by status and search term."""
        status_filter = self.status_filter_var.get()
        search_term = self.search_var.get().lower()
        
        # Show/hide items based on filter
        for item in self.all_tree_items:
            try:
                values = self.results_tree.item(item)
                status = self.results_tree.set(item, "Status")
                obj = values['text']
                db_name = self.results_tree.set(item, "DB")
                
                # Check status filter (exact match on combobox value, not substring)
                status_match = (status_filter == "All" or status == status_filter)
                
                # Check search filter
                search_match = (not search_term or 
                               search_term in obj.lower() or 
                               search_term in db_name.lower() or
                               search_term in status.lower())
                
                # Show or hide item
                if status_match and search_match:
                    self.results_tree.reattach(item, "", 0)  # Show item
                else:
                    self.results_tree.detach(item)  # Hide item
            except:
                pass  # Item might have been deleted
                
    def _clear_filter(self):
        """Clear all filters and show all results."""
        self.status_filter_var.set("All")
        self.search_var.set("")
        
        # Re-attach all items
        for item in self.all_tree_items:
            try:
                self.results_tree.reattach(item, "", 0)
            except:
                pass

