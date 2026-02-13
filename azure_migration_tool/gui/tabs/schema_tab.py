# Author: S@tish Chauhan

"""
Schema Backup/Migration Tab
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from pathlib import Path
import threading
import sys
import os
import json
import pandas as pd

# Add parent directories to path
parent_dir = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(parent_dir))

from gui.utils.excel_utils import read_excel_file, create_sample_excel
from gui.widgets.connection_widget import ConnectionWidget
from gui.widgets.schema_tree import SchemaTree
from gui.widgets.schema_diff_viewer import SchemaDiffViewer
from gui.utils.schema_comparison import compare_schemas
from gui.utils.schema_script_generator import generate_script_for_objects
from gui.utils.database_utils import connect_with_msal_cache, connect_to_any_database
import logging

# Import pick_sql_driver
try:
    from src.utils.database import pick_sql_driver
except ImportError:
    try:
        from src.utils.database import pick_sql_driver
    except ImportError:
        def pick_sql_driver(logger=None):
            import pyodbc
            drivers = [x for x in pyodbc.drivers() if x.startswith("ODBC Driver")]
            if drivers:
                return drivers[0]
            return "ODBC Driver 18 for SQL Server"

# Import schema_backup
try:
    from src.backup import schema_backup as sechma_backup
except ImportError:
    try:
        import schema_backup as sechma_backup
    except ImportError:
        try:
            import sechma_backup
        except ImportError:
            sechma_backup = None

# Import restore_schema
try:
    from src.restore import schema_restore as restore_schema
except ImportError:
    try:
        import restore_schema
    except ImportError:
        restore_schema = None


class SchemaTab:
    """Schema backup and migration tab."""
    
    def __init__(self, parent, main_window):
        self.main_window = main_window
        self.frame = ttk.Frame(parent)
        self.project_path = None
        
        self._create_widgets()
        
    def set_project_path(self, project_path):
        """Set the current project path."""
        self.project_path = project_path
        
    def _create_widgets(self):
        """Create UI widgets."""
        # Create scrollable canvas
        canvas = tk.Canvas(self.frame)
        scrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        def on_canvas_configure(event):
            canvas_width = event.width
            canvas.itemconfig(canvas_frame, width=canvas_width)
        
        canvas.bind('<Configure>', on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Title
        title_label = tk.Label(
            scrollable_frame,
            text="Schema Backup & Migration",
            font=("Arial", 16, "bold")
        )
        title_label.pack(pady=10)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Store reference for later use
        self.scrollable_frame = scrollable_frame
        
        # Create notebook: Backup | Restore | Bulk (Excel) | Compare
        notebook = ttk.Notebook(scrollable_frame)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 1. Backup sub-tab (step-based, single DB only)
        backup_frame = ttk.Frame(notebook)
        notebook.add(backup_frame, text="Backup")
        self._create_backup_widgets(backup_frame)
        
        # 2. Restore sub-tab (step-based, single DB only; preview off by default)
        restore_frame = ttk.Frame(notebook)
        notebook.add(restore_frame, text="Restore")
        self._create_restore_widgets(restore_frame)
        
        # 3. Bulk (Excel) sub-tab - one view, bulk backup + bulk restore + one log
        bulk_frame = ttk.Frame(notebook)
        notebook.add(bulk_frame, text="Bulk (Excel)")
        self._create_bulk_tab_widgets(bulk_frame)
        
        # 4. Compare sub-tab (schema comparison)
        comparison_frame = ttk.Frame(notebook)
        notebook.add(comparison_frame, text="Compare")
        self._create_comparison_widgets(comparison_frame)
        
        # Store latest backup path for auto-fill
        self.latest_backup_path = None
        self.src_conn = None
        self.dest_conn = None
        self.comparison_result = None
        
    def _create_comparison_widgets(self, parent):
        """Create schema comparison interface."""
        # Connection panels (side-by-side at top)
        conn_paned = ttk.PanedWindow(parent, orient=tk.HORIZONTAL)
        conn_paned.pack(fill=tk.X, padx=5, pady=5)
        
        # Source connection frame
        src_conn_frame = ttk.LabelFrame(conn_paned, text="Source Database", padding=10)
        conn_paned.add(src_conn_frame, weight=1)
        
        self.src_server_var = tk.StringVar()
        self.src_db_var = tk.StringVar()
        self.src_auth_var = tk.StringVar(value="entra_mfa")
        self.src_user_var = tk.StringVar()
        self.src_password_var = tk.StringVar()
        self.src_db_type_var = tk.StringVar(value="sqlserver")
        self.src_port_var = tk.StringVar(value="50000")
        self.src_schema_var = tk.StringVar(value="")
        
        self.src_connection_widget = ConnectionWidget(
            parent=src_conn_frame,
            server_var=self.src_server_var,
            db_var=self.src_db_var,
            auth_var=self.src_auth_var,
            user_var=self.src_user_var,
            password_var=self.src_password_var,
            label_text="",
            row_start=0,
            db_type_var=self.src_db_type_var,
            port_var=self.src_port_var,
            schema_var=self.src_schema_var
        )
        
        # Get the last row used by ConnectionWidget (it uses rows 0-7 with db_type, port, and schema)
        last_row = 8
        src_btn_frame = ttk.Frame(src_conn_frame)
        src_btn_frame.grid(row=last_row, column=0, columnspan=2, pady=5, sticky=tk.W)
        self.src_connect_btn = ttk.Button(src_btn_frame, text="Connect", command=self._connect_source)
        self.src_connect_btn.pack(side=tk.LEFT, padx=5)
        self.src_status_label = tk.Label(src_conn_frame, text="Not connected", fg="gray")
        self.src_status_label.grid(row=last_row+1, column=0, columnspan=2, sticky=tk.W, padx=5)
        
        # Destination connection frame
        dest_conn_frame = ttk.LabelFrame(conn_paned, text="Destination Database", padding=10)
        conn_paned.add(dest_conn_frame, weight=1)
        
        self.dest_server_var = tk.StringVar()
        self.dest_db_var = tk.StringVar()
        self.dest_auth_var = tk.StringVar(value="entra_mfa")
        self.dest_user_var = tk.StringVar()
        self.dest_password_var = tk.StringVar()
        self.dest_db_type_var = tk.StringVar(value="sqlserver")
        self.dest_port_var = tk.StringVar(value="50000")
        self.dest_schema_var = tk.StringVar(value="")
        
        self.dest_connection_widget = ConnectionWidget(
            parent=dest_conn_frame,
            server_var=self.dest_server_var,
            db_var=self.dest_db_var,
            auth_var=self.dest_auth_var,
            user_var=self.dest_user_var,
            password_var=self.dest_password_var,
            label_text="",
            row_start=0,
            db_type_var=self.dest_db_type_var,
            port_var=self.dest_port_var,
            schema_var=self.dest_schema_var
        )
        
        # Get the last row used by ConnectionWidget (it uses rows 0-7 with db_type, port, and schema)
        last_row = 8
        dest_btn_frame = ttk.Frame(dest_conn_frame)
        dest_btn_frame.grid(row=last_row, column=0, columnspan=2, pady=5, sticky=tk.W)
        self.dest_connect_btn = ttk.Button(dest_btn_frame, text="Connect", command=self._connect_destination)
        self.dest_connect_btn.pack(side=tk.LEFT, padx=5)
        self.dest_status_label = tk.Label(dest_conn_frame, text="Not connected", fg="gray")
        self.dest_status_label.grid(row=last_row+1, column=0, columnspan=2, sticky=tk.W, padx=5)
        
        # Compare button
        compare_btn_frame = ttk.Frame(parent)
        compare_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        self.compare_btn = ttk.Button(compare_btn_frame, text="Compare Schemas", 
                                     command=self._compare_schemas, state=tk.DISABLED)
        self.compare_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_comparison_btn = ttk.Button(compare_btn_frame, text="Export Report", 
                                               command=self._export_comparison_report, state=tk.DISABLED)
        self.export_comparison_btn.pack(side=tk.LEFT, padx=5)
        
        # Comparison tree (side-by-side)
        tree_paned = ttk.PanedWindow(parent, orient=tk.HORIZONTAL)
        tree_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Source tree
        src_tree_frame = ttk.LabelFrame(tree_paned, text="Source Objects", padding=5)
        tree_paned.add(src_tree_frame, weight=1)
        self.src_tree = SchemaTree(src_tree_frame, title="")
        self.src_tree.pack(fill=tk.BOTH, expand=True)
        
        # Destination tree
        dest_tree_frame = ttk.LabelFrame(tree_paned, text="Destination Objects", padding=5)
        tree_paned.add(dest_tree_frame, weight=1)
        self.dest_tree = SchemaTree(dest_tree_frame, title="")
        self.dest_tree.pack(fill=tk.BOTH, expand=True)
        
        # Selection controls
        selection_frame = ttk.LabelFrame(parent, text="Selection Controls", padding=10)
        selection_frame.pack(fill=tk.X, padx=5, pady=5)
        
        btn_row = ttk.Frame(selection_frame)
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="Select All", command=self.src_tree.select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_row, text="Select Missing", command=self.src_tree.select_missing).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_row, text="Deselect All", command=self.src_tree.deselect_all).pack(side=tk.LEFT, padx=5)
        
        # Script generation and preview
        script_frame = ttk.LabelFrame(parent, text="Script Generation", padding=10)
        script_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        script_btn_frame = ttk.Frame(script_frame)
        script_btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(script_btn_frame, text="Generate Script", command=self._generate_script).pack(side=tk.LEFT, padx=5)
        self.preview_btn = ttk.Button(script_btn_frame, text="Preview Script", command=self._preview_script, state=tk.DISABLED)
        self.preview_btn.pack(side=tk.LEFT, padx=5)
        self.save_btn = ttk.Button(script_btn_frame, text="Save Script", command=self._save_script, state=tk.DISABLED)
        self.save_btn.pack(side=tk.LEFT, padx=5)
        self.execute_btn = ttk.Button(script_btn_frame, text="Execute Script", command=self._execute_script, state=tk.DISABLED)
        self.execute_btn.pack(side=tk.LEFT, padx=5)
        
        self.script_preview = scrolledtext.ScrolledText(script_frame, height=10, wrap=tk.WORD, font=("Consolas", 9))
        self.script_preview.pack(fill=tk.BOTH, expand=True)
        
        self.generated_script = None
    
    def _create_bulk_tab_widgets(self, parent):
        """Bulk (Excel) tab: load configs, Start Bulk Backup / Start Bulk Restore, one log."""
        tk.Label(parent, text="Bulk Backup & Restore (Excel)", font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Load config section
        load_frame = ttk.LabelFrame(parent, text="Load configuration", padding=10)
        load_frame.pack(fill=tk.X, padx=5, pady=5)
        
        row0 = ttk.Frame(load_frame)
        row0.pack(fill=tk.X)
        ttk.Button(row0, text="Download backup template", command=lambda: self._download_template("schema_backup")).pack(side=tk.LEFT, padx=5)
        ttk.Button(row0, text="Upload Excel (backup)", command=self._upload_excel_backup).pack(side=tk.LEFT, padx=5)
        tk.Label(load_frame, textvariable=self.backup_excel_file_var, fg="gray").pack(anchor=tk.W)
        
        row1 = ttk.Frame(load_frame)
        row1.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(row1, text="Download restore template", command=lambda: self._download_template("schema_restore")).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="Upload Excel (restore)", command=self._upload_excel_restore).pack(side=tk.LEFT, padx=5)
        tk.Label(load_frame, textvariable=self.restore_excel_file_var, fg="gray").pack(anchor=tk.W)
        
        # Backup output folder for bulk backup
        folder_frame = ttk.LabelFrame(parent, text="Backup output folder (for bulk backup)", padding=10)
        folder_frame.pack(fill=tk.X, padx=5, pady=5)
        path_row = ttk.Frame(folder_frame)
        path_row.pack(fill=tk.X)
        ttk.Entry(path_row, textvariable=self.backup_output_var, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(path_row, text="Browse...", command=self._browse_backup_output).pack(side=tk.LEFT, padx=5)
        
        # Buttons
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(pady=10)
        self.backup_bulk_btn = ttk.Button(btn_frame, text="Start Bulk Backup", command=self._start_bulk_backup, width=20, state=tk.DISABLED)
        self.backup_bulk_btn.pack(side=tk.LEFT, padx=5)
        self.restore_bulk_btn = ttk.Button(btn_frame, text="Start Bulk Restore", command=self._start_bulk_restore, width=20, state=tk.DISABLED)
        self.restore_bulk_btn.pack(side=tk.LEFT, padx=5)
        
        # Single log for bulk operations
        log_frame = ttk.LabelFrame(parent, text="Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.bulk_log = scrolledtext.ScrolledText(log_frame, height=12, wrap=tk.WORD)
        self.bulk_log.pack(fill=tk.BOTH, expand=True)

    def _connect_source(self):
        """Connect to source database."""
        def connect():
            try:
                self.src_conn = connect_to_any_database(
                    server=self.src_server_var.get(),
                    database=self.src_db_var.get(),
                    auth=self.src_auth_var.get(),
                    user=self.src_user_var.get(),
                    password=self.src_password_var.get() or None,
                    db_type=self.src_db_type_var.get(),
                    port=int(self.src_port_var.get() or 50000),
                    timeout=30
                )
                self.frame.after(0, lambda: self.src_status_label.config(text="Connected", fg="green"))
                self._update_compare_button()
            except Exception as e:
                err_msg = str(e)
                self.frame.after(0, lambda msg=err_msg: messagebox.showerror("Connection Error", f"Failed to connect to source: {msg}"))
                self.frame.after(0, lambda: self.src_status_label.config(text="Connection failed", fg="red"))
        
        threading.Thread(target=connect, daemon=True).start()
    
    def _connect_destination(self):
        """Connect to destination database."""
        def connect():
            try:
                self.dest_conn = connect_to_any_database(
                    server=self.dest_server_var.get(),
                    database=self.dest_db_var.get(),
                    auth=self.dest_auth_var.get(),
                    user=self.dest_user_var.get(),
                    password=self.dest_password_var.get() or None,
                    db_type=self.dest_db_type_var.get(),
                    port=int(self.dest_port_var.get() or 50000),
                    timeout=30
                )
                self.frame.after(0, lambda: self.dest_status_label.config(text="Connected", fg="green"))
                self._update_compare_button()
            except Exception as e:
                err_msg = str(e)
                self.frame.after(0, lambda msg=err_msg: messagebox.showerror("Connection Error", f"Failed to connect to destination: {msg}"))
                self.frame.after(0, lambda: self.dest_status_label.config(text="Connection failed", fg="red"))
        
        threading.Thread(target=connect, daemon=True).start()
    
    def _update_compare_button(self):
        """Update compare button state based on connections."""
        if self.src_conn and self.dest_conn:
            self.compare_btn.config(state=tk.NORMAL)
        else:
            self.compare_btn.config(state=tk.DISABLED)
    
    def _compare_schemas(self):
        """Compare source and destination schemas."""
        if not self.src_conn or not self.dest_conn:
            messagebox.showerror("Error", "Please connect to both source and destination databases first!")
            return
        
        # Get schema values for DB2 filtering
        src_schema = self.src_schema_var.get().strip() if self.src_schema_var.get() else None
        dest_schema = self.dest_schema_var.get().strip() if self.dest_schema_var.get() else None
        
        # Store connection info for export
        self.comparison_info = {
            'src_server': self.src_server_var.get(),
            'src_db': self.src_db_var.get(),
            'src_schema': src_schema,
            'src_db_type': self.src_db_type_var.get(),
            'dest_server': self.dest_server_var.get(),
            'dest_db': self.dest_db_var.get(),
            'dest_schema': dest_schema,
            'dest_db_type': self.dest_db_type_var.get()
        }
        
        def compare():
            try:
                import logging
                logger = logging.getLogger(__name__)
                comparison = compare_schemas(
                    self.src_conn, 
                    self.dest_conn, 
                    src_schema=src_schema,
                    dest_schema=dest_schema,
                    logger=logger
                )
                self.comparison_result = comparison
                # Load into trees
                self.frame.after(0, lambda: self.src_tree.load_comparison(comparison))
                self.frame.after(0, lambda: self.dest_tree.load_comparison(comparison))
                self.frame.after(0, lambda: self.export_comparison_btn.config(state=tk.NORMAL))
                self.frame.after(0, lambda: messagebox.showinfo("Success", "Schema comparison completed!"))
            except Exception as e:
                err_msg = str(e)
                self.frame.after(0, lambda msg=err_msg: messagebox.showerror("Error", f"Comparison failed: {msg}"))
        
        threading.Thread(target=compare, daemon=True).start()
    
    def _export_comparison_report(self):
        """Export schema comparison report to Excel."""
        if not hasattr(self, 'comparison_result') or not self.comparison_result:
            messagebox.showwarning("Warning", "No comparison results to export! Please run comparison first.")
            return
        
        # Ask user if they want to export only differences
        export_only_diffs = messagebox.askyesno(
            "Export Options",
            "Export only differences (exclude matched objects)?\n\n"
            "Yes = Only differences (smaller file)\n"
            "No = All objects (including matches)"
        )
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("JSON files", "*.json"), ("All files", "*.*")],
            initialfilename=f"schema_comparison_{self.comparison_info.get('src_db', 'source')}_{self.comparison_info.get('dest_db', 'dest')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            from datetime import datetime
            
            # Store preference
            self.export_only_differences = export_only_diffs
            
            if filename.endswith('.xlsx'):
                self._export_to_excel(filename)
            else:
                self._export_to_json(filename)
                
        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Failed to export report: {str(e)}\n\n{traceback.format_exc()}")
    
    def _export_to_excel(self, filename):
        """Export comparison to Excel with multiple sheets and enhanced details."""
        import pandas as pd
        from datetime import datetime
        from openpyxl.styles import PatternFill, Font
        
        comparison = self.comparison_result
        info = getattr(self, 'comparison_info', {})
        show_only_differences = getattr(self, 'export_only_differences', False)
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Summary sheet with statistics
            summary_data = []
            total_objects = 0
            total_matched = 0
            total_missing = 0
            total_extra = 0
            
            for obj_type, result in comparison.items():
                if isinstance(result, dict):
                    if 'missing' in result:
                        # List-based comparison
                        matching_count = len(result.get('matching', []))
                        missing_count = len(result.get('missing', []))
                        extra_count = len(result.get('extra', []))
                        
                        total_objects += matching_count + missing_count
                        total_matched += matching_count
                        total_missing += missing_count
                        total_extra += extra_count
                        
                        match_pct = (matching_count / (matching_count + missing_count) * 100) if (matching_count + missing_count) > 0 else 0
                        summary_data.append({
                            'Object Type': obj_type,
                            'Source Count': matching_count + missing_count,
                            'Dest Count': matching_count + extra_count,
                            'Matching': matching_count,
                            'Missing in Dest': missing_count,
                            'Extra in Dest': extra_count,
                            'Match %': f"{match_pct:.1f}%",
                            'Status': 'Match' if missing_count == 0 and extra_count == 0 else 'Differences'
                        })
                    elif 'status' in result:
                        # SQL-based comparison
                        summary_data.append({
                            'Object Type': obj_type,
                            'Source Count': 'N/A',
                            'Dest Count': 'N/A',
                            'Matching': 'N/A',
                            'Missing in Dest': 'N/A',
                            'Extra in Dest': 'N/A',
                            'Match %': 'N/A',
                            'Status': result.get('status', 'unknown')
                        })
            
            # Add overall summary row
            if total_objects > 0:
                overall_match_pct = (total_matched / total_objects * 100)
                summary_data.append({
                    'Object Type': 'TOTAL',
                    'Source Count': total_objects,
                    'Dest Count': total_matched + total_extra,
                    'Matching': total_matched,
                    'Missing in Dest': total_missing,
                    'Extra in Dest': total_extra,
                    'Match %': f"{overall_match_pct:.1f}%",
                    'Status': 'Match' if total_missing == 0 and total_extra == 0 else 'Differences'
                })
            
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Details for each object type
            for obj_type, result in comparison.items():
                if isinstance(result, dict) and 'missing' in result:
                    details = []
                    
                    # Matching objects (only if not filtering)
                    if not show_only_differences:
                        for obj in result.get('matching', []):
                            obj_name = '.'.join(str(o) for o in obj) if isinstance(obj, tuple) else str(obj)
                            details.append({
                                'Object': obj_name,
                                'Status': 'Match',
                                'Source': 'Exists',
                                'Destination': 'Exists',
                                'Action Required': 'No Action'
                            })
                    
                    # Missing in destination
                    for obj in result.get('missing', []):
                        obj_name = '.'.join(str(o) for o in obj) if isinstance(obj, tuple) else str(obj)
                        details.append({
                            'Object': obj_name,
                            'Status': 'Missing in Dest',
                            'Source': 'Exists',
                            'Destination': 'Missing',
                            'Action Required': 'Create Object'
                        })
                    
                    # Extra in destination
                    for obj in result.get('extra', []):
                        obj_name = '.'.join(str(o) for o in obj) if isinstance(obj, tuple) else str(obj)
                        details.append({
                            'Object': obj_name,
                            'Status': 'Extra in Dest',
                            'Source': 'Missing',
                            'Destination': 'Exists',
                            'Action Required': 'Review/Drop'
                        })
                    
                    if details:
                        details_df = pd.DataFrame(details)
                        # Truncate sheet name to 31 chars (Excel limit)
                        sheet_name = obj_type[:31]
                        details_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Connection info sheet with statistics
            overall_match_pct = (total_matched / total_objects * 100) if total_objects > 0 else 0
            info_data = [
                {'Property': 'Export Date', 'Value': datetime.now().isoformat()},
                {'Property': 'Source Server', 'Value': info.get('src_server', '')},
                {'Property': 'Source Database', 'Value': info.get('src_db', '')},
                {'Property': 'Source Schema', 'Value': info.get('src_schema', 'All')},
                {'Property': 'Source Type', 'Value': info.get('src_db_type', '')},
                {'Property': 'Dest Server', 'Value': info.get('dest_server', '')},
                {'Property': 'Dest Database', 'Value': info.get('dest_db', '')},
                {'Property': 'Dest Schema', 'Value': info.get('dest_schema', 'All')},
                {'Property': 'Dest Type', 'Value': info.get('dest_db_type', '')},
                {'Property': '', 'Value': ''},
                {'Property': 'Total Objects', 'Value': total_objects},
                {'Property': 'Matched', 'Value': total_matched},
                {'Property': 'Missing in Dest', 'Value': total_missing},
                {'Property': 'Extra in Dest', 'Value': total_extra},
                {'Property': 'Match Rate', 'Value': f"{overall_match_pct:.1f}%"},
                {'Property': 'Export Mode', 'Value': 'Differences Only' if show_only_differences else 'All Objects'}
            ]
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, sheet_name='Connection Info', index=False)
            
            # Apply color coding to all sheets
            workbook = writer.book
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Find Status column
                status_col = None
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value and 'Status' in str(cell.value):
                        status_col = col_idx
                        break
                
                if status_col:
                    # Apply colors based on status
                    for row_idx in range(2, sheet.max_row + 1):
                        cell = sheet.cell(row=row_idx, column=status_col)
                        status_val = str(cell.value).upper() if cell.value else ''
                        if 'MATCH' in status_val and 'MISMATCH' not in status_val:
                            cell.fill = green_fill
                        elif 'EXTRA' in status_val or 'DIFFERENCE' in status_val:
                            cell.fill = yellow_fill
                        elif 'MISSING' in status_val or 'MISMATCH' in status_val or 'ERROR' in status_val:
                            cell.fill = red_fill
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        
        messagebox.showinfo("Success", f"Report exported to Excel:\n{filename}\n\nSheets: Summary, Details by Type, Connection Info")
    
    def _export_to_json(self, filename):
        """Export comparison to JSON."""
        import json
        from datetime import datetime
        
        comparison = self.comparison_result
        info = getattr(self, 'comparison_info', {})
        
        # Convert tuples to strings for JSON serialization
        serializable_comparison = {}
        for obj_type, result in comparison.items():
            if isinstance(result, dict):
                serializable_result = {}
                for key, value in result.items():
                    if isinstance(value, list):
                        serializable_result[key] = [
                            '.'.join(str(o) for o in item) if isinstance(item, tuple) else str(item)
                            for item in value
                        ]
                    else:
                        serializable_result[key] = value
                serializable_comparison[obj_type] = serializable_result
            else:
                serializable_comparison[obj_type] = result
        
        report = {
            'timestamp': datetime.now().isoformat(),
            'source': {
                'server': info.get('src_server', ''),
                'database': info.get('src_db', ''),
                'schema': info.get('src_schema', ''),
                'type': info.get('src_db_type', '')
            },
            'destination': {
                'server': info.get('dest_server', ''),
                'database': info.get('dest_db', ''),
                'schema': info.get('dest_schema', ''),
                'type': info.get('dest_db_type', '')
            },
            'comparison': serializable_comparison
        }
        
        with open(filename, 'w') as f:
            json.dump(report, f, indent=2)
        
        messagebox.showinfo("Success", f"Report exported to JSON:\n{filename}")
    
    def _generate_script(self):
        """Generate script for selected objects."""
        if not self.src_conn:
            messagebox.showerror("Error", "Please connect to source database first!")
            return
        
        selected = self.src_tree.get_selected_objects()
        if not selected:
            messagebox.showwarning("Warning", "Please select objects to generate script for!")
            return
        
        def generate():
            try:
                import logging
                logger = logging.getLogger(__name__)
                script = generate_script_for_objects(self.src_conn, selected, logger)
                
                # Schedule UI updates on the main thread
                def update_ui(script_text=script, count=len(selected)):
                    self.generated_script = script_text
                    self.script_preview.delete("1.0", tk.END)
                    self.script_preview.insert("1.0", script_text)
                    # Enable preview/save/execute buttons
                    self.preview_btn.config(state=tk.NORMAL)
                    self.save_btn.config(state=tk.NORMAL)
                    self.execute_btn.config(state=tk.NORMAL)
                    messagebox.showinfo("Success", f"Generated script for {count} object(s)!")
                
                self.frame.after(0, update_ui)
            except Exception as e:
                err_msg = str(e)
                self.frame.after(0, lambda msg=err_msg: messagebox.showerror("Error", f"Script generation failed: {msg}"))
        
        threading.Thread(target=generate, daemon=True).start()
    
    def _preview_script(self):
        """Preview generated script in a dialog."""
        if not self.generated_script:
            messagebox.showwarning("Warning", "No script generated yet!")
            return
        
        # Create a preview dialog window
        preview_window = tk.Toplevel(self.frame)
        preview_window.title("Script Preview")
        preview_window.geometry("900x600")
        preview_window.transient(self.frame.winfo_toplevel())
        
        # Add a text widget with the script
        text_frame = ttk.Frame(preview_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        preview_text = scrolledtext.ScrolledText(text_frame, wrap=tk.WORD, font=("Consolas", 10))
        preview_text.pack(fill=tk.BOTH, expand=True)
        preview_text.insert("1.0", self.generated_script)
        preview_text.config(state=tk.DISABLED)
        
        # Add close button
        btn_frame = ttk.Frame(preview_window)
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        ttk.Button(btn_frame, text="Close", command=preview_window.destroy).pack(side=tk.RIGHT)
    
    def _save_script(self):
        """Save generated script to file."""
        if not self.generated_script:
            messagebox.showwarning("Warning", "No script generated yet!")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Save Script",
            defaultextension=".sql",
            filetypes=[("SQL files", "*.sql"), ("All files", "*.*")]
        )
        if file_path:
            try:
                Path(file_path).write_text(self.generated_script, encoding="utf-8")
                messagebox.showinfo("Success", f"Script saved to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save script: {str(e)}")
    
    def _execute_script(self):
        """Execute generated script on destination database."""
        if not self.dest_conn:
            messagebox.showerror("Error", "Please connect to destination database first!")
            return
        if not self.generated_script:
            messagebox.showwarning("Warning", "No script generated yet!")
            return
        
        if not messagebox.askyesno("Confirm", "Execute script on destination database?"):
            return
        
        def execute():
            try:
                cur = self.dest_conn.cursor()
                # Split script by GO statements
                batches = [b.strip() for b in self.generated_script.split("GO") if b.strip()]
                for batch in batches:
                    cur.execute(batch)
                self.dest_conn.commit()
                messagebox.showinfo("Success", "Script executed successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Script execution failed: {str(e)}")
        
        threading.Thread(target=execute, daemon=True).start()
    
    def _create_backup_widgets(self, parent):
        """Backup sub-tab: Step 1 Source, Step 2 What to backup, Step 3 Where to save, Start Backup, Log."""
        tk.Label(parent, text="Schema Backup", font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Step 1: Source database
        conn_frame = ttk.LabelFrame(parent, text="Step 1: Source database", padding=10)
        conn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.backup_server_var = tk.StringVar()
        self.backup_db_var = tk.StringVar()
        self.backup_auth_var = tk.StringVar(value="entra_mfa")
        self.backup_user_var = tk.StringVar()
        self.backup_password_var = tk.StringVar()
        
        self.backup_connection_widget = ConnectionWidget(
            parent=conn_frame,
            server_var=self.backup_server_var,
            db_var=self.backup_db_var,
            auth_var=self.backup_auth_var,
            user_var=self.backup_user_var,
            password_var=self.backup_password_var,
            label_text="",
            row_start=0
        )
        
        # Step 2: What to backup
        options_frame = ttk.LabelFrame(parent, text="Step 2: What to backup", padding=10)
        options_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.backup_tables_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Tables", variable=self.backup_tables_var).pack(anchor=tk.W)
        self.backup_programmables_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Programmables (Views, Procedures, Functions)", variable=self.backup_programmables_var).pack(anchor=tk.W)
        self.backup_constraints_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Constraints (Foreign Keys, Check)", variable=self.backup_constraints_var).pack(anchor=tk.W)
        self.backup_indexes_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Indexes", variable=self.backup_indexes_var).pack(anchor=tk.W)
        
        # Step 3: Where to save
        out_frame = ttk.LabelFrame(parent, text="Step 3: Where to save", padding=10)
        out_frame.pack(fill=tk.X, padx=5, pady=5)
        self.backup_output_var = tk.StringVar()
        output_row = ttk.Frame(out_frame)
        output_row.pack(fill=tk.X)
        ttk.Entry(output_row, textvariable=self.backup_output_var, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(output_row, text="Browse...", command=self._browse_backup_output).pack(side=tk.LEFT, padx=5)
        tk.Label(out_frame, text="(Leave empty to use project folder / backups)", fg="gray").pack(anchor=tk.W)
        
        self.backup_excel_file_var = tk.StringVar()
        self.backup_configs = []
        
        # Start Backup
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(pady=10)
        self.backup_btn = ttk.Button(btn_frame, text="Start Backup", command=self._start_backup, width=20)
        self.backup_btn.pack(side=tk.LEFT, padx=5)
        
        # Log
        log_frame = ttk.LabelFrame(parent, text="Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.backup_log = scrolledtext.ScrolledText(log_frame, height=10, wrap=tk.WORD)
        self.backup_log.pack(fill=tk.BOTH, expand=True)
        
    def _create_restore_widgets(self, parent):
        """Restore sub-tab: Step 1 Backup path, Step 2 Destination, Step 3 Options (preview off by default), Start Restore, Log."""
        tk.Label(parent, text="Schema Restore", font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Step 1: Backup to restore
        backup_path_frame = ttk.LabelFrame(parent, text="Step 1: Backup to restore", padding=10)
        backup_path_frame.pack(fill=tk.X, padx=5, pady=5)
        tk.Label(backup_path_frame, text="Backup folder (run folder from a completed backup):").pack(anchor=tk.W)
        path_frame = ttk.Frame(backup_path_frame)
        path_frame.pack(fill=tk.X, pady=5)
        self.restore_backup_path_var = tk.StringVar()
        ttk.Entry(path_frame, textvariable=self.restore_backup_path_var, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(path_frame, text="Browse...", command=self._browse_restore_backup).pack(side=tk.LEFT, padx=5)
        tk.Label(backup_path_frame, text="Tip: Use the path shown in Backup log after a successful run.", fg="gray").pack(anchor=tk.W)
        
        # Step 2: Destination database
        dest_frame = ttk.LabelFrame(parent, text="Step 2: Destination database", padding=10)
        dest_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.restore_server_var = tk.StringVar()
        self.restore_db_var = tk.StringVar()
        self.restore_auth_var = tk.StringVar(value="entra_mfa")
        self.restore_user_var = tk.StringVar()
        self.restore_password_var = tk.StringVar()
        
        self.restore_connection_widget = ConnectionWidget(
            parent=dest_frame,
            server_var=self.restore_server_var,
            db_var=self.restore_db_var,
            auth_var=self.restore_auth_var,
            user_var=self.restore_user_var,
            password_var=self.restore_password_var,
            label_text="",
            row_start=0
        )
        
        # Step 3: What to restore + options
        restore_options_frame = ttk.LabelFrame(parent, text="Step 3: What to restore", padding=10)
        restore_options_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.restore_tables_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(restore_options_frame, text="Tables", variable=self.restore_tables_var).pack(anchor=tk.W)
        self.restore_programmables_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(restore_options_frame, text="Programmables", variable=self.restore_programmables_var).pack(anchor=tk.W)
        self.restore_constraints_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(restore_options_frame, text="Constraints", variable=self.restore_constraints_var).pack(anchor=tk.W)
        self.restore_indexes_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(restore_options_frame, text="Indexes", variable=self.restore_indexes_var).pack(anchor=tk.W)
        self.restore_continue_on_error_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(restore_options_frame, text="Continue on error (recommended)", variable=self.restore_continue_on_error_var).pack(anchor=tk.W)
        self.restore_mirror_source_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            restore_options_frame,
            text="Mirror source (attempt all batches; no Azure filter — use for on-prem or when target supports all objects)",
            variable=self.restore_mirror_source_var,
        ).pack(anchor=tk.W)
        self.restore_show_preview_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(restore_options_frame, text="Show preview before each batch (FKs, indexes, etc.) — for debugging", 
                       variable=self.restore_show_preview_var).pack(anchor=tk.W)
        
        self.restore_excel_file_var = tk.StringVar()
        self.restore_configs = []
        
        # Start Restore
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(pady=10)
        self.restore_btn = ttk.Button(btn_frame, text="Start Restore", command=self._start_restore, width=20)
        self.restore_btn.pack(side=tk.LEFT, padx=5)
        
        # Log
        log_frame = ttk.LabelFrame(parent, text="Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.restore_log = scrolledtext.ScrolledText(log_frame, height=10, wrap=tk.WORD)
        self.restore_log.pack(fill=tk.BOTH, expand=True)
        
    def _browse_backup_output(self):
        """Browse for backup output folder."""
        folder = tk.filedialog.askdirectory(title="Select Backup Output Folder")
        if folder:
            self.backup_output_var.set(folder)
            
    def _browse_restore_backup(self):
        """Browse for backup folder to restore."""
        folder = tk.filedialog.askdirectory(title="Select Backup Folder")
        if folder:
            self.restore_backup_path_var.set(folder)
            
    def _start_backup(self):
        """Start schema backup in a separate thread."""
        if not sechma_backup:
            messagebox.showerror("Error", "Schema backup module not available!")
            return
            
        # Validate inputs
        if not self.backup_server_var.get():
            messagebox.showerror("Error", "Server is required!")
            return
        if not self.backup_db_var.get():
            messagebox.showerror("Error", "Database is required!")
            return
            
        self.backup_btn.config(state=tk.DISABLED)
        self.backup_log.delete("1.0", tk.END)
        self.backup_log.insert(tk.END, "Starting backup...\n")
        
        def run_backup():
            try:
                backup_root = self.backup_output_var.get() or (self.project_path / "backups" if self.project_path else "backups")
                
                cfg = {
                    "server": self.backup_server_var.get(),
                    "database": self.backup_db_var.get(),
                    "auth": self.backup_auth_var.get(),
                    "user": self.backup_user_var.get(),
                    "password": self.backup_password_var.get() or None,
                    "backup_root": str(backup_root),
                    "log_table_sample": 20,
                    "export_defaults_separately": True
                }
                
                self.backup_log.insert(tk.END, f"Connecting to {cfg['server']}...\n")
                self.backup_log.see(tk.END)
                
                summary = sechma_backup.run_backup(cfg)
                
                if summary["status"] == "success":
                    self.backup_log.insert(tk.END, f"\n✓ Backup completed successfully!\n")
                    self.backup_log.insert(tk.END, f"Run ID: {summary.get('run_id', 'N/A')}\n")
                    
                    # Auto-fill restore backup path with the backup that was just created
                    # Construct backup path from summary data
                    backup_root = self.backup_output_var.get() or (self.project_path / "backups" if self.project_path else "backups")
                    run_id = summary.get('run_id', '')
                    server = summary.get('server', '')
                    database = summary.get('database', '')
                    
                    if run_id and server and database:
                        try:
                            from src.utils.paths import short_slug
                        except ImportError:
                            from utils.paths import short_slug
                        server_slug = short_slug(server)
                        db_slug = short_slug(database)
                        backup_path = str(Path(backup_root) / server_slug / db_slug / "runs" / run_id)
                        
                        self.latest_backup_path = backup_path
                        # Update restore backup path field
                        self.restore_backup_path_var.set(backup_path)
                        self.backup_log.insert(tk.END, f"\n💡 Backup path auto-filled in the Restore tab.\n")
                        self.backup_log.insert(tk.END, f"   Path: {backup_path}\n")
                    
                    self.frame.after(0, lambda: messagebox.showinfo("Success", "Backup completed successfully!\n\nSwitch to the Restore tab to restore to a destination."))
                else:
                    self.backup_log.insert(tk.END, f"\n✗ Backup failed!\n")
                    self.backup_log.insert(tk.END, f"Errors: {summary.get('errors', [])}\n")
                    self.frame.after(0, lambda: messagebox.showerror("Error", "Backup failed! Check log for details."))
            except Exception as e:
                self.backup_log.insert(tk.END, f"\n✗ Error: {str(e)}\n")
                self.frame.after(0, lambda msg=str(e): messagebox.showerror("Error", f"Backup failed: {msg}"))
            finally:
                self.frame.after(0, lambda: self.backup_btn.config(state=tk.NORMAL))
                
        threading.Thread(target=run_backup, daemon=True).start()
        
    def _start_restore(self):
        """Start schema restore in a separate thread."""
        if not restore_schema:
            messagebox.showerror("Error", "Schema restore module not available!")
            return
            
        # Validate inputs
        if not self.restore_backup_path_var.get():
            messagebox.showerror("Error", "Backup path is required!")
            return
        if not self.restore_server_var.get():
            messagebox.showerror("Error", "Destination server is required!")
            return
        if not self.restore_db_var.get():
            messagebox.showerror("Error", "Destination database is required!")
            return
            
        self.restore_btn.config(state=tk.DISABLED)
        self.restore_log.delete("1.0", tk.END)
        self.restore_log.insert(tk.END, "Starting restore...\n")
        
        def run_restore():
            try:
                preview_callback = None
                if self.restore_show_preview_var.get():
                    try:
                        from gui.dialogs.restore_preview_dialog import create_preview_callback
                        preview_callback = create_preview_callback(self.frame.winfo_toplevel())
                    except ImportError:
                        pass
                
                cfg = {
                    "backup_path": self.restore_backup_path_var.get(),
                    "dest_server": self.restore_server_var.get(),
                    "dest_db": self.restore_db_var.get(),
                    "dest_auth": self.restore_auth_var.get(),
                    "dest_user": self.restore_user_var.get(),
                    "dest_password": self.restore_password_var.get() or None,
                    "restore_tables": self.restore_tables_var.get(),
                    "restore_programmables": self.restore_programmables_var.get(),
                    "restore_constraints": self.restore_constraints_var.get(),
                    "restore_indexes": self.restore_indexes_var.get(),
                    "continue_on_error": self.restore_continue_on_error_var.get(),
                    "mirror_source": self.restore_mirror_source_var.get(),
                    "dry_run": False,
                    "preview_callback": preview_callback,
                }
                
                def log(msg):
                    self.restore_log.insert(tk.END, msg)
                    self.restore_log.see(tk.END)
                
                log(f"Restoring to {cfg['dest_server']} / {cfg['dest_db']}...\n")
                
                summary = restore_schema.run_restore(cfg)
                
                # Append per-type summary if available
                stats = summary.get("statistics") or {}
                if stats:
                    log("\n--- Summary ---\n")
                    log(f"  Batches executed: {stats.get('batches_executed', 0)}\n")
                    log(f"  Already existed (skipped): {stats.get('batches_already_existed', 0)}\n")
                    log(f"  Failed: {stats.get('batches_failed', 0)}\n")
                    log(f"  Other skips: {stats.get('batches_skipped', 0) - stats.get('batches_already_existed', 0)}\n")
                    filtered_azure = stats.get("batches_filtered_azure", 0)
                    if filtered_azure > 0:
                        log(f"  Azure-incompatible (not restored): {filtered_azure}\n")
                        note = summary.get("azure_filter_note")
                        if note:
                            log(f"  Note: {note}\n")
                
                if summary["status"] == "success":
                    log("\n✓ Restore completed successfully!\n")
                    self.frame.after(0, lambda: messagebox.showinfo("Success", "Restore completed successfully!"))
                else:
                    log("\n✗ Restore completed with errors.\n")
                    for err in (summary.get("errors") or [])[:20]:
                        log(f"  {err}\n")
                    if len(summary.get("errors") or []) > 20:
                        log(f"  ... and {len(summary['errors']) - 20} more\n")
                    self.frame.after(0, lambda: messagebox.showwarning("Warning", "Restore completed with errors. Check log for details."))
            except Exception as e:
                self.restore_log.insert(tk.END, f"\n✗ Error: {str(e)}\n")
                self.restore_log.see(tk.END)
                self.frame.after(0, lambda msg=str(e): messagebox.showerror("Error", f"Restore failed: {msg}"))
            finally:
                self.frame.after(0, lambda: self.restore_btn.config(state=tk.NORMAL))
                
        threading.Thread(target=run_restore, daemon=True).start()
        
    def _download_template(self, template_type: str):
        """Download sample Excel template."""
        try:
            file_path = create_sample_excel(template_type)
            if file_path:
                messagebox.showinfo("Success", f"Template saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template: {str(e)}")
            
    def _upload_excel_backup(self):
        """Upload Excel file for bulk backup."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not file_path:
            return
            
        try:
            configs = read_excel_file(
                file_path,
                required_columns=["src_server", "src_db"],
                default_user=self.backup_user_var.get() or None
            )
            
            self.backup_configs = configs
            self.backup_excel_file_var.set(f"Loaded {len(configs)} configuration(s) from {Path(file_path).name}")
            self.backup_bulk_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Loaded {len(configs)} configuration(s) from Excel file!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {str(e)}")
            
    def _upload_excel_restore(self):
        """Upload Excel file for bulk restore."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not file_path:
            return
            
        try:
            configs = read_excel_file(
                file_path,
                required_columns=["dest_server", "dest_db", "backup_path"],
                default_user=self.restore_user_var.get() or None
            )
            
            self.restore_configs = configs
            self.restore_excel_file_var.set(f"Loaded {len(configs)} configuration(s) from {Path(file_path).name}")
            self.restore_bulk_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Loaded {len(configs)} configuration(s) from Excel file!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {str(e)}")
            
    def _start_bulk_backup(self):
        """Start bulk backup from Excel configurations (uses Bulk tab log)."""
        if not self.backup_configs:
            messagebox.showwarning("Warning", "No configurations loaded! Upload an Excel file in the Bulk (Excel) tab first.")
            return
        if not sechma_backup:
            messagebox.showerror("Error", "Schema backup module not available!")
            return
        
        self.backup_bulk_btn.config(state=tk.DISABLED)
        log_widget = getattr(self, "bulk_log", self.backup_log)
        log_widget.delete("1.0", tk.END)
        log_widget.insert(tk.END, f"Starting bulk backup for {len(self.backup_configs)} configuration(s)...\n")
        log_widget.see(tk.END)
        
        def run_bulk_backup():
            success_count = 0
            fail_count = 0
            for idx, cfg in enumerate(self.backup_configs, 1):
                def log(msg):
                    log_widget.insert(tk.END, msg)
                    log_widget.see(tk.END)
                log(f"\n[{idx}/{len(self.backup_configs)}] {cfg.get('src_server')}/{cfg.get('src_db')}...\n")
                try:
                    backup_root = self.backup_output_var.get() or (self.project_path / "backups" if self.project_path else "backups")
                    backup_cfg = {
                        "server": cfg.get("src_server"),
                        "database": cfg.get("src_db"),
                        "auth": cfg.get("src_auth", self.backup_auth_var.get()),
                        "user": cfg.get("src_user", cfg.get("user", self.backup_user_var.get())),
                        "password": cfg.get("src_password", self.backup_password_var.get() or None),
                        "backup_root": str(backup_root),
                        "log_table_sample": 20,
                        "export_defaults_separately": True
                    }
                    summary = sechma_backup.run_backup(backup_cfg)
                    if summary["status"] == "success":
                        log(f"✓ {cfg.get('src_db')}\n")
                        success_count += 1
                    else:
                        log(f"✗ {cfg.get('src_db')}: {summary.get('errors', [])}\n")
                        fail_count += 1
                except Exception as e:
                    log(f"✗ {cfg.get('src_db')}: {str(e)}\n")
                    fail_count += 1
            log(f"\n{'='*60}\nBulk backup: {success_count} succeeded, {fail_count} failed\n")
            self.frame.after(0, lambda: messagebox.showinfo("Bulk Backup Complete", f"Completed: {success_count} succeeded, {fail_count} failed"))
            self.frame.after(0, lambda: self.backup_bulk_btn.config(state=tk.NORMAL))
        threading.Thread(target=run_bulk_backup, daemon=True).start()
        
    def _start_bulk_restore(self):
        """Start bulk restore from Excel configurations (uses Bulk tab log; no preview)."""
        if not self.restore_configs:
            messagebox.showwarning("Warning", "No configurations loaded! Upload an Excel file in the Bulk (Excel) tab first.")
            return
        if not restore_schema:
            messagebox.showerror("Error", "Schema restore module not available!")
            return
        
        self.restore_bulk_btn.config(state=tk.DISABLED)
        log_widget = getattr(self, "bulk_log", self.restore_log)
        log_widget.delete("1.0", tk.END)
        log_widget.insert(tk.END, f"Starting bulk restore for {len(self.restore_configs)} configuration(s)...\n")
        log_widget.see(tk.END)
        
        def run_bulk_restore():
            success_count = 0
            fail_count = 0
            for idx, cfg in enumerate(self.restore_configs, 1):
                def log(msg):
                    log_widget.insert(tk.END, msg)
                    log_widget.see(tk.END)
                log(f"\n[{idx}/{len(self.restore_configs)}] {cfg.get('dest_server')}/{cfg.get('dest_db')}...\n")
                try:
                    restore_cfg = {
                        "backup_path": cfg.get("backup_path"),
                        "dest_server": cfg.get("dest_server"),
                        "dest_db": cfg.get("dest_db"),
                        "dest_auth": cfg.get("dest_auth", self.restore_auth_var.get()),
                        "dest_user": cfg.get("dest_user", cfg.get("user", self.restore_user_var.get())),
                        "dest_password": cfg.get("dest_password", self.restore_password_var.get() or None),
                        "restore_tables": cfg.get("restore_tables", self.restore_tables_var.get()),
                        "restore_programmables": cfg.get("restore_programmables", self.restore_programmables_var.get()),
                        "restore_constraints": cfg.get("restore_constraints", self.restore_constraints_var.get()),
                        "restore_indexes": cfg.get("restore_indexes", self.restore_indexes_var.get()),
                        "continue_on_error": cfg.get("continue_on_error", self.restore_continue_on_error_var.get()),
                        "mirror_source": cfg.get("mirror_source", self.restore_mirror_source_var.get()),
                        "dry_run": False,
                        "preview_callback": None,
                    }
                    summary = restore_schema.run_restore(restore_cfg)
                    if summary["status"] == "success":
                        log(f"✓ {cfg.get('dest_db')}\n")
                        success_count += 1
                    else:
                        errs = summary.get("errors", [])
                        log(f"✗ {cfg.get('dest_db')}: {len(errs)} error(s)\n")
                        fail_count += 1
                except Exception as e:
                    log(f"✗ {cfg.get('dest_db')}: {str(e)}\n")
                    fail_count += 1
            log(f"\n{'='*60}\nBulk restore: {success_count} succeeded, {fail_count} failed\n")
            self.frame.after(0, lambda: messagebox.showinfo("Bulk Restore Complete", f"Completed: {success_count} succeeded, {fail_count} failed"))
            self.frame.after(0, lambda: self.restore_bulk_btn.config(state=tk.NORMAL))
        threading.Thread(target=run_bulk_restore, daemon=True).start()

